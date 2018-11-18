from __future__ import division
import re,os,random,codecs



def num2col(i):
	import string
	alpha=string.ascii_uppercase
	if i<len(alpha):
		code=alpha[i]
	else:
		import math
		offset=int(math.floor(i / len(alpha)))
		code1=alpha[offset - 1]
		offset2=i - (len(alpha) * offset)
		code2=alpha[offset2]
		code=code1+code2
	return code

def yank(text,tag,none=None):
	if type(tag)==type(''):
		tag=tagname2tagtup(tag)

	try:
		return text.split(tag[0])[1].split(tag[1])[0]
	except IndexError:
		return none


import re
import htmlentitydefs

def safeunicode(x):
	y=u''
	for xx in x:
		if xx=="'":
			y+="-"
			continue
		try:
			y+=xx.encode('utf-8')
		except:
			y+='-'
	return y

def list2freqs(list,tfy=False):
	x=[_x.strip() for _x in list.split('\n')]
	return toks2freq(x,tfy=tfy)

def convertentity(m):
	if m.group(1)=='#':
		try:
			return chr(int(m.group(2)))
		except ValueError:
			return '&#%s;' % m.group(2)
	try:
		return htmlentitydefs.entitydefs[m.group(2)]
	except KeyError:
		return '&%s;' % m.group(2)

def converthtml(s):
	return re.sub(r'&(#?)(.+?);',convertentity,s)

def escape(s):
	return converthtml(s).replace('\ufffd','')


def yanks(text,tag):
	if type(tag)==type(''):
		tag=tagname2tagtup(tag)

	return [ x.split(tag[1])[0] for x in text.split(tag[0])[1:] ]

def yanks2(text,tag):
	if type(tag)==type(''):
		tag=tagname2tagtup(tag)

	ys=[]
	#return [ tag[0][-1].join(x.split(tag[0][:-1])[1].split(tag[0][-1])[1:]) for x in text.split(tag[1])[:-1] ]

	for x in text.split(tag[1])[:-1]:
		try:
			x=x.split(tag[0][:-1])[1].split(tag[0][-1])[1:]
			x=tag[0][-1].join(x)
		except IndexError:
			pass
		ys.append(x)
	return ys

class String(unicode):
	def __str__(self):
		return self.str
	def __repr__(self):
		return self.str
	def lines(self,lb='\n'):
		return self.str.split(lb)
	def __init__(self,str,**meta):
		for k,v in meta.items(): setattr(self,k,v)
		self.str=str

class StringObject(object):
	def __str__(self):
		return self.str.encode('utf-8')
	def __unicode__(self):
		return self.str
	def __repr__(self):
		return self.str
	def lines(self,lb='\n'):
		return self.str.split(lb)
	def __init__(self,str,**meta):
		for k,v in meta.items(): setattr(self,k,v)
		self.str=str
	def split(self,x):
		return self.str.split(x)

def pull1(text,tag,returnType='u',html=True):
	x=pull(text,tag,returnType)
	if not len(x): return None
	xx=x[0]
	if not html: xx=unhtml(xx)
	return xx


# def tags(html):
# 	def next(tag):
# 		if
#
# 	import bs4
# 	bs=bs4.BeautifulSoup(html)


def pull(text,tag,returnType='u'):
	#print tag,returnType
	#if not returnType: returnType='u'
	if type(tag)==type(''): tag=tagname2tagtup(tag)
	ys=[]

	## split by closing tag, excluding the last [b/c outside of tag]
	for x in text.split(tag[1])[:-1]:
		## now try to split by opening tag omitting closing ">"
		x=x.split(tag[0][:-1]+' ')

		if len(x)>1:	# then there is such a thing as "<tag "
			x=x[1]
			## now split by closing ">"
			y=x.split(tag[0][-1])
			## get attrs
			attrs=y[0]
			adict={}
			if attrs:
				key=None
				for ax in attrs.split():
					if '=' in ax:
						axs=ax.split('=')
						k=axs[0]
						v='='.join(axs[1:])
						#if not "'" in v and not '"' in v:
						adict[k]=v.replace('"','').replace("'",'')
						key=k
					elif key:
						adict[key]+=' '+ax.replace('"','').replace("'",'')

			for k in adict:
				adict[k]=adict[k].strip()
				if adict[k].isdigit(): adict[k]=int(adict[k])

			## rejoin text by closing ">", without attrs
			text=tag[0][-1].join(y[1:])
		else:
			text=x[0].split(tag[0])[1]
			adict={}

		if returnType=='u':
			string=text
		else:
			string=StringObject(text)
			for k,v in adict.items(): setattr(string,k,v)
		ys.append(string)
	return ys



def tagname2tagtup(tagname):
	return ('<'+tagname+'>','</'+tagname+'>')

def safestr(string):
	try:
		return str(string)
	except UnicodeEncodeError:
		return str(string.encode('utf-8','replace'))
	except:
		return "<????>"

def is_safe(string):
	try:
		return str(string)==ascii(string)
	except:
		return False


def simple(str):
	o=[]
	for x in str:
		try:
			unicode(x)
			o+=[x]
		except:
			pass
	return ''.join(o)

def ascii(inputstring):
	o=[]
	for x in inputstring:
		try:
			str(x)
			o+=[x]
		except:
			pass
	return ''.join(o)

def rascii(inputstring,woops='.'):
	o=[]
	for x in inputstring:
		try:
			str(x)
			o+=[x]
		except:
			o+=[woops]
	return ''.join(o)

def dict2xml(d,root="xml"):
	o=[]
	for k,v in sorted(d.items(),reverse=False):
		o+=["<"+k+">"+v+"</"+k+">"]
	return "<"+root+">\n\t"+ "\n\t".join(o) + "\n</"+root+">"


def neginback(strnum):
	if strnum.startswith("-"):
		return strnum[1:]+"-"
	else:
		return strnum

def thetime():
	from time import localtime, strftime
	return strftime("%Y%m%d.%H%M", localtime())

# these two lists serves as building blocks to construt any number
# just like coin denominations.
# 1000->"M", 900->"CM", 500->"D"...keep on going
decimalDens=[1000,900,500,400,100,90,50,40,10,9,5,4,1]
romanDens=["M","CM","D","CD","C","XC","L","XL","X","IX","V","IV","I"]


def roman(dec):
	"""
	Perform sanity check on decimal and throws exceptions when necessary
	"""
        if dec <=0:
	  raise ValueError, "It must be a positive"
         # to avoid MMMM
	elif dec>=4000:
	  raise ValueError, "It must be lower than MMMM(4000)"

	return decToRoman(dec,"",decimalDens,romanDens)

def decToRoman(num,s,decs,romans):
	"""
	  convert a Decimal number to Roman numeral recursively
	  num: the decimal number
	  s: the roman numerial string
	  decs: current list of decimal denomination
	  romans: current list of roman denomination
	"""
	if decs:
	  if (num < decs[0]):
	    # deal with the rest denomination
	    return decToRoman(num,s,decs[1:],romans[1:])
	  else:
	    # deduce this denomation till num<desc[0]
	    return decToRoman(num-decs[0],s+romans[0],decs,romans)
	else:
	  # we run out of denomination, we are done
	  return s


def flatten_ld(metald,flatten_prefix='window'):
	ld2=[]
	for d in metald:
		flatten=[k for k in d.keys() if k.startswith(flatten_prefix)]
		include_with_flatten=[]
		not_to_flatten=list(set(d.keys())-set(flatten))
		for k in [kx for kx in flatten]:
			for k2 in [kx for kx in flatten if kx.startswith(k) and kx!=k]:
				flatten.remove(k2)
				include_with_flatten.append(k2)
		for k in flatten:
			d2={}
			for dk in not_to_flatten: d2[dk]=d[dk]
			d2[flatten_prefix+'Type']=noPunc(k.replace(flatten_prefix,''))
			d2[flatten_prefix+'Value']=d[k]
			for dk in include_with_flatten:
				if not dk.startswith(k): continue
				dkx=noPunc(dk.replace(k,''))
				dkx=flatten_prefix+dkx[0].upper()+dkx[1:]
				d2[dkx]=d[dk]
			ld2.append(d2)
	return ld2



def ynk(text,start,end,inout=""):
	if (not start in text) or (not end in text):
		return ""


	try:
		if (inout=="in" or inout==0):
			return text.split(start)[1].split(end)[0]
		elif (inout=="out" or inout==1):
			return text.split(end)[0].split(start)[-1]
		else:
			o=[]
			for x in text.split(start):
				#if x.count(">")>1:
				#	x=x[x.index(">")+1:]

				xx=x.split(end)[0].strip()
				if not xx: continue
				if xx.startswith("<!DOCTYPE"): continue		# NYT hack
				if xx.startswith("<NYT_"): continue
				if xx.startswith("<script"): continue

				o.append(xx.replace("\n"," ").replace("\r"," "))
			return "\n\n".join(o)
	except:
		return ""


def nupos2desc(pos):
	translation={'a':'adverb/conj/prep as adverb','av':'adverb','c':'adverb/conj/prep as conj','cc':'coordinating conjunction','cr':'numeral','cs':'subordinating conjunction / that as conj','d':'determiner [that, much]','d.':'d.?','dc':'comparative determiner [less]','dg':"determiner as possessive [the latter's]",'ds':'superlative determiner','dt':'article','dx':'negative determiner as adverb','fw':'foreign word','j':'adjective','j.':'j.?','jc':'comparative adjective','jp':'proper adjective','js':'superlative adjective','n':'adjective/participle as noun','n1':'singular noun','n2':'plural noun','ng':'possessive noun','nj':'proper adjective [Roman]','np':'proper noun','or':'ordinal number','p':'adj/conj/prep as prep','pc':'adj/conj/prep as particle','pi':'indefinite pronoun [one]','pn':'personal pronoun','po':'possessive pronoun','pp':'preposition','px':'reflexive pronoun','q':'wh-word, interrogative use','r':'wh-word, relative use','sy':'symbol','uh':'interjection','vb':'to be, any tense','vd':'to do, any tense','vh':'to have, any tense','vm':'modal verb, any tense','vv':'verb','xx':'negative','zz':'unknown token'}

	return translation.get(pos,pos)

def nupos():
	ld=tsv2ld('/Users/ryan/inbox/python/nupos.txt')
	nd={}
	for dx in ld:
		nd[dx['tag']]=dx
	return nd

def bunch_ld(ld,key):
	last_val=None
	newld=[]
	newd={}
	for d in ld:
		if not last_val or d.get(key,None)!=last_val:
			if newd: newld+=[newd]
			newd=dict(d.items())
		else:
			for k in d:
				if not k in newd or newd[k]!=d[k]:
					if type(newd[k])==list:
						newd[k].append(d[k])
					else:
						newd[k]=[newd[k], d[k]]
		last_val=d.get(key,None)
	if newd: newld+=[newd]
	return newld



def xls2ld(fn,header=[],sheetname=True,keymap={},keymap_all=unicode):
	import time
	now=time.time()
	print '>> reading as xls:',fn
	import xlrd
	if '*' in keymap: keymap_all=keymap['*']
	headerset=True if len(header) else False
	f=xlrd.open_workbook(fn)
	ld=[]
	def _boot_xls_sheet(sheet,header=[]):
		ld2=[]
		for y in range(sheet.nrows):
			if not header:
				for xi in range(sheet.ncols):
					cell=sheet.cell_value(rowx=y,colx=xi)
					header+=[cell]
				continue
			d={}
			for key in header:
				try:
					value=sheet.cell_value(rowx=y, colx=header.index(key))
					#print '??',value,type(value),key
					if keymap_all:
						func=keymap_all
						if func in [str,unicode] and type(value) in [float]:
							if value == int(value): value=int(value)
						d[key]=keymap_all(value)
					elif keymap and key in keymap:
						func=keymap[key]
						if func in [str,unicode] and type(value) in [float]:
							if value == int(value): value=int(value)
						d[key]=keymap[key](value)
					else:
						d[key]=value
					#print key,value,y,header.index(key),row[header.index(key)]
				except Exception as e:
					print '!! ERROR:',e
					print '!! on key =',key,'& value =',value, type(value)
					#print "!! "+key+" not found in "+str(sheet)
					#d[key]=''
					pass
			if len(d):
				if sheetname: d['sheetname']=sheet.name
				ld2.append(d)
		return ld2


	if f.nsheets > 1:
		sheetnames=sorted(f.sheet_names())
		for sheetname in sheetnames:
			sheet=f.sheet_by_name(sheetname)
			for d in _boot_xls_sheet(sheet,header=header if headerset else []):
				ld.append(d)
	else:
		sheet = f.sheet_by_index(0)
		ld.extend(_boot_xls_sheet(sheet,header=header if headerset else []))

	nownow=time.time()
	print '>> done ['+str(round(nownow-now,1))+' seconds]'

	return ld


def xls2dld(fn,header=[]):
	return ld2dld(xls2ld(fn,header=header,sheetname=True), 'sheetname')

def levenshtein(s1, s2):
	l1 = len(s1)
	l2 = len(s2)

	matrix = [range(l1 + 1)] * (l2 + 1)
	for zz in range(l2 + 1):
		matrix[zz] = range(zz,zz + l1 + 1)
	for zz in range(0,l2):
		for sz in range(0,l1):
			if s1[sz] == s2[zz]:
				matrix[zz+1][sz+1] = min(matrix[zz+1][sz] + 1, matrix[zz][sz+1] + 1, matrix[zz][sz])
			else:
				matrix[zz+1][sz+1] = min(matrix[zz+1][sz] + 1, matrix[zz][sz+1] + 1, matrix[zz][sz] + 1)
	return matrix[l2][l1]

def xlsx2ld(fn,header=[],numsheets=1):
	from openpyxl.reader.excel import load_workbook
	header_set=bool(len(header))
	wb=load_workbook(filename=fn)
	ld=[]
	for sheet in wb.worksheets[:numsheets]:
		if not header_set: header=[]
		#header=[]
		for rownum,row in enumerate(sheet.rows):
			values=[]
			for cell in row:
				value=cell.value
				if value is None:
					value=''

				try:
					value=float(value)/0
				except:
					value=value
					if not isinstance(value, unicode):
						value=unicode(value)
				values.append(value)
			if not rownum and not len(header):
				header=values
			else:
				d=dict((header[i],values[i]) for i in range(len(values)))
				ld+=[d]
	return ld

def dl2ld(dl,kcol='group'):
	ld=[]
	for k in dl:
		for d in dl[k]:
			d[kcol]=k
			ld+=[d]
	return ld

def ld2dl(ld):
	keys = ld[0].keys()
	dl={}
	for k in keys:
		dl[k] = [d[k] for d in ld]
	return dl

def fn2ld(fn,header=[],sep='\t',nsep='\n'):
	import codecs
	f=codecs.open(fn,encoding='utf-8')
	for line in f:
		line=line.strip()
		if not header:
			header=line.split(sep)
			continue
		dx={}
		for i,val in enumerate(line.split(sep)):
			key=header[i] if len(header)>i else 'key_'+str(i)
			dx[key]=val
		yield dx

def goog2tsv(googsrc):
	import bs4
	dom=bs4.BeautifulSoup(googsrc,'html.parser')
	header=[th.text for th in dom('thead')[0]('th')]
	header=header if True in [bool(hx) for hx in header] else None
	old=[]
	for row in dom('tbody')[0]('tr'):
		rowdat=[cell.text for cell in row('td')]
		if not header:
			header=rowdat
			#print ">> HEADER:",header
			continue
		odx=dict(zip(header,rowdat))
		old+=[odx]
	return old


def tsv2ld(fn,tsep='\t',nsep='\n',u=True,header=[],keymap={},zero='',removeEmpties=False):
	import time
	now=time.time()
	if tsep=='\t':
		print '>> reading as tsv:',fn
	elif tsep==',':
		print '>> reading as csv:',fn

	import os
	if fn.startswith('http'):
		print '>> reading webpage...'
		import urllib
		f=urllib.urlopen(fn)
		t=f.read()
		if fn.endswith('/pubhtml'):
			return goog2tsv(t)
		f.close()
	elif not os.path.exists(fn):
		t=fn
	elif u:
		import codecs
		f=codecs.open(fn,encoding='utf-8')
		t=f.read()
		f.close()
	else:
		f=open(fn,'r')
		t=f.read()
		f.close()
	t=t.replace('\r\n','\n')
	t=t.replace('\r','\n')

	#header=[]
	listdict=[]


	for line in t.split(nsep):
		if not line.strip(): continue
		line=line.replace('\n','')
		ln=line.split(tsep)
		#print ln
		if not header:
			header=ln
			for i,v in enumerate(header):
				if v.startswith('"') and v.endswith('"'):
					header[i]=v[1:-1]
			continue
		edict={}
		for i in range(len(ln)):
			try:
				k=header[i]
			except IndexError:
				#print "!! unknown column for i={0} and val={1}".format(i,ln[i])
				continue
			v=ln[i].strip()

			if '*' in keymap:
				v=keymap['*'](v)
			elif k in keymap:
				#print v, type(v)
				v=keymap[k](v)
				#print v, type(v)
			else:
				if v.startswith('"') and v.endswith('"'):
					v=v[1:-1]
				try:
					v=float(v)
				except ValueError:
					v=v

			if type(v) in [str,unicode] and not v:
				if zero=='' and removeEmpties:
					continue
				else:
					v=zero
			edict[k]=v
		if edict:
			listdict.append(edict)

	nownow=time.time()
	print '>> done ['+str(round(nownow-now,1))+' seconds]'

	return listdict

def dkey(d,extra={}):
	#import pytxt
	#kv='__'.join(['{0}_{1}'.format(unicode(k),unicode(v)) for k,v in sorted(d.items()+extra.items())])
	import cPickle
	kv=cPickle.dumps(dict(d.items()+extra.items()))
	return kv

def unhtml(data):
	if not data: return data
	try:
		from lxml.html import fromstring
		return fromstring(data).text_content()
	except:
		return remove_html_tags(data)

def remove_html_tags(data):
	#data=safestr(data)
	p=re.compile(r'<.*?>',re.UNICODE)
	try:
		y=str(p.sub('',data)).strip().split('">')
	except UnicodeEncodeError:
		y=unicode(p.sub('',data)).strip().split('">')
	while(('&' in y) and (';' in y)):
		y=y[:y.index('&')]+y[y.index(';')+1:]
	try:
		return y[1].strip()
	except:
		return y[0]

def htm2txt(element):
    import types
    text = ''
    for elem in element.recursiveChildGenerator():
        if isinstance(elem, types.StringTypes):
            txt=elem.strip().replace('\r\n',' ').replace('\r',' ').replace('\n',' ')
            while '  ' in txt: txt=txt.replace('  ',' ')
            text += txt
        elif elem.name in ['br','p']:
            text += '\n'
    return text

def extractTags(text,leavetexttags=[u"placeName"]):
	tags=[]
	tags_milestone=[]
	yankeds=[]

	if "</" in text:
		for x in text.split("</")[1:]:
			tags.append(x.split(">")[0])

	if "/>" in text:
		for x in text.split("/>")[:-1]:
			x=x.split("<")[-1]
			try:
				x=x.split()[0]
			except IndexError:
				x=x
			#if "/" in x: continue
			#if not x: continue
			tags_milestone.append(x)

	for tag in tags_milestone:
		yanked=yank(text,("<"+tag,"/>"))
		while yanked.strip():
			ydat="<"+tag+yanked+"/>"
			#yankeds.append(ydat)
			text=text.replace(ydat,' ')
			yanked=yank(text,("<"+tag,"/>"))

	for tag in tags:
		yanked=yank(text,("<"+tag,"</"+tag+">"))
		while yanked and yanked.strip():
			ydat="<"+tag+yanked+"</"+tag+">"
			if tag in leavetexttags:
				text=text.replace(ydat,remove_html_tags(yanked.split(">")[-1]))
			else:
				yankeds.append(ydat)
				text=text.replace(ydat,' ')
			yanked=yank(text,("<"+tag,"</"+tag+">"))
	return (text.replace("\n","").replace("\r",""),yankeds)

def gleanPunc2(aToken):
	aPunct0 = ''
	aPunct1 = ''
	while(len(aToken) > 0 and not aToken[0].isalnum()):
		aPunct0 = aPunct0+aToken[:1]
		aToken = aToken[1:]
	while(len(aToken) > 0 and not aToken[-1].isalnum()):
		aPunct1 = aToken[-1]+aPunct1
		aToken = aToken[:-1]

	return (aPunct0, aToken, aPunct1)



def gleanPunc(aToken):
	aPunct = None
	while(len(aToken) > 0 and not aToken[0].isalnum()):
		aPunct = aToken[:1]
		aToken = aToken[1:]
	while(len(aToken) > 0 and not aToken[-1].isalnum()):
		aPunct = aToken[-1]
		aToken = aToken[:-1]
	return (aToken, aPunct)

def isPunc(token):
	return (len(noPunc(token))==0)

def alphas(token):
	return ''.join([x for x in token if x.isalpha() or x in [' ']])

"""
def noPunc(token):
	if not token: return token
	x=gleanPunc(token)[0]
	x=x.split('&')[0]
	y=x.split(';')
	try:
		x=y[1]
	except IndexError:
		pass
	x=x.split('\\')[0]
	return x
"""

def noPunc(token):
	from string import punctuation
	return token.strip(punctuation)


def bigrams(l):
	return ngram(l,2)

def ngram(l,n=3):
	grams=[]
	gram=[]
	for x in l:
		gram.append(x)
		if len(gram)<n: continue
		g=tuple(gram)
		grams.append(g)
		gram.reverse()
		gram.pop()
		gram.reverse()
	return grams


def fn2dict(fn,sep='\t'):
	return readDict(fn,sep=sep)

def readDict(fn,sep='\t'):
	try:
		d={}
		#f=open(fn)
		import codecs
		f=codecs.open(fn,encoding='utf-8')
		for line in f:
			ln=line.split(sep)
			k=ln[0].strip()
			v=ln[1].strip()

			if v.isdigit():
				d[k]=int(v)
			else:
				d[k]=v

		if len(d):
			return d
		else:
			return None

	except IOError:
		return {}

def writeDict(fn,d,sep="\t",toprint=True):
	o=""
	for k,v in d.items():
		o+=sep.join(str(x) for x in [k,v])+"\n"
	write(fn,o,toprint)


def dict2tuple(d):
	return tuple((k,v if type(v)!=list else tuple(v)) for k,v in sorted(d.items()))


def extractTagsAsDict(text,leavetexttags=[u"placeName"]):
	text,tags=extractTags(text,leavetexttags)
	tagdict={}
	for tag in tags:

		opentag=tag.split(">")[0].split("<")[1].strip()
		tagbody=unhtml(tag).strip()

		if not tagbody: continue

		if " " in opentag:
			spaces=opentag.split()
			tagname=spaces[0]
			for space in spaces[1:2]:
				if not space.strip(): continue
				dat=space.strip().split("=")
				k=dat[0]
				try:
					v=dat[1]
				except:
					print "error with extractTagsAsDict in pytxt"
					continue
				v=v.replace('"','').replace("'","").strip()

				try:
					tagdict[tagname][k][v]=tagbody
				except KeyError:
					try:
						tagdict[tagname][k]={}
						tagdict[tagname][k][v]=tagbody
					except KeyError:
						tagdict[tagname]={}
						tagdict[tagname][k]={}
						tagdict[tagname][k][v]=tagbody

		else:
			tagname=opentag
			tagdict[tagname]=tagbody

	return tagdict


def writeToFile(folder,fn,data,extension="tsv"):
	#ofolder=os.path.join(folder,'results','stats','corpora',name)

	if not os.path.exists(folder):
		os.makedirs(folder)

	ofn=os.path.join(folder,'.'.join([fn,extension]))
	print ">> saved:",ofn
	of = open(ofn,'w')
	of.write(data)
	of.close()


def chunk(fn,num=10):
	import codecs,os
	t=codecs.open(fn,encoding='utf-8').read()
	words=t.split()
	numwords=len(words)
	wordperseg=int(numwords/num)
	for segnum in range(num):
		segwords=words[segnum*wordperseg : (segnum+1)*wordperseg]
		segtext=' '.join(segwords)
		segfn=os.path.basename(fn).replace('fulltext',str(segnum+1))
		segfn=segfn.split('_')[0] + '_' + segfn.split('_')[1].zfill(2) + '_' + '_'.join(segfn.split('_')[2:])
		write(segfn,segtext)


def split_texts(infolder,outfolder,lim=1000):
	for fn in os.listdir(infolder):
		import codecs
		text=codecs.open(os.path.join(infolder,fn),encoding='utf-8').read().split()
		for n,txt in enumerate(segment(text,lim)):
			ofn=fn.replace('.txt','.'+str(n).zfill(4)+'.txt')
			write(os.path.join(outfolder,ofn), ' '.join(txt))



def segment(l,num=200):
	import math
	segments_needed=int(math.ceil(len(l)/num))
	for n in range(segments_needed+1):
		yield l[n*num:(n+1)*num]

def dld2dll(dld):
	for k in dld:
		dld[k]=ld2ll(dld[k])
	return dld

def write_xls(fn,data,sheetname='index',toprint=True,limFields=None,widths=[],zero=''):
	import xlwt
	wb=xlwt.Workbook(encoding='utf-8')

	if datatype(data).startswith('ld'):
		dd={}
		dd[sheetname]=ld2ll(data,zero=zero)
	elif type(data)!=type({}):
		dd={}
		dd[sheetname]=data
	elif datatype(data).startswith('dld'):
		dd=dld2dll(data)
	else:
		dd=data

	for sheetname,data in sorted(dd.items()):
		ws=wb.add_sheet(sheetname)
		nr=-1
		#style = xlwt.easyxf('align: wrap True')
		#style=xlwt.easyxf('')
		for row in data:
			nc=-1
			nr+=1
			for cell in row:
				nc+=1
				if not (type(cell)==type(1) or type(cell)==type(1.0)):
					try:
						ws.row(nr).set_cell_text(nc,cell)
					except TypeError:
						ws.row(nr).set_cell_text(nc,unicode(cell))
				else:
					ws.row(nr).set_cell_number(nc,cell)
	wb.save(fn)
	if toprint:
		print ">> saved:",fn


def tmp(data):
	import tempfile
	f=tempfile.NamedTemporaryFile()
	f.write(data)
	#f.close()
	return f

def write_tmp(data,suffix=''):
	import time
	fn='/Lab/Processing/tmp/'+str(time.time()).replace('.','')+suffix
	write(fn,data)
	return fn

def ld2html(ld):
	keys=ld2keys(ld)
	headerrow=['<th>%s</th>'%k for k in keys]
	rows=[]
	rows+=['\n\t\t'.join(headerrow)]
	for d in ld:
		row=['<td>%s</td>'%d.get(k,'') for k in keys]
		rows+=['\n\t\t'.join(row)]
	ostr=u"<table>\n\t<tr>\n\t\t" + u'\n\t</tr>\n\t<tr>\n\t\t'.join(rows) + u"\n\t</tr>\n</table>"
	return ostr

def ld2keys(ld):
	keys=[]
	for d in ld:
		for k in d:
			keys+=[k]
	keys=list(sorted(list(set(keys))))
	return keys

def ld2ll(ld,zero='',tostr=False,uni=True):
	keys=[]
	for d in ld:
		for k in d:
			keys+=[k]
	keys=sorted(list(set(keys)))
	o=[keys]
	for d in ld:
		l=[]
		for k in keys:
			v=d.get(k,zero)
			if tostr:
				v=unicode(v) if uni else str(v)
			l+=[v]
		o+=[l]
	return o


def write_ld(fn,ld,zero='',timestamp=None):
	return write(fn,ld2ll(ld,zero=zero),timestamp=timestamp)

def dd2ld(dd,rownamecol='rownamecol'):
	if not rownamecol:
		return [ (dict(v.items())) for k,v in dd.items() ]
	else:
		return [ (dict(v.items() + [(rownamecol,k)])) for k,v in dd.items() ]

def dld2ld(dld,key='rownamecol'):
	ld=[]
	for k in dld:
		for d in dld[k]:
			d[key]=k
			ld+=[d]
	return ld

def ld_resample(ld,key='rownamecol',n=None):
	import random
	dld=ld2dld(ld,key)
	minlen_actually=min([len(dld[k]) for k in dld])
	minlen=minlen_actually if not n or n>minlen_actually else n
	ld2=[]
	print '>> resampling to minimum length of:',minlen
	for k in sorted(dld):
		print '>>',k,len(dld[k]),'-->',minlen
		ld2+=random.sample(dld[k],minlen)
	return ld2

def ld2dld(ld,key='rownamecol'):
	dld={}
	for d in ld:
		if not d[key] in dld: dld[d[key]]=[]
		dld[d[key]]+=[d]
	return dld

def ld2dd(ld,rownamecol='rownamecol'):
	dd={}
	for d in ld:
		dd[d[rownamecol]]=d
		#del dd[d[rownamecol]][rownamecol]
	return dd

def datatype(data,depth=0,v=False):
	def echo(dt):
		if not v: return
		for n in range(depth): print "\t",
		print '['+dt[0]+']'+dt[1:],
		try:
			print "[{0} records]".format(len(data),dt)
		except:
			print

	if type(data) in [str,unicode]:
		echo('string')
		return 's'
	elif type(data) in [float,int]:
		echo('number')
		return 'n'
	elif type(data) in [list]:
		echo('list')
		if not len(data):
			return 'l'
		else:
			return 'l'+datatype(data[0],depth=depth+1,v=v)
	elif type(data) in [dict]:
		echo('dictionary')
		if not len(data):
			return 'd'
		else:
			return 'd'+datatype(data.values()[0],depth=depth+1,v=v)
	else:
		#print "WHAT TYPE OF DATA IS THIS:"
		#print data
		#print type(data)
		#print
		return '?'


def limcols(ld,limcol=255):
	keyd={}
	keys=set()
	for d in ld:
		dkeys=set(d.keys())
		for key in dkeys-keys:
			keyd[key]=0
		keys|=dkeys
		for k in d:
			if d[k]:
				keyd[k]+=1

	cols=set(sorted(keyd.keys(), key=lambda _k: (-keyd[_k],_k))[:limcol])

	for d in ld:
		dkeys=set(d.keys())
		for key in dkeys-cols:
			del d[key]

	return ld

def ld2str(ld,**data):
	if data['limcol']:
		print ">> limiting columns"
		limcol=data['limcol']
		ld=limcols(ld,limcol)
	if 'limcol' in data:
		del data['limcol']
	return ll2str(ld2ll(ld),**data)

def d2ll(d):
	try:
		return [[k,v] for k,v in sorted(d.items(),key=lambda lt: -lt[1])]
	except:
		return [[k,v] for k,v in d.items()]

def d2str(d,uni=True):
	return ll2str(d2ll(d),uni=uni)

def strmake(x,uni=True):
	if uni and type(x) in [unicode]:
		return x
	elif uni and type(x) in [str]:
		return x.decode('utf-8',errors='replace')
	elif uni:
		return unicode(x)
	elif not uni and type(x) in [str]:
		return x
	elif not uni and type(x) in [unicode]:
		return x.encode('utf-8',errors='replace')

	print [x],type(x)
	return str(x)


def ll2str(ll,uni=True,join_line=u'\n',join_cell=u'\t'):
	if not uni:
		join_line=str(join_line)
		join_cell=str(join_cell)
		quotechar='"' if join_cell==',' else ''
	else:
		quotechar=u'"' if join_cell==',' else u''

	for line in ll:
		lreturn=join_cell.join([quotechar+strmake(cell,uni=uni)+quotechar for cell in line])+join_line
		yield lreturn

def l2str(l,uni=True,join_line=u'\n',join_cell=u'\t',quotechar=''):
	for line in l: yield strmake(line)+join_line

def write_ld2(fn,gen1,gen2,uni=True,badkeys=[]):
	def find_keys(gen):
		keys=set()
		for d in gen:
			keys=keys|set(d.keys())
		keys=keys-set(badkeys)
		return keys

	keys=list(sorted(list(find_keys(gen1))))
	numk=len(keys)

	import codecs
	of=codecs.open(fn,'w',encoding='utf-8')
	of.write('\t'.join([strmake(x) for x in keys]) + '\n')

	for d in gen2:
		data=[d.get(key,'') for key in keys]
		of.write('\t'.join([strmake(x) for x in data]) + '\n')
	of.close()
	print ">> saved:",fn


def write2(fn,data,uni=True,join_cell=u'\t',join_line=u'\n',limcol=None):
	## pass off to other write functions if necessary
	if fn.endswith('.xls'): return write_xls(fn,data)
	if fn.endswith('.csv'): join_cell=','

	## get datatyoe
	dt=datatype(data)

	## get str output for datatype
	if dt.startswith('ld'):
		o=ld2str(data,join_cell=join_cell,limcol=limcol)
	elif dt.startswith('dl'):
		o=dl2str(data,uni=uni)
	elif dt.startswith('ll'):
		o=ll2str(data,uni=uni)
	elif dt.startswith('dd'):
		o=dd2str(data,uni=uni)
	elif dt.startswith('l'):
		o=l2str(data,uni=uni)
	elif dt.startswith('d'):
		o=d2str(data,uni=uni)
	else:
		o=data

	## write
	import codecs
	of = codecs.open(fn,'w',encoding='utf-8') if True else open(fn,'w')
	for line in o: of.write(line)
	of.close()
	print '>> saved:',fn


def now(now=None,seconds=True):
	import datetime as dt
	if not now:
		now=dt.datetime.now()
	elif type(now) in [int,float,str]:
		now=dt.datetime.fromtimestamp(now)

	return '{0}{1}{2}-{3}{4}{5}'.format(now.year,str(now.month).zfill(2),str(now.day).zfill(2),str(now.hour).zfill(2),str(now.minute).zfill(2),'-'+str(now.second).zfill(2) if seconds else '')

def striphtml(data):
	import re
	p = re.compile(r'<.*?>')
	return p.sub('', data)




def write(fn,data,uni=True,toprint=True,join_line='\n',join_cell='\t',timestamp=None):
	if timestamp:
		from datetime import datetime
		ts=datetime.now().strftime('%Y-%m-%d_%H%M')
		fn='.'.join(fn.split('.')[:-1]) + '.' + ts + '.' + fn.split('.')[-1]

	if not uni:
		of = open(fn,'w')
	else:
		join_line=u'\n'
		join_cell=u'\t'
		import codecs
		of = codecs.open(fn,'w',encoding='utf-8')

	if type(data) in [list,tuple]:
		o=""
		for x in data:
			if type(x) in [list,tuple]:
				z=[]
				for y in x:
					if not uni and type(y)==type(u''):
						y=y.encode('utf-8')
					z+=[y]
				x=z

				try:
					line=join_cell.join(x)
				except TypeError:
					line=[]
					for y in x:
						try:
							yx=y.decode('utf-8')
						except:
							try:
								yx=str(y)
							except:
								yx=y
						line+=[yx]
					line=join_cell.join(line)
					#
					# if not uni:
					# 	line=join_cell.join(str(y) for y in x)
					# else:
					# 	line=join_cell.join(unicode(y) for y in x)
			else:
				try:
					line=str(x)
				except UnicodeEncodeError:
					line=x.encode('utf-8')
			line=line.replace('\r','').replace('\n','')
			#print [line+join_line], type(line)
			#print [join_line], type(join_line)
			of.write(line + join_line)
	else:
		try:
			o=str(data)
		except UnicodeEncodeError:
			o=unicode(data)
		of.write(o)
	of.close()
	if toprint:
		print ">> saved:",fn.encode('utf-8')

def uwrite(fn,data,toprint=True,join_line='\n',join_cell='\t'):
	if type(data)==type([]):
		o=u""
		for x in data:
			if type(x)==type([]):
				z=[]
				for y in x:
					if type(y)!=type(u''):
						try:
							y=y.decode('utf-8')
						except AttributeError:
							y=unicode(y)
					z+=[y]
				x=z
				line=join_cell.join(x)

			else:
				if type(x)!=type(u''):
					try:
						line=x.decode('utf-8')
					except AttributeError:
						line=unicode(x)
				else:
					line=x
			line=line.replace('\n','\r').replace('\n','')
			o+=line+join_line
	else:
		o=unicode(data)

	import codecs
	of = codecs.open(fn,'w',encoding='utf-8')
	of.write(o)
	of.close()
	if toprint: print ">> saved:",fn



def makeminlength(string,numspaces):
	if len(string) < numspaces:
		for i in range(len(string),numspaces):
			string += " "
	return string

def get_class( kls ):
    parts = kls.split('.')
    module = ".".join(parts[:-1])
    m = __import__( module )
    for comp in parts[1:]:
        m = getattr(m, comp)
    return m

def gleanPunc(aToken):
	aPunct = None
	while(len(aToken) > 0 and not aToken[0].isalnum()):
		aPunct = aToken[:1]
		aToken = aToken[1:]
	while(len(aToken) > 0 and not aToken[-1].isalnum()):
		aPunct = aToken[-1]
		aToken = aToken[:-1]
	return (aToken, aPunct)


def count(string, look_for):
    start   = 0
    matches = 0

    while True:
        start = string.find (look_for, start)
        if start < 0:
            break

        start   += 1
        matches += 1

    return matches

def slice(l,num_slices=None,slice_length=None,runts=True,random=False):
	"""
	Returns a new list of n evenly-sized segments of the original list
	"""
	if random:
		import random
		random.shuffle(l)
	if not num_slices and not slice_length: return l
	if not slice_length: slice_length=int(len(l)/num_slices)
	newlist=[l[i:i+slice_length] for i in range(0, len(l), slice_length)]
	if runts: return newlist
	return [lx for lx in newlist if len(lx)==slice_length]


def select2(dict_with_key_as_option_and_value_as_prob):
	import random
	import bisect
	import collections

	def cdf(weights):
	    total=sum(weights)
	    result=[]
	    cumsum=0
	    for w in weights:
	        cumsum+=w
	        result.append(cumsum/total)
	    return result

	def choice(population,weights):
	    assert len(population) == len(weights)
	    cdf_vals=cdf(weights)
	    x=random.random()
	    idx=bisect.bisect(cdf_vals,x)
	    return population[idx]

	items=dict_with_key_as_option_and_value_as_prob.items()
	weights=[v for k,v in items]
	population=[k for k,v in items]
	return choice(population,weights)

def select(dict_with_key_as_option_and_value_as_prob):
	import random
	d=dict_with_key_as_option_and_value_as_prob
	r = random.uniform(0, sum(d.itervalues()))
	s = 0.0
	for k, w in d.iteritems():
		s += w
		if r < s: return k
	return k


def choose(optionlist,msg="please select from above options [using commas for individual selections and a hyphen for ranges]:\n"):
	seldict={}

	selnum=0
	print
	print

	if type(optionlist)==type([]):
		for option in optionlist:
			selnum+=1
			seldict[selnum]=option
			print "\t"+"\t".join(str(x) for x in [selnum,option])
	elif type(optionlist)==type({}):
		for option,desc in optionlist.items():
			selnum+=1
			seldict[selnum]=option
			print "\t"+"\t".join(str(x) for x in [selnum,option,desc])

	inp=raw_input("\n\t>> "+msg+"\n\t").strip()
	sels=[]
	for np in inp.split(","):
		np=np.strip()
		if "-" in np:
			try:
				nn=np.split("-")
				for n in range(int(nn[0]),int(nn[1])+1):
					sels.append(seldict[n])
			except:
				continue
		else:
			try:
				sels.append(seldict[int(np)])
			except:
				continue

	return sels

def toks2str(tlist,uni=False):
	toks=[]
	putleft=False
	#print tlist
	for tk in tlist:
		tk=tk.strip()
		if not tk: continue
		tk = tk.split()[-1]
		if not tk: continue
		if (not len(toks)):
			toks+=[tk]
		elif putleft:
			toks[-1]+=tk
			putleft=False
		elif tk=='`':
			toks+=[tk]
			putleft=True
		elif tk=='-LRB-':
			toks+=['(']
			putleft=True
		elif tk=='-RRB-':
			toks[-1]+=')'
		elif len(tk)>1 and tk[0]=="'":
			toks[-1]+=tk
		elif tk[0].isalnum():
			toks+=[tk]
		elif tk.startswith('<') and '>' in tk:
			toks+=[tk]
		else:
			toks[-1]+=tk
	if uni: return u' '.join(toks)
	return ' '.join(toks)

def unescape(text,delete_remaining=True):
    def fixup(m):
        text = m.group(0)
        if text[:2] == "&#":
            # character reference
            try:
                if text[:3] == "&#x":
                    return unichr(int(text[3:-1], 16))
                else:
                    return unichr(int(text[2:-1]))
            except ValueError:
                pass
        else:
            # named entity
            try:
                text = unichr(htmlentitydefs.name2codepoint[text[1:-1]])
            except KeyError:
                pass
        return text

    x=re.sub("&#?\w+;", fixup, text)

    sgml_replacements=[
    ('&thorne;','th'),
    ('&hyphen;','-'),
    ('&wblank;','  '),
    ('&abar;','a'),
    ('&obar;','o'),
    ('&ebar;','e'),
    ('&ibar;','i'),
    ('&ubar;','u'),
    ('&nbar;','n'),
    ('&thornu;','th'),
    ('&thornt;','th'),
    ('&sblank;',' '),
    ('&indent;','     '),
    ('&lblank;',' '),
    ('&blank;',' '),
    ('&ast;',' ')
    ]


    for a,b in sgml_replacements: x=x.replace(a,b)

    if delete_remaining:
    	x=re.sub('&.*?;', '', x)

    return x


def escape_punc(punc,keep=['_']):
	if len(punc)>1:
		newword=[]
		for li,letter in enumerate(punc):
			newletter=escape_punc(letter)
			newword+=[newletter]
			if newletter!=letter and li+1!=len(punc):
				newword+=['_']
		return ''.join(newword)
	if punc in keep: return punc
	puncd={"'":"p_apos", ",":"p_comma", "!":"p_exclam", "-":"p_hyphen", ".":"p_period", "?":"p_ques", '"':"p_quote", ";":"p_semi"}
	puncd['_']='p_underscore'
	puncd[u'\u2014']='p_emdash'
	puncd['|']='p_pipe'
	puncd[':']='p_colon'
	puncd['&']='p_ampersand'
	puncd[u'\u201d']='p_rightquote'
	puncd[u'\u201c']='p_leftquote'
	puncd[']']='p_rightbracket'
	puncd['[']='p_leftbracket'
	puncd[')']='p_rightparenthesis'
	puncd['(']='p_leftparenthesis'
	puncd[u'\u2018']='p_leftquote_single'
	puncd[u'\u2019']='p_rightquote_single'
	puncd['`']='p_tilde'
	puncd['$']='p_dollarsign'
	if punc.isdigit():
		return 'num_'+str(punc)
	return puncd.get(punc,punc)

def mdwgo(folder='.'):
	for fn in os.listdir(folder):
		if not 'fmt_long' in fn: continue
		if not fn.endswith('.txt'): continue
		#mdw2viz(fn,cellcol='p-value_indiv',classes=True,minobs=30)
		mdw2viz(fn,cellcol='obs/exp',classes=True,minobs=30)


def mdw2net2(fn,minobs=3):
	import networkx as nx
	if '*' in fn:
		a,b=fn.split('*',1)
		fndir=os.path.split(fn)[0]
		if not fndir: fndir='.'
		G=nx.DiGraph()
		for _fn in os.listdir(fndir):
			if _fn.startswith(a) and _fn.endswith(b):
				print _fn
				_g=mdw2net2(os.path.join(fndir,_fn),minobs=minobs)
				for node in _g.nodes():
					if not G.has_node(node):
						G.add_node(node,**_g.node[node])

				for a,b,d in _g.edges(data=True): G.add_edge(a,b,**d)

		nx.write_gexf(G, fn+'.gexf')
		return G

	print ">> converting:",fn
	ld=tsv2ld(fn)
	g=nx.DiGraph()

	if 'Gender-Gender' in fn and 'document' in ld[0]:
		ld=[d for d in ld if d['document'] in ['Male-Male','Male-Female','Female-Female','Female-Male']]

	for d in ld:
		if d['obs']<minobs: continue
		for k in ['document','word']:
			nodeType=k
			node=d[k]
			if not g.has_node(node): g.add_node(node,nodeType=nodeType)
			d['weight']=1-d['p-value_indiv']
			d['weightstr']=str(round(d['weight']*100,1))+'%'
			g.add_edge(d['document'], d['word'], **d)
	nx.write_gexf(g, fn+'.minobs={0}.gexf'.format(minobs))
	return g

def mdw2net(fn,minp=0.1):
	import networkx as nx
	g=nx.Graph()
	metaphors='plague beast whore black feed laugh poison weep gold lips fiend madness'.split()
	for d in tsv2ld(fn):
		#if minp and d['p-value']>minp: continue
		for k in ['document','word']:
			nodeType=k if not d[k] in metaphors else 'word_metaphor'
			node=d[k]
			if not g.has_node(node): g.add_node(node,nodeType=nodeType)
		if d['obs/exp']>1:
			g.add_edge(d['document'], d['word'], **d)
	nx.write_gexf(g, fn+'.gexf')


def dd2mdw(fn,dd,sample=True,save=True,minnum=None):
	if sample:
		import random
		minlen=min([sum(dd[k].values()) for k in dd])
		minlen=minnum if minnum and minnum<minlen else minlen

		dl={}
		dd2={}
		for k in dd:
			words=[]
			for word,freq in dd[k].items():
				for n in range(freq): words+=[word]
			words=random.sample(words,minlen)
			dd2[k]=toks2freq(words)

		ld=dd2ld(dd2)
	else:
		print ">> dd2ld..."
		ld=dd2ld(dd)

	return ld2mdw(fn,ld,save=save)

def dl2mdw(fn,dl,sample=True):
	if sample:
		import random
		minlen=min([len(dl[k]) for k in dl])
		for k in dl:
			dl[k]=random.sample(dl[k],minlen)

	print ">> MDW:"
	for group in dl:
		print "\t>>",group,"-->",len(dl[group]),'words'
		dl[group]=toks2freq(dl[group])
	ld=dd2ld(dl)
	return ld2mdw(fn,ld)


def dd2mdw2(fn,dd,sample=True,save=False,minnum=None,maxnum=1000000):
	print ">>","dd2mdw2"
	if sample:
		import random
		#minlen=min([sum(dd[k].values()) for k in dd])
		#minlen=minnum if minnum and minnum<minlen else minlen
		#if minlen>maxnum: minlen=maxnum
		minlen=maxnum
		from scipy import stats


		for k in dd.keys():
			items=dd[k].items()
			weights=[x[1] for x in items]
			sumweight=float(sum(weights))
			weights=[w/sumweight for w in weights]
			population=[x[0] for x in items]
			pop_ids=[i for i,x in enumerate(population)]
			newd={}
			custm=stats.rv_discrete(name='custm', values=(pop_ids, weights))
			choices=custm.rvs(size=minlen)
			for i,x in enumerate(choices):
				xk=population[x]
				if not xk in newd: newd[xk]=0
				newd[xk]+=1

			dd[k]=newd
			print "RESAMPLED",k,len(dd[k]),sum(dd[k].values())

	ld=dd2ld(dd)
	print ld[0]
	if save:
		old=list(ld2mdw(fn,ld))
		write2(fn,old)
	else:
		return ld2mdw(fn,ld)



def ld2mdw(fn,ld,save=True,allcols=True):
	import rpyd2,rpy2
	print ">> ld2mdw init..."
	rekey={}
	for d in ld:
		for k in d.keys():
			try:
				if type(k)==str and rpyd2.rfy(k)==k: continue
			except rpy2.rinterface.RRuntimeError:
				pass
			#"""
			strk = k.encode('utf-8') if type(k) in [unicode] else str(k)
			hashk = 'X'+hash(strk)
			rekey[hashk]=k
			d[hashk]=d[k]
			del d[k]
	print ">> making rpyd2..."
	#print ld[0]
	r=rpyd2.RpyD2(ld,rownamecol='rownamecol',allcols=allcols)
	print r
	print ">> doing mdw..."
	"""
	old1,old2=r.mdw(returnType='both')
	for d in old1: d['word']=rekey.get(d['word'],d['word'])
	for d in old2: d['word']=rekey.get(d['word'],d['word'])
	if save:
		print ">> writing..."
		write2(fn.replace('.txt','.fmt_long.txt'),old1)
		write2(fn.replace('.txt','.fmt_wide.txt'),old2)
	return old1
	"""
	for d in r.mdw():
		d['word']=rekey.get(d['word'],d['word'])
		yield d

def mdw_long2wide(ld,rowcol='word',colcol='document',cellcol='obs/exp',pmin=None,minobs=3,empties=True):
	import pytxt
	dld=ld2dld(ld, rowcol)
	old=[]
	for rowname,ld in dld.items():
		if pmin and True in [d['p-value']>pmin for d in ld]: continue
		if minobs and True in [d['obs_total']<minobs for d in ld]: continue
		if not empties and True in [not d['obs_min'] for d in ld]: continue
		od={'word':rowname}
		for d in ld:
			colname=d[colcol]
			od[colname]=d[cellcol]
		old+=[od]

	return old

def mdw2wordnet(fn,minp=0.01,weight='p-value_indiv'):
	import networkx as nx,pytxt,pystats
	ld=fn if datatype(fn)=='ld' else tsv2ld(fn)
	g=nx.Graph()
	g2=nx.Graph()
	for i,d in enumerate(ld):
		if minp and d['p-value_indiv']>minp: continue
		print i
		g.add_edge(d['word'],d['document'],weight=d['p-value_indiv'])
	print ">> number of nodes, edges:",g.order(),g.size()
	print ">> computing shortest paths..."
	paths=nx.shortest_path(g,weight=weight)
	print ">> done."
	newedges={}
	for a in paths:
		for b in paths[a]:
			path=paths[a][b]
			print a, b, path
			if len(path)<2: continue
			pathweight=0
			for _a,_b in pytxt.bigrams(path):
				pathweight+=g[_a][_b]['weight']
			newedges[(a,b)]=pathweight

	zweight=pystats.zfy(newedges)
	for (a,b),weight in newedges.items():
		zw=zweight[(a,b)]
		if zw<0: continue
		g2.add_edge(a,b,weight=zw,sumweight=newedges[(a,b)])

	nx.write_gexf(g2, fn+'.wordnet.gexf')

def stanford2toks(fn,not_pos=['DT','CC','IN'],not_words=[],punc=False,lemmatize=True,replacements={},digits=False):
	import codecs,bs4
	t=codecs.open(fn,encoding='utf-8').read() if fn.endswith('.xml') else fn
	t=t.lower()
	bs=bs4.BeautifulSoup(t)
	words=[]
	for tok in bs.find_all('token'):
		pos=tok.find('pos').text.upper()
		if not_pos:
			if pos in not_pos: continue

		if not punc:
			if not pos[0].isalpha(): continue

		word=tok.find('word').text if not lemmatize else tok.find('lemma').text
		word=replacements.get(word,word)
		if not digits and word.isdigit(): continue
		if not_words and word in not_words: continue

		if not word[0].isalpha() and words:
			words[-1]+=word
		else:
			words+=[word]
	return words



def mdwlong2nozero(infolder,outfolder):
	for fn in os.listdir(infolder):
		ld=tsv2ld(os.path.join(infolder,fn))
		ld=[d for d in ld if d['obs'] and d['obs']!=0]
		write2(os.path.join(outfolder,fn), ld)


def mdwviz(folder='.',suffix='.fmt_long.txt',classes=True,minobs=3):
	import os
	for fn in sorted(os.listdir(folder)):
		if not fn.endswith(suffix): continue
		print fn,'...'
		#if fn<'Shakespeare.com.1598.much_ado': continue
		if os.path.exists(fn.replace(suffix,'.obsexp.minobs=5.words.pca.pdf')) and (not classes or os.path.exists(fn.replace(suffix,'.obsexp.minobs=5.classes.pvclust.pdf'))): continue
		try:
			mdw2viz(os.path.join(folder,fn),classes=classes,minobs=minobs)
		except Exception as e:
			print "!!"*50
			print "!!",e
			print "!!"*50


def mdw2viz(fn,classes=False,log=True,maxscore=5.0,cellcol='obs/exp',minobs=100):
	import rpyd2,math
	ld=tsv2ld(fn)
	if 'Gender-Gender' in fn and 'document' in ld[0]:
		print len(ld)
		ld=[d for d in ld if d['document'] in ['Male-Male','Male-Female','Female-Female','Female-Male']]
		print len(ld)
		print set(d['document'] for d in ld)
		print

	if 'document' in ld[0]: ld=mdw_long2wide(ld,cellcol=cellcol,minobs=minobs)
	ld=sorted(ld,key=lambda _dx: -sum([abs(math.log10(v)) for v in _dx.values() if type(v)==float and v]))
	todel=['inf']
	for d in ld:
		if type(d['word'])==float:
			try:
				d['word']=str(int(d['word']))
			except (ValueError,OverflowError) as e:
				continue

		d['word']=escape_punc(d['word'])
		if not d['word']: continue
		if 'greek' in fn and ascii(d['word'])==d['word']: todel+=[d['word']]
		for k in d:
			if not d[k]: d[k]=1.0
			if type(d[k])==float:
				print k, d[k]
				d[k]=math.log10(d[k]) if d[k] else 1
				print k, d[k]
				print
				if d[k]>maxscore: d[k]=maxscore
	ld=[d for d in ld if not d['word'] in todel]

	fn=fn.replace('fmt_wide.','').replace('fmt_long.','')
	for d in ld:
		for k in d.keys():
			if not k: del d[k]

	fn=fn.replace('.txt','.{0}.minobs={1}.txt'.format(cellcol.replace('/',''), minobs))
	r=rpyd2.RpyD2(ld,rownamecol='word',allcols=True,zero=1)
	r.pca(fn=fn.replace('.txt','.words.pca.pdf'))

	if classes:
		#r.t().kclust(fn=fn.replace('.txt','.classes.kclust.pdf'),k=2)
		r.pvclust(fn=fn.replace('.txt','.classes.pvclust.pdf'))

		#ld2=ld[:len(ld[0])-1]
		#r2=rpyd2.RpyD2(ld2,rownamecol='word',allcols=True,zero=1)
		#r2.t().pca(fn.replace('.txt','.classes.pca.pdf'))



##############################################################

def ld2bayes(ld,textkey='text',metakeys=[],samplekey=None,windowsize=100,ngrams=[1,2],save=True):
	## really, an ld2featsets
	feat_sets={}
	ld2=[]
	def addslice(feat,sliceid):
		if not feat in feat_sets: feat_sets[feat]=[]
		feat_sets[feat]+=[sliceid]

	for d in ld:
		if not textkey in d: continue
		print d
		slicenum=0
		words=tokenize2(d[textkey]) if type(d[textkey])!=list else d[textkey]
		for txt in slice(words,slice_length=windowsize,runts=False):
			slicenum+=1
			sliceid=(d['filename'], slicenum)

			for mkey in metakeys:
				mkeyval=unicode(mkey)+u':'+unicode(d[mkey])
				addslice(mkeyval,sliceid)

			## words
			for ngramlen in ngrams:
				for gram in ngram(txt,ngramlen):
					gramstr=u"_".join(gram)
					feat=u'{0}-gram:{1}'.format(ngramlen,gramstr)
					addslice(feat,sliceid)

	for feat in feat_sets:
		feat_sets[feat]=set(feat_sets[feat])

	if save:
		import cPickle
		cPickle.dump(feat_sets,open('feat_sets.{0}.pickle'.format('_'.join(sorted(ld[0].keys()))[:25]),'wb'))
	else:
		return sets2bayes(feat_sets)

def sets2bayes(feat_sets, ofn=None,filter_by=None, minobs=10, stopwords=['1-gram:p','2-gram:p_m','1-gram:m']):
	import os,cPickle,pystats,networkx as nx,random
	for sw in stopwords:
		if not sw in feat_sets: continue
		del feat_sets[sw]

	def cmp(k1,k2,feat_sets,allnums=None):
		numposs=len(allnums)
		## test whether k1's being the case increases the odds of k2's occurring
		ids_k1_occurring=set(feat_sets[k1])
		#allnums=set(range(numposs))
		ids_k1_NOToccurring=allnums-ids_k1_occurring

		minnum=min([len(ids_k1_occurring), len(ids_k1_NOToccurring)])
		if not minnum or minnum<30: return
		#control_NOTk1=set(random.sample(list(ids_k1_NOToccurring),minnum))
		#study_k1=set(random.sample(list(ids_k1_occurring),minnum))

		study_k1=ids_k1_occurring
		control_NOTk1=ids_k1_NOToccurring

		ids_k2_occurring=set(feat_sets[k2])
		k2_given_k1=len(study_k1 & ids_k2_occurring)
		notk2_given_k1=minnum-k2_given_k1
		p_k2_given_k1=k2_given_k1/len(study_k1)
		k2_given_notk1=len(control_NOTk1 & ids_k2_occurring)
		notk2_given_notk1=minnum-k2_given_notk1
		p_k2_given_notk1=k2_given_notk1/len(control_NOTk1)
		dx={}
		dx['Y_name']=k2
		dx['X_name']=k1
		dx['odds_Y_given_X']='{0} out of {1}'.format(k2_given_k1,len(study_k1))
		dx['odds_Y_given_notX']='{0} out of {1}'.format(k2_given_notk1,len(control_NOTk1))
		dx['odds_ratio2']=p_k2_given_k1/p_k2_given_notk1 if p_k2_given_notk1 else 0
		dx['odds_diff']=p_k2_given_k1 - len(ids_k2_occurring)/numposs
		dx['prob_X']=len(ids_k1_occurring)/numposs
		dx['prob_Y']=len(ids_k2_occurring)/numposs
		dx['prob_Y_given_X']=p_k2_given_k1
		dx['prob_Y_given_notX']=p_k2_given_notk1
		dx['num_X']=len(feat_sets[k1])
		dx['num_Y']=len(feat_sets[k2])
		dx['num_X_and_Y']=len(feat_sets[k1] & feat_sets[k2])
		if not dx['prob_Y_given_X'] or not dx['prob_Y_given_notX']: return
		return dx


	def do_featset(setfn, feat_sets):
		G=nx.DiGraph()
		print len(feat_sets)
		allnums=set(item for sublist in feat_sets.values() for item in sublist)
		numposs=len(allnums)
		i=0
		ii=0
		ilim=len(feat_sets)
		for k1,k2 in sorted(pystats.product(feat_sets.keys(),feat_sets.keys())):
			print k1,k2,len(feat_sets[k1]),len(feat_sets[k2]),'...'
			if k1==k2: continue
			k1x=set(k1.split(':',1)[1].split('_'))
			k2x=set(k2.split(':',1)[1].split('_'))
			k1prefix=k1.split(':')[0]
			k2prefix=k2.split(':')[0]
			#if k1prefix == k2prefix:
			#	continue

			if {k1prefix,k2prefix} in [ {'Type','Era'}, {'Era','Booth'}, {'Type','Booth'}, {'Space','Era'}, {'Space','Type'} ]:
				continue

			if len(k1x&k2x):
				print ">> skipping:",k1,k2
				continue


			d=cmp(k1,k2,feat_sets,allnums=allnums)
			if not d: continue
			#print d,'...'
			print ilim,ii,i
			print
			i+=1
			if i>ilim:
				i=0
				ii+=1
			if d['odds_diff']<0: continue
			for kx in [k1,k2]:
				_a,_b=kx.split(':',1)
				if not G.has_node(kx): G.add_node(kx,nodeType=_a,nodeName=_b,nodeFreq=len(feat_sets[kx]))
			G.node[k1]['nodeProb']=d['prob_X']
			G.node[k2]['nodeProb']=d['prob_Y']
			print G.node[k1]
			G.add_edge(k1,k2,weight=d['odds_ratio2'],Label=d['odds_ratio2'],**d)
		nx.write_gexf(G, setfn+'.gexf')

		return G


	if type(feat_sets) in [str,unicode]:
		print ">> loading sets:"
		feat_sets=cPickle.load(open(feat_sets))
		print ">> done."



	if not ofn: ofn='bayesnet.'+now()

	if filter_by:
		relevkeys=[k for k in feat_sets.keys() if k.startswith(filter_by)]
		minlen=min([len(feat_sets[key]) for key in relevkeys])
		for key in relevkeys:
			featsetnow={}
			allowablekeys=set(random.sample(list(feat_sets[key]), minlen))
			for kx in feat_sets.keys():
				setnow=set(list(feat_sets[kx]))&allowablekeys
				#if len(setnow)<lenlim/2: continue
				if len(setnow)<minobs: continue
				print kx, len(setnow)
				featsetnow[kx]=setnow
			print len(featsetnow)
			return do_featset(ofn+'.'+key, featsetnow)
	else:
		for k in feat_sets.keys():
			if len(feat_sets[k])<minobs:
				del feat_sets[k]

		print ">> # keys to analyze:", len(feat_sets)
		print feat_sets.keys()

		return do_featset(ofn, feat_sets)


##############################################################

def binyear(year,binsize,start=0):
	if binsize==1: return '{0}'.format(year)
	for a,b in bigrams(range(start,3000,binsize)):
		if year in range(a,b):
			return '{0}-{1}'.format(a,b-1)
	return '????-????'

def bin(num,step,start=0.0,max=None,zfill=3):
	numnow=start
	while numnow + step < num:
		if max and numnow + step>max: break
		numnow+=step


	if int(numnow) == numnow: numnow=int(numnow)
	if max and numnow + step >max: return str(numnow).zfill(zfill)+'-'
	return str(numnow).zfill(zfill)+'-'+str(numnow+step-1)




def toks2freq(toks,tfy=False):
	tokd={}
	for tok in toks:
		try:
			tokd[tok]+=1
		except:
			tokd[tok]=1

	if tfy:
		import pystats
		return pystats.tfy(tokd)
	else:
		return tokd

def text2toks(text):
	toks=[]
	for w in text.lower().split():
		for ww in w.split('-'):
			www=noPunc(ww)
			toks.append(www)
	return toks

def hash(string):
	import hashlib
	if type(string)==unicode:
		string=string.encode('utf-8')
	return str(hashlib.sha224(string).hexdigest())


def tokenize(speech,regex=r'(\s+|\w+|\S+|\W+)'):
	c=re.compile(regex,re.U)
	l=[]
	for w in [noPunc(x).lower() for x in c.findall(speech) if noPunc(x).strip()]:
		for ww in w.split('.'):
			if ww: l+=[ww]
	return l

def tokenize2(speech,punc=False,regex=r'(\s+|\w+|\S+|\W+)',lower=True):
	c=re.compile(regex,re.U)
	l=[]
	for w in [(x.lower() if lower else x) for x in c.findall(speech) if x.strip()]:
		wx=''
		if (w.startswith("'") or w.startswith("-")) and len(l):
			l[-1]+=w
			continue

		for ww in w:
			if ww.isalpha() or (wx and ww in [u"'",u"-"]):
				wx+=ww
			else:
				if wx: l+=[wx]
				if punc: l+=[ww]
				wx=''
		if wx: l+=[wx]
	return l

def sortd(dict):
	return sorted(dict.items(), key=lambda lx: -lx[1])



def crunch(objects,function_or_methodname,ismethod=None,nprocs=8,args=[],kwargs={}):
	import time,random
	ismethod=type(function_or_methodname) in [str,unicode] if ismethod is None else ismethod

	def do_preparse(text,args=[],kwargs={}):
		threadid=os.getpid()
		time.sleep(random.uniform(0,5))
		print "[{2}] Starting working on {0} at {1}".format(text if False else 'ObjectX', now(), threadid)
		#print ismethod,function_or_methodname,args,kwargs
		if ismethod:
			x=getattr(text,function_or_methodname)(*args,**kwargs)
		else:
			x=function_or_methodname(text, *args, **kwargs)

		print "[{2}] Finished working on {0} at {1}".format(text if False else 'ObjectX', now(), threadid)
		return x

	import thread,multiprocessing,os
	from multiprocessing import Process, Pipe
	from itertools import izip

	def spawn(f):
		def fun(q_in,q_out):
			numdone=0
			while True:
				numdone+=1
				i,x = q_in.get()
				if i == None:
					break
				q_out.put((i,f(x,args=args,kwargs=kwargs)))
		return fun

	def parmap(f, X, nprocs = multiprocessing.cpu_count()):
		q_in   = multiprocessing.Queue(1)
		q_out  = multiprocessing.Queue()

		proc = [multiprocessing.Process(target=spawn(f),args=(q_in,q_out)) for _ in range(nprocs)]
		for p in proc:
			p.daemon = True
			p.start()

		sent = [q_in.put((i,x)) for i,x in enumerate(X)]
		[q_in.put((None,None)) for _ in range(nprocs)]
		res = [q_out.get() for _ in range(len(sent))]

		[p.join() for p in proc]

		return [x for i,x in sorted(res)]

	parmap(do_preparse, objects, nprocs=nprocs)
	return True




import re

#Define exceptions
class RomanError(Exception): pass
class OutOfRangeError(RomanError): pass
class NotIntegerError(RomanError): pass
class InvalidRomanNumeralError(RomanError): pass

#Define digit mapping
romanNumeralMap = (('M',  1000),
				   ('CM', 900),
				   ('D',  500),
				   ('CD', 400),
				   ('C',  100),
				   ('XC', 90),
				   ('L',  50),
				   ('XL', 40),
				   ('X',  10),
				   ('IX', 9),
				   ('V',  5),
				   ('IV', 4),
				   ('I',  1))

def toRoman(n):
	"""convert integer to Roman numeral"""
	if not (0 < n < 5000):
		raise OutOfRangeError, "number out of range (must be 1..4999)"
	if int(n) != n:
		raise NotIntegerError, "decimals can not be converted"

	result = ""
	for numeral, integer in romanNumeralMap:
		while n >= integer:
			result += numeral
			n -= integer
	return result

#Define pattern to detect valid Roman numerals
romanNumeralPattern = re.compile("""
	^					# beginning of string
	M{0,4}				# thousands - 0 to 4 M's
	(CM|CD|D?C{0,3})	# hundreds - 900 (CM), 400 (CD), 0-300 (0 to 3 C's),
						#			 or 500-800 (D, followed by 0 to 3 C's)
	(XC|XL|L?X{0,3})	# tens - 90 (XC), 40 (XL), 0-30 (0 to 3 X's),
						#		 or 50-80 (L, followed by 0 to 3 X's)
	(IX|IV|V?I{0,3})	# ones - 9 (IX), 4 (IV), 0-3 (0 to 3 I's),
						#		 or 5-8 (V, followed by 0 to 3 I's)
	$					# end of string
	""" ,re.VERBOSE)

def fromRoman(s):
	"""convert Roman numeral to integer"""
	if not s:
		return None
		raise InvalidRomanNumeralError, 'Input can not be blank'
	if not romanNumeralPattern.search(s):
		return None
		raise InvalidRomanNumeralError, 'Invalid Roman numeral: %s' % s

	result = 0
	index = 0
	for numeral, integer in romanNumeralMap:
		while s[index:index+len(numeral)] == numeral:
			result += integer
			index += len(numeral)
	return result



#!/usr/bin/env python
#coding:utf-8
# Author: Alejandro Nolla - z0mbiehunt3r
# Purpose: Example for detecting language using a stopwords based approach
# Created: 15/05/13



#----------------------------------------------------------------------
def _calculate_languages_ratios(text):
	"""
	Calculate probability of given text to be written in several languages and
	return a dictionary that looks like {'french': 2, 'spanish': 4, 'english': 0}

	@param text: Text whose language want to be detected
	@type text: str

	@return: Dictionary with languages and unique stopwords seen in analyzed text
	@rtype: dict
	"""

	from nltk import wordpunct_tokenize
	from nltk.corpus import stopwords


	languages_ratios = {}

	'''
	nltk.wordpunct_tokenize() splits all punctuations into separate tokens

	>>> wordpunct_tokenize("That's thirty minutes away. I'll be there in ten.")
	['That', "'", 's', 'thirty', 'minutes', 'away', '.', 'I', "'", 'll', 'be', 'there', 'in', 'ten', '.']
	'''

	tokens = wordpunct_tokenize(text)
	words = [word.lower() for word in tokens]

	# Compute per language included in nltk number of unique stopwords appearing in analyzed text
	for language in stopwords.fileids():
		stopwords_set = set(stopwords.words(language))
		words_set = set(words)
		common_elements = words_set.intersection(stopwords_set)

		languages_ratios[language] = len(common_elements) # language "score"

	return languages_ratios


#----------------------------------------------------------------------
def detect_language(text):
	"""
	Calculate probability of given text to be written in several languages and
	return the highest scored.

	It uses a stopwords based approach, counting how many unique stopwords
	are seen in analyzed text.

	@param text: Text whose language want to be detected
	@type text: str

	@return: Most scored language guessed
	@rtype: str
	"""

	import sys
	ratios = _calculate_languages_ratios(text)

	most_rated_language = max(ratios, key=ratios.get)

	return most_rated_language







def morphadorn(text):

	"""text=Mary+had+a+lytle+lamb+whose+vertue+was+strykinge+as+snowe&
	corpusConfig=eme&
	includeInputText=true&
	media=json&
	xmlOutputType=outputPlainXML&
	adorn=Adorn"""

	from simplejson import JSONDecodeError
	import requests
	rd={}
	rd['corpusConfig']='eme'
	rd['text']=text
	rd['media']='json'
	rd['adorn']='Adorn'
	rd['xmlOutputType']='outputPlainXML'
	rd['includeInputText']=False
	import random
	from requests import ConnectionError
	while True:
		port=random.choice(['8182','8183'])
		print ">> morphadorning on port",port
		try:
			r=requests.post('http://localhost:'+port+'/partofspeechtagger',data=rd)
			return r.json()
		except (JSONDecodeError,ConnectionError) as e:
			print "!!",e
			print "!! failed to get data from morphadorn server, trying again after a short nap"
			from random import randint
			from time import sleep
			naptime=randint(1,60)
			print ">> napping for {0}sec...".format(naptime)
			sleep(naptime)
			print ">> awake."


def mongosizes():
	import pymongo
	db=pymongo.Connection().litlab
	for cname in db.collection_names():
		cdb=getattr(db,cname)


def sample_bycount(ld,feature,wanted_count,acceptable_range=10000):
	dec_ld=ld
	ready=False
	while not ready:
		#print "looping..."
		random.shuffle(dec_ld)
		chosen_texts=[]
		for _d in dec_ld:
			chosen_texts+=[_d]
			summ=sum([__d[feature] for __d in chosen_texts])

			diff=summ-wanted_count
			diff_abs=abs(diff)
			#print summ, decade2numword[decade], diff
			if diff_abs<acceptable_range:
				ready=True
				break
			elif diff>acceptable_range:
				break

	return summ,chosen_texts

def stopwords():
	import codecs
	return set(codecs.open('/Users/ryan/PHD/DH/TOOLS/stopwords.txt',encoding='utf-8').read().split('\n'))



color_hexes=['#4B0082', '#0000CD', '#228B22', '#C71585', '#FF00FF', '#9400D3', '#20B2AA', '#90EE90', '#32CD32', '#FF6347', '#9370DB', '#A0522D', '#8A2BE2', '#191970', '#6A5ACD', '#808080', '#483D8B', '#5F9EA0', '#DC143C', '#DB7093', '#FF0000', '#3CB371', '#8B008B', '#EE82EE', '#FFB6C1', '#008080', '#48D1CC', '#AFEEEE', '#8FBC8F', '#00FF00', '#708090', '#00FFFF', '#9ACD32', '#7FFFD4', '#8B4513', '#7B68EE', '#FF00FF', '#00FA9A', '#FF7F50', '#9932CC', '#DEB887', '#CD853F', '#BDB76B', '#0000FF', '#6B8E23', '#2F4F4F', '#66CDAA', '#FF8C00', '#B8860B', '#FFC0CB', '#FA8072', '#ADD8E6', '#FFA500', '#696969', '#87CEEB', '#40E0D0', '#B0C4DE', '#CD5C5C', '#F08080', '#4682B4', '#FF69B4', '#00FFFF', '#8B0000', '#800000', '#BC8F8F', '#7FFF00', '#00FF7F', '#00BFFF', '#7CFC00', '#006400', '#008000', '#FFDAB9', '#D2691E', '#008B8B', '#4169E1', '#556B2F', '#B22222', '#FFDEAD', '#C0C0C0', '#F0E68C', '#F5DEB3', '#E9967A', '#DA70D6', '#1E90FF', '#ADFF2F', '#00008B', '#B0E0E6', '#FAF0E6', '#FFE4B5', '#FFFAFA', '#000000', '#F5F5F5', '#E6E6FA', '#FFEFD5', '#FFA07A', '#00CED1', '#FFFFE0', '#BA55D3', '#F5F5DC', '#FFFF00', '#DAA520', '#808000', '#000080', '#A9A9A9', '#F4A460', '#FF1493', '#A52A2A', '#D8BFD8', '#DDA0DD', '#800080', '#98FB98', '#2E8B57', '#D2B48C', '#87CEFA', '#FF4500', '#6495ED', '#FAFAD2', '#778899']


fields={}
fields['HardSeed']="come,go,drop,stand,touch,see,pick,pull,put,get,lie,look,keep,run,stoop,fetch,jump,scratch,knock,watch,catch,roll,crack,work,tilt,fling,crawl,lean,swing,trip,kick,move,show,creep,stroll,tap,walk,glimpse,turn,grow,wait,eat,quiver,let,hurt,smell,slice,slip,break,leap,sweep,cramp,open,close,lift,hang,shiver,sit,bend,crush,stride,grind,red,white,blue,green,brown,yellow,black,purple,orange,down,out,back,up,over,under,above,off,behind,inside,outside,through,between,front,along,three,five,two,seven,eight,ten,twenty,four,thirty,six,sixty,nine,forty,fifty,hard,rough,flat,round,clear,liquid,bushy,hot,sharp,clean,wet,heavy,stiff,low,backward,wooden,ripe,transparent,bare,straight,dusky,tight,crooked,empty,slow,wide,apart,big,dry,loose,thin,thick,scalp,jowl,jowls,forehead,brow,eyelid,eyelids,cheek,cheeks,jaw,limb,limbs,neck,skull,nape,throat,gorge,breast,torso,waist,buttocks,loin,loins,hip,belly,navel,lap,pubes,groin,arm,arms,armpit,armpits,forearm,forearms,wrist,wrists,foot,feet,thigh,thighs,shin,shins,ankle,ankles,palm,palms,toe,toes,instep,forefinger,thumb,thumbs,skin,dimple,dimples,nipple,nipples,eyebrow,eyebrows,eyelash,eyelashes,nostril,nostrils,moustache,beard,bone,bones,flesh,marrow,knuckle,knuckles,skeleton,spine,backbone,pelvis,collarbone,rib,ribs,muscle,muscles,sinew,sinews,tendon,tendons,tongue,chest,eyeball,eyeballs,gut,womb,teeth,molar,molars,stomach,bowel,bowels,intestine,intestines,gland,glands,liver,spleen,kidney,kidneys,genitals,penis,vagina,lung,lungs,vein,veins,artery,arteries,blood,pus,nerves,nerve,finger,fingers,face,hair,chin,hand,hands,fist,head,nose,forehead,lip,lips,shoulder,shoulders,elbow,elbows,tooth,mouth,ear,ears,eye,eyes,knee,knees,heel,heels,leg,legs,white,silver,black,sable,red,ruddy,rubor,scarlet,green,verdure,sinople,verd,yellow,topaz,azure,blue,jonquil,gold,orange,aventurine,saffron,buff,nankeen,chamois,jessamy,sapphire,delphinium,indigo,bice,fesse,watchet,azury,cerulean,celeste,ciel,aquamarine,turqoise,cyan,aqua,teal,tango,purple,gridelin,puce,grey,gray,putty,beige,grege,kasha,taupe,canescence,hoariness,whiteness,redness,blackness,dun,brown,burnet,sorrel,russet,mordore,castor,teak,filemot,sandalwood,tan,tawny,umber,fallow,coccyn,cocke,vermilion,vermelet,vermeil,ponceau,fuchsia,crimson,murrey,tuly,crimson,sanguine,rubine,carmine,incarnadine,rufe,chaudron,nacarat,pink,verdour,corbeau,celadon,reseda,feuille,mignonette,chartreuse,0,1,2,3,4,5,6,7,8,9,10,11,one,two,three,four,five,six,seven,eight,nine,ten,eleven,twelve,thirteen,fourteen,fifteen,sixteen,seventeen,eighteen,nineteen,twenty,thirty,forty,fifty,sixty,seventy,eighty,ninety,hundred,thousand,million,billion,trillion,dozen,couple,pair,came,coming,comes,went,gone,going,goes,dropped,dropping,dropt,drops,stood,standing,stands,touched,touches,touching,saw,seeing,seen,sees,picked,picking,picks,pulled,pulling,pulls,putting,puts,got,getting,gotten,gets,lay,lying,lies,lain,looked,looking,looks,kept,keeping,keeps,ran,running,runs,stooped,stooping,stoops,fetched,fetching,fetches,jumped,jumping,jumps,scratched,scratches,scratching,knocked,knocking,knocks,watched,watching,watches,caught,catching,catches,rolled,rolling,rolls,cracked,cracking,cracks,worked,working,works,tilted,tilting,tilts,flung,flinging,flings,crawled,crawling,crawls,leaned,leaning,leans,swung,swinging,swings,tripped,tripping,trips,kicked,kicking,kicks,moved,moving,moves,showed,shown,showing,shows,creeped,crept,creeping,creeps,strolled,strolling,strolls,tapped,tapping,taps,walked,walking,walks,glimpsed,glimpsing,glimpses,turned,turning,turns,grew,grown,growing,grows,waited,waiting,waits,ate,eaten,eating,eats,quivered,quivering,quivers,letting,lets,hurting,hurts,smelled,smelling,smells,sliced,slicing,slices,slipped,slipping,slips,broke,broken,breaking,breaks,leaped,leaping,leapt,leaps,swept,sweeping,sweeps,cramped,cramping,cramps,opened,opening,opens,closed,closing,closes,lifted,lifting,lifts,hung,hanged,hanging,hangs,shivered,shivering,shivers,sat,sitting,sits,bent,bending,bends,crushed,crushing,crushes,strode,stridden,strides,ground,grinding,grinds".split(',')
fields['AbstractValues']="modesty,delicacy,elegant,restraint,sensibility,prudent,sensible,softness,ostentation,vanity,extravagance,excess,indulge,indulgence,profligate,zeal,unrelenting,superiority,insensibility,moderation,self-control,decorum,propriety,proper,sober,sobriety,coarse,temperate,lewd,simplicity,indecent,indecency,decency,decent,demure,chaste,moderate,presumption,presumptuous,arrogance,arrogant,pride,proud,subdued,reserve,reserved,deference,seemly,unseemly,indecorous,decorous,appropriate,inappropriate,appropriateness,inappropriateness,conformity,orderly,disorderly,modest,immodest,polite,politeness,becoming,unbecoming,indelicate,indelicacy,cautious,caution,circumspect,circumspection,restrained,unrestrained,self-restraint,sensitive,sensitiveness,sensible,passion,passionate,reasonable,judicious,judiciousness,discreet,discretion,courteous,courtesy,rash,reckless,recklessness,impetuous,impetuousness,impulsive,impulsiveness,considerate,inconsiderate,mildness,mild,gentle,gentleness,ostentation,vainglorious,pompous,pompousness,boastful,boast,bragging,humble,humility,haughtiness,haughty,supercilious,superciliousness,conceited,conceit,conceitedness,excessive,extravagant,prodigality,prodigal,emotional,scrupulous,indecent,manners,morality,moral,immoral,sin,sinful,indecent,manners,morality,moral,immoral,sin,sinful,shame,shameful,etiquette,reputation,reputable,disreputable,virtuous,temper,licentious,licentiousness,debauched,debauchery,conduct,character,virtue,faultless,excellence,esteem,admiration,nobility,dignity,dignified,principle,depravity,unworthy,infamy,despise,vice,prejudice,prejudiced,prejudiced,partiality,partial,impartial,objective,objectivity,objectively,sober-minded,correct,impartial,detached,impartiality,dispassionate,unpassionate,disinterested,unbiased,subjective,subjectively,bias,biased,prepossession,prepossessed,prejudgement,preconceit,preconception,inveterate,prejudicial,prejudicate,prejudicated,prejudicately,prejudicially,prepossess,passion,passionate,passionately,impassioned,emotion,emotional, emotionally,sensible,sensibleness,sensibility,sentiment,sentimental,sentimentally,sentimentality,sentimentalizing,sentimentalist,sentimentalizer,sentimentalize,mawkish,mawkishness,sentimentalism,feeling,affect,heart,pang,affective,emotive,affectional,affectively,feelingly,bosom,heartward,maudlin,emotionalism,maudlinly,mawkishly,zealous,blushing,passionateness,master-passion,passioning,ebullition,ecstasy,roiled,pathetic,full-hearted,passioned,passionable,vehement,hot-blooded,passionful,vehemently,pathetically,impassionately,impassion,fervour,ardor,ardour,ardurous,fervent,fervency,fervently,fervid,fervidly,perfervor,perfervour,ardent,ardently,burningly,enkindle,dispassion,dispassionate,dispassionately,unsentimental,unsentimentality,dispassionateness,passionlessness,impassive,impassiveness,insensibility,unfeeling,unfeelingness,detachment,frigidity,cold-hearted,callous,callousness,heartless,heartlessness,unpassionate,passionless,unimpassioned,throbless,feelingless,disimpassioned,emotionless,unemotional,unzealous,hard-hearted,hardheartedness,stony-hearted,dedolent,unblushing,cold-blooded,stolid,tearless,inexpressive,expressionless,passionlessly,cold-bloodedly,stolidly,unemotionally,coldly,icily,frigidly,heartlessly,unzealously,callously,unsentimentalize".split(',')
fields['AbstractValues_SocialRestraint']='modesty,delicacy,elegant,restraint,prudent,softness,ostentation,vanity,extravagence,excess,indulge,indulgence,profligate,unrelenting,superiority,moderation,self-control,decorum,propriety,proper,sober,sobriety,coarse,temperate,simplicity,indecent,indecency,decency,decent,demure,moderate,presumption,presumptuous,arrogance,arrogant,pride,proud,subdued,reserve,reserved,deference,seemly,unseemly,indecorous,decorous,appropriate,inappropriate,appropriateness,inappropriateness,conformity,orderly,disorderly,modest,immodest,polite,politeness,becoming,unbecoming,indelicate,indelicacy,cautious,caution,circumspect,circumspection,restrained,unrestrained,self-restraint,sensitive,sensitiveness,reasonable,judicious,judiciousness,discreet,discretion,courteous,courtesy,rash,reckless,recklessness,impetuous,impetuousness,impulsive,impulsiveness,considerate,inconsiderate,mildness,mild,gentle,gentleness,ostentation,vainglorious,pompous,pompousness,boastful,boast,bragging,humble,humility,haughtiness,haughty,supercilious,superciliousness,conceited,conceit,conceitedness,excessive,extravagant,prodigality,prodigal,scrupulous,sensible,sensibleness,restrain,civility,agreeable,gracious,genteel,urbanity,nicety,refined,refinement,uncouth,intemperate,intemperance,immoderate,wanton,wantonness,extravagance,insolent,insolence,rudeness,incivility,condescension,condescending,thoughtless,indiscreet,ungracious,misbehaviour,impropriety,rusticity,grossness,vulgar,improper,ungenteel,gross,solicitous,sobriety,overbearing,petulant,petulance,impertinent,impertinence,affront,impudence,impudent,affectation,unassuming'.split(',')
fields['AbstractValues_MoralValuation']='manners,morality,moral,immoral,sin,sinful,indecent,manners,morality,moral,immoral,sin,sinful,shame,shameful,etiquette,reputation,reputable,disreputable,virtuous,temper,licentious,licentiousness,debauched,debauchery,conduct,character,virtue,faultless,excellence,esteem,admiration,nobility,dignity,dignified,principle,depravity,unworthy,infamy,despise,vice,lewd,lewdness,badness,sinful,iniquity,iniquitous,malice,malicious,maliciousness,malignity,malignant,flagrant,villainy,heinous,turpitude,perverse,perverseness,depravity,depraved,deprave,wicked,pernicious,wickedness,outrageous,inexcusable,unpardonable,reprobate,debauch,misconduct,shameful,unwholesome,foul,depravation,debased,degenerate,corrupt,corruption,reprobate,corrupted,misconduct,guilt,guilty,ribaldry,goodness,valour,respectable,honour,magnanimity,magnanimous,integrity,irreproachable,reproach,rectitude,righteous,righteousness,undefiled,uncorrupt,uncorrupted,tainted,untainted,incorruptible,corruptible,corrupting,innocent,innocence,guiltless,respect,admire,admiration,admired,esteemed,respected,worthy,merit,dignified,contempt,contemptible,despised,despicable,laudable,sordid,disdained,disgrace,infamy,infamous,notorious,scandalous,reproof,reprove,ignominy,disgraceful,ignominious,disgraced,ashamed,baseness,chastity,chaste'.split(',')
fields['AbstractValues_Sentiment']='passion,passionate,passionately,impassioned,emotion,emotional, emotionally,sensibility,sentiment,sentimental,sentimentally,sentimentality,sentimentalizing,sentimentalist,sentimentalizer,sentimentalize,mawkish,mawkishness,sentimentalism,feeling,affect,heart,pang,affective,emotive,affectional,affectively,feelingly,bosom,heartward,maudlin,emotionalism,maudlinly,mawkishly,zealous,blushing,passionateness,master-passion,passioning,ebullition,ecstasy,roiled,pathetic,full-hearted,passioned,passionable,vehement,hot-blooded,passionful,vehemently,pathetically,impassionately,impassion,fervour,ardor,ardour,ardurous,fervent,fervency,fervently,fervid,fervidly,perfervor,perfervour,ardent,ardently,burningly,enkindle,dispassion,dispassionate,dispassionately,unsentimental,unsentimentality,dispassionateness,passionlessness,impassive,impassiveness,insensibility,unfeeling,unfeelingness,frigidity,cold-hearted,callous,callousness,heartless,heartlessness,unpassionate,passionless,unimpassioned,throbless,feelingless,disimpassioned,emotionless,unemotional,unzealous,hard-hearted,hardheartedness,stony-hearted,dedolent,unblushing,cold-blooded,stolid,tearless,inexpressive,expressionless,passionlessly,cold-bloodedly,stolidly,unemotionally,coldly,icily,frigidly,heartlessly,unzealously,callously,unsentimentalize'.split(',')
fields['AbstractValues_Partiality']='prejudice,prejudiced,prejudiced,partiality,partial,impartial,objective,objectivity,objectively,sober-minded,correct,impartial,detached,impartiality,disinterested,unbiased,subjective,subjectively,bias,biased,prepossession,prepossessed,prejudgement,preconceit,preconception,inveterate,prejudicial,prejudicate,prejudicated,prejudicately,prejudicially,prepossess,bigotry,bigot,detachment,disinterestedness'.split(',')
fields['HardSeed_ActionVerbs']='come,go,drop,stand,touch,see,pick,pull,put,get,lie,look,keep,run,stoop,fetch,jump,scratch,knock,watch,catch,roll,crack,work,tilt,fling,crawl,lean,swing,trip,kick,move,show,creep,stroll,tap,walk,glimpse,turn,grow,wait,eat,quiver,let,hurt,smell,slice,slip,break,leap,sweep,cramp,open,close,lift,hang,shiver,sit,bend,crush,stride,grind,came,coming,comes,went,gone,going,goes,dropped,dropping,dropt,drops,stood,standing,stands,touched,touches,touching,saw,seeing,seen,sees,picked,picking,picks,pulled,pulling,pulls,putting,puts,got,getting,gotten,gets,lay,lying,lies,lain,looked,looking,looks,kept,keeping,keeps,ran,running,runs,stooped,stooping,stoops,fetched,fetching,fetches,jumped,jumping,jumps,scratched,scratches,scratching,knocked,knocking,knocks,watched,watching,watches,caught,catching,catches,rolled,rolling,rolls,cracked,cracking,cracks,worked,working,works,tilted,tilting,tilts,flung,flinging,flings,crawled,crawling,crawls,leaned,leaning,leans,swung,swinging,swings,tripped,tripping,trips,kicked,kicking,kicks,moved,moving,moves,showed,shown,showing,shows,creeped,crept,creeping,creeps,strolled,strolling,strolls,tapped,tapping,taps,walked,walking,walks,glimpsed,glimpsing,glimpses,turned,turning,turns,grew,grown,growing,grows,waited,waiting,waits,ate,eaten,eating,eats,quivered,quivering,quivers,letting,lets,hurting,hurts,smelled,smelling,smells,sliced,slicing,slices,slipped,slipping,slips,broke,broken,breaking,breaks,leaped,leaping,leapt,leaps,swept,sweeping,sweeps,cramped,cramping,cramps,opened,opening,opens,closed,closing,closes,lifted,lifting,lifts,hung,hanged,hanging,hangs,shivered,shivering,shivers,sat,sitting,sits,bent,bending,bends,crushed,crushing,crushes,strode,stridden,strides,ground,grinding,grinds'.split(',')
fields['HardSeed_BodyParts']='scalp,jowl,jowls,forehead,brow,eyelid,eyelids,cheek,cheeks,jaw,limb,limbs,neck,skull,nape,throat,gorge,breast,torso,waist,buttocks,loin,loins,hip,belly,navel,lap,pubes,groin,arm,arms,armpit,armpits,forearm,forearms,wrist,wrists,foot,feet,thigh,thighs,shin,shins,ankle,ankles,palm,palms,toe,toes,instep,forefinger,thumb,thumbs,skin,dimple,dimples,nipple,nipples,eyebrow,eyebrows,eyelash,eyelashes,nostril,nostrils,moustache,beard,bone,bones,flesh,marrow,knuckle,knuckles,skeleton,spine,backbone,pelvis,collarbone,rib,ribs,muscle,muscles,sinew,sinews,tendon,tendons,tongue,chest,eyeball,eyeballs,gut,womb,teeth,molar,molars,stomach,bowel,bowels,intestine,intestines,gland,glands,liver,spleen,kidney,kidneys,genitals,penis,vagina,lung,lungs,vein,veins,artery,arteries,blood,pus,nerves,nerve,finger,fingers,face,hair,chin,hand,hands,fist,head,nose,forehead,lip,lips,shoulder,shoulders,elbow,elbows,tooth,mouth,ear,ears,eye,eyes,knee,knees,heel,heels,leg,legs'.split(',')
fields['HardSeed_Colors']='white,silver,black,sable,red,ruddy,rubor,scarlet,green,verdure,sinople,verd,yellow,topaz,azure,blue,jonquil,gold,orange,aventurine,saffron,buff,nankeen,chamois,jessamy,sapphire,delphinium,indigo,bice,fesse,watchet,azury,cerulean,celeste,ciel,aquamarine,turqoise,cyan,aqua,teal,tango,purple,gridelin,puce,grey,gray,putty,beige,grege,kasha,taupe,canescence,hoariness,whiteness,redness,blackness,dun,brown,burnet,sorrel,russet,mordore,castor,teak,filemot,sandalwood,tan,tawny,umber,fallow,coccyn,cocke,vermilion,vermelet,vermeil,ponceau,fuchsia,crimson,murrey,tuly,crimson,sanguine,rubine,carmine,incarnadine,rufe,chaudron,nacarat,pink,verdour,corbeau,celadon,reseda,feuille,mignonette,chartreuse'.split(',')
fields['HardSeed_Numbers']='0,1,2,3,4,5,6,7,8,9,10,11,two,three,four,five,six,seven,eight,nine,ten,eleven,twelve,thirteen,fourteen,fifteen,sixteen,seventeen,eighteen,nineteen,twenty,thirty,forty,fifty,sixty,seventy,eighty,ninety,hundred,thousand,million,billion,trillion,dozen,couple,pair'.split(',')
fields['HardSeed_SpatialPrepositions']='down,out,back,up,over,under,above,off,behind,inside,outside,through,between,front,along,within,amid,around,underneath,beneath,ahead,alongside,beside,across,beyond,throughout,toward,away'.split(',')
fields['HardSeed_PhysicalAdjectives']='hard,rough,flat,round,clear,liquid,bushy,hot,sharp,clean,wet,heavy,stiff,low,backward,wooden,ripe,transparent,bare,straight,dusky,tight,crooked,empty,slow,wide,apart,big,dry,loose,thin,thick'.split(',')
fields['WordOrigin_0850-1150']='eighty,eastward,ark,heels,as,all,one,so,more,may,any,other,some,such,time,only,great,two,most,made,first,said,then,now,man,can,could,do,same,must,much,many,mr,men,well,god,own,life,good,day,new,little,see,even,did,might,years,long,how,never,here,while,make,king,three,old,each,lord,thus,still,way,whole,house,another,work,church,last,too,world,say,far,given,again,come,right,give,water,know,few,year,name,nothing,less,mind,hand,came,thought,death,cannot,others,things,think,england,small,always,let,side,find,means,does,body,seen,go,father,TRUE,head,english,love,end,soon,young,often,words,either,known,done,land,heart,light,high,best,thing,back,half,following,indeed,times,better,christ,son,set,near,word,until,went,full,night,says,feet,kind,till,truth,sent,rather,next,children,together,gave,become,town,held,book,ground,hands,home,sometimes,sea,eyes,knowledge,saw,became,free,twenty,greater,short,north,works,hope,enough,white,friend,friends,look,thousand,read,six,south,making,further,length,told,mother,alone,rest,strong,nearly,morning,show,open,fire,wife,lady,third,room,something,master,afterwards,ought,ten,soul,tell,written,leave,lost,school,holy,born,felt,child,woman,care,began,west,late,knew,beyond,queen,speak,answer,going,heaven,eye,arms,lay,followed,forth,bring,fall,laid,met,dead,black,east,sun,women,living,looked,hear,bishop,deep,strength,none,dear,self,longer,hold,island,books,live,feeling,instead,greatest,iron,mean,road,asked,wish,kingdom,gold,fear,feel,field,gives,ready,seven,evil,led,months,need,red,eight,stood,hath,later,shown,meet,houses,makes,fell,horse,don,coming,table,foot,stand,street,hence,unless,giving,bear,anything,fair,hard,thirty,loss,sight,cold,higher,stone,gone,board,door,meeting,reached,middle,drawn,miss,sin,rich,ship,forward,wood,names,follow,trees,wrote,heat,looking,becomes,lead,help,food,highest,lands,comes,die,otherwise,upper,fully,rights,understand,green,filled,pope,lie,doing,hardly,silver,earl,twelve,writing,fifty,sides,bodies,thoughts,walls,greatly,ask,mouth,opened,broken,wild,follows,rose,saying,spring,meaning,tree,offered,leaving,run,leaves,hall,feelings,speech,named,winter,send,lives,answered,fish,takes,speaking,fresh,horses,worthy,marked,understood,wisdom,month,wind,heavy,worship,reading,summer,main,wall,minds,write,worth,writer,everything,glass,temple,reader,standing,deal,freedom,mark,altogether,western,wide,seeing,youth,churches,nine,height,play,shows,schools,ships,wise,goods,reach,week,northern,fit,sister,taught,highly,sat,spoke,writers,cloth,step,leading,southern,learn,waters,lies,hearts,showed,grey,pounds,wholly,dry,sons,plants,sweet,yes,mankind,notwithstanding,kings,share,spoken,arm,draw,working,hair,heads,narrow,stream,learning,beneath,towns,likewise,port,islands,wished,sought,offer,goes,knows,ones,toward,eastern,hills,kinds,growth,stands,stock,ways,boat,smith,seek,wine,bearing,birds,fallen,lords,warm,sold,sleep,ladies,talk,meant,weeks,hopes,flesh,streets,priest,tears,yellow,seldom,sake,weather,glad,break,shape,sword,fifteen,oh,fight,ends,strongly,wonder,shot,thick,driven,salt,souls,fast,drew,stones,bishops,steps,ear,worse,corn,tongue,sit,plant,smaller,grow,sixty,won,gods,priests,growing,coal,snow,lose,grounds,ran,utmost,watch,friendship,moon,thoroughly,sunday,path,slowly,rain,falls,sick,grew,sand,deeply,ice,bird,neck,romans,eat,forget,mighty,storm,tower,arise,fathers,teeth,looks,speaks,archbishop,thin,thereby,tells,song,drink,grown,bore,verse,woods,wishes,bones,forgotten,sorrow,falling,sharp,teach,arose,grass,masters,floor,gate,fly,freely,stars,gathered,marks,thousands,sheep,ears,yards,oath,sooner,worked,fill,elsewhere,yield,behold,wit,southeast,deed,ring,hoped,slow,inner,lordship,boats,murder,deemed,nearer,greeks,needed,triumph,goodness,clothes,roads,seed,milk,writes,bone,devil,fought,wound,doth,flight,bears,friendly,lime,flow,thinks,blessing,star,rooms,breath,clouds,hereafter,sixth,dust,roof,shortly,truths,finds,neighbouring,talents,dare,thanks,tail,gates,hast,shadow,worst,worn,outward,neighbours,gen,knight,doors,quick,evils,upwards,ended,followers,holds,younger,welcome,disciples,sell,sail,fled,breaking,winds,eleven,leads,thank,stronger,righteousness,hung,widow,loud,runs,sees,wear,hon,hat,kindly,deeds,older,arising,cloud,drive,bow,steel,tale,springs,bare,played,ere,lion,speed,yard,meat,bath,fox,hell,elder,shade,sing,wire,sixteen,wheat,tide,streams,shame,beg,sorry,ghost,burnt,silk,tables,empty,clean,lovely,wheel,verses,heavens,reaching,idle,beat,lee,sabbath,weary,seeking,tin,shell,sending,nearest,wore,fingers,dwell,overcome,fears,thine,key,tis,laugh,wives,hole,timber,arises,heathen,eggs,philosopher,laying,lo,nose,singing,thither,planted,neighbour,torn,leaf,ward,finger,seas,telling,greatness,naked,shoulders,watched,holiness,talent,limbs,swift,scottish,wet,writ,deeper,bit,staff,sets,ride,slain,bark,shoulder,sailed,grows,fired,eldest,philosophers,answers,burned,hollow,thirteen,songs,almighty,strongest,sisters,hidden,spare,blessings,seeds,epistle,yon,wrath,saturday,thickness,asking,net,monks,yielded,chest,playing,owe,knights,temples,harm,win,plays,feels,beheld,rests,mild,openly,ore,oak,mad,wounds,washed,sheet,flood,thorough,watching,throat,dried,gather,meal,warning,loves,burn,laughed,knees,inward,offers,shillings,ports,wool,ridge,speeches,noon,midnight,lamb,canon,smallest,worldly,highness,shells,heights,foe,rested,tie,nights,homes,wheels,hate,thunder,shining,lighted,liver,keen,forgive,shalt,horn,deer,righteous,guests,rightly,steep,tied,stem,yearly,crop,hide,spared,stretched,salts,nest,ninth,newly,merry,holes,lasted,endless,fires,dwelled,shook,gathering,borrowed,crops,hay,cheap,dared,needle,workmen,monk,brook,upright,unwilling,yea,fastened,bid,weapons,hare,owned,linen,tear,cat,seventeen,satan,boards,slept,cares,shake,foul,abbot,rope,hang,thread,helped,flows,oldest,shoes,awakened,begged,shadows,wells,harry,forgot,reaches,towers,dealt,bladder,hook,wash,upward,flies,limb,arisen,sixteenth,fold,cheek,shorter,beam,eaten,whither,hunger,thirst,rings,shield,laughter,guest,weighed,feathers,fullness,sung,kent,anchor,cow,dearest,deepest,knee,opens,melted,oft,manly,deadly,losses,burns,shooting,sworn,feeding,hither,thereto,egg,mid,twelfth,stir,westward,sends,swelling,sailing,pit,gladly,wider,hid,mouths,owed,string,handle,wishing,sprung,wax,mar,hated,seventeenth,ties,fifteenth,eighteenth,shine,wonders,thereon,cheeks,teaches,paths,meets,seamen,heavily,township,kiss,bowed,borough,nos,swear,honey,shoot,martyr,lays,storms,cart,psalm,woe,cared,bade,shades,lightly,craft,drank,wisely,yoke,frost,sore,shone,nigh,sorrows,breaks,sang,sails,tongues,horns,fourteenth,elders,beard,draws,tools,seeks,richly,tooth,thirds,hue,grove,dost,selling,swords,roses,lend,epistles,northwest,beams,loses,worms,answering,deem,mothers,christendom,worm,candle,wills,saddle,worshipped,penny,wolf,sheets,foes,asks,dish,mourning,mood,shaken,thyself,weapon,priesthood,arrows,sinful,weeping,heap,stolen,wondered,yields,canons,thirteenth,apple,stirring,deaths,forgiveness,oaths,fewer,pan,stretch,cock,ripe,stirred,pool,stepped,lock,richest,flowed,youngest,kissed,spake,weigh,letting,sweetness,stake,token,swell,southward,freed,hoping,ox,sown,chalk,arrow,steward,worlds,rains,ford,nut,sits,warmly,disciple,deaf,wedding,hie,dew,glasses,weighing,psalms,heating,oxen,reads,weep,danish,thankful,dreary,wept,highway,cliffs,hood,martyrs,keys,mist,cows,marsh,nought,wiser,wandered,hanged,wright,hairs,breathed,fain,na,sands,southwest,planting,easter,nineteenth,spear,folk,roar,earned,wires,owes,whites,fills,tread,sundry,stretching,darling,fearing,baths,popes,missed,locked,loudly,lean,folds,moors,shapes,cliff,shilling,heed,ditch,doomed,ribs,groves,locks,borrow,steal,richer,reed,welcomed,wake,deacon,thirdly,threats,folks,bows,mantle,triumphs,stead,sharply,grounded,tithes,bye,doom,reared,flint,halls,folded,ridges,moor,idleness,wines,strings,melting,mirth,devils,thief,weeds,awaken,nails,stocks,wisest,heath,spark,tor,mars,lions,wander,cheaper,lust,ins,leap,longest,mast,tame,swollen,wretch,threads,heavier,thereafter,roofs,stems,ores,thanked,feather,hats,harder,forwarded,forgiven,earn,reckless,fowl,rue,swore,wits,fisher,martyrdom,mare,lordships,heaps,shoe,crept,hearth,forgetting,forbear,hardness,thieves,candles,underneath,fullest,thumb,coals,gall,workman,shrink,blossoms,thickly,provost,helps,whale,turf,fathoms,goat,calf,oats,toes,lengths,begging,wearied,fan,sow,hangs,tides,brooks,wears,hedge,threshold,apples,mat,goats,thorn,shoots,pillow,shines,dishes,thigh,meals,forsake,weed,ness,hawk,hounds,ropes,leaned,lap,ringing,homeward,nests,sunlight,kin,didst,toe,grim,dearly,lame,birch,flee,twentieth,hides,blossom,pharisees,morn,forsaken,leapt,youths,dwells,wolves,freshness,tar,earls,healed,heal,waking,faster,fowls,carts,swim,bathing,bits,floods,errand,clung,richness,mourn,flax,hooks,reap,harp,creeping,drinks,readings,greeted,beads,therewith,sparks,ram,handful,womb,sweat,chill,marshes,reft,melt,canst,tool,web,anchored,aldermen,nail,fairest,marcus,widows,hart,foreseen,slew,understands,sap,nuts,shields,thorns,theft,drives,deals,warmed,swiftly,threat,span,boyhood,spur,lain,blacks,pharaoh,harper,watches,bishopric,pits,bids,alderman,sorrowful,loveliness,dale,shipped,crow,warmer,shots,swelled,sundays,lasts,heaped,swan,oar,tokens,cluster,din,comb,sweetest,gladness,fright,withstand,spears,nets,liking,nuns,meanest,sinned,shaw,witty,witch,call,cut,bad,trust,boy,cast,wealth,shore,formerly,girl,acknowledge,kill,sunshine,ugly,moonlight,trustworthy,charleston,place,days,since,away,brought,hundred,least,earth,early,heard,placed,business,places,daughter,title,forty,gospel,daily,fourth,witness,bread,showing,earlier,bringing,northeast,alike,apostles,apostle,earliest,crowd,lot,angels,mill,witnesses,yesterday,brings,fourteen,daughters,seventh,angel,listen,drove,hundreds,hunting,placing,bless,listened,feed,witnessed,childhood,flowing,titles,ashes,clothed,mills,englishman,losing,harvest,straw,listening,offspring,inland,englishmen,likeness,thursday,wednesday,hearers,hungry,gospels,betwixt,clad,ash,brightness,lots,crowds,barely,whisper,shower,hears,holiday,ale,stole,slide,belly,camels,danes,barn,camel,bowl,coldness,lore,beans,showers,aft,sights,shameful,part,law,ever,four,put,almost,five,parts,latter,line,miles,laws,blood,keep,kept,brother,turned,turn,evening,lines,dark,beginning,weight,else,husband,rise,bed,cross,bound,hill,health,ago,lake,inches,brown,FALSE,settled,thrown,bright,castle,top,truly,pretty,struck,grave,broad,january,partly,chosen,hot,walk,carefully,thereof,bottom,december,begin,november,circle,bridge,pride,finding,wounded,fleet,february,soft,steam,inch,brethren,fifth,fields,belong,brain,guilty,earnest,mile,lately,fever,saint,lips,careful,copper,stop,throw,jerusalem,temper,burning,quickly,buried,clay,organs,israel,breast,sad,proud,begun,stopped,putting,crossed,threw,bold,readers,edge,walked,game,clerk,beside,alive,choose,brothers,shut,box,frame,blind,bell,lest,saints,rough,begins,acres,cap,organ,wherever,therein,needs,altar,heavenly,bitter,eighteen,beds,landed,burden,threatened,strike,amid,bosom,afterward,bought,turns,drop,burst,pipe,rises,smoke,offering,pre,hebrew,cup,tall,busy,fat,cool,guilt,settle,loving,earthly,lively,feared,pound,buy,earnestly,eighth,sickness,dropped,scholars,inn,breach,creed,load,lakes,brass,flying,scholar,throwing,stern,bloody,lungs,lane,monday,purple,riding,cook,paradise,bench,edges,ashamed,risen,pine,kingdoms,hunt,prime,bounds,fulfilled,brow,drops,puts,stick,acre,ninety,posts,forbidden,asleep,wandering,stairs,keeps,circles,bride,lent,sink,reckoned,forehead,flock,butter,leather,pipes,friday,landlord,walks,pulled,sinking,boldly,bred,kitchen,bind,maiden,shaft,careless,sandy,ending,bridges,raw,dim,bee,crystals,bend,beer,nowhere,framed,restless,awake,weights,boxes,pull,fuller,bush,fortnight,den,hinder,hen,breed,graves,foremost,forbid,earnestness,undergo,belt,strikes,foster,carved,flocks,tops,lip,bitterness,stiff,warned,games,herein,glowing,lined,hymn,blast,eleventh,posted,crosses,stopping,frozen,throws,bees,clerks,baker,leaning,glow,sank,nineteen,crystal,hired,fare,abide,bending,hymns,undergone,barley,sadly,meadows,bitterly,amen,blowing,meadow,longing,moss,trap,healing,dumb,swallowed,bury,smoking,bushes,bliss,husbands,burdens,boldness,castles,axes,offerings,ant,hebrews,goldsmith,anew,softly,cheese,swallow,soap,reckon,shrine,ass,ware,lung,backs,hammer,knocked,cooling,hull,choosing,sticks,fin,mules,snake,blade,float,highlands,aught,chin,doe,brighter,herd,plaster,hire,stuck,mortar,godly,manifold,warn,shrubs,stricken,drain,darker,blew,herds,pulling,lily,crucified,alms,awoke,fetch,childish,freeman,blaze,cradle,highland,altars,forbade,kneeling,brains,asunder,shafts,tiger,bequeathed,awhile,throng,bursting,blindness,elbow,sadness,fork,dropping,mint,evenings,buying,climb,ladder,swimming,dip,cups,rainy,headache,freemen,beset,zion,fleets,floated,knot,sinks,loads,watery,stony,brightest,gleam,threaten,loft,bite,players,player,lighting,hale,ham,climbing,roughly,archdeacon,nostrils,frames,creeds,confessor,knock,lid,reins,mule,strand,fasting,begotten,frog,goose,breasts,fevers,stops,cedar,pines,broader,pepper,duck,goodly,roaring,boughs,drained,coolness,bridegroom,handling,sack,hunted,tearing,chooses,beaver,falsely,shift,berry,stillness,cooled,dealers,hindered,shifting,carelessness,ants,climbed,bloodshed,sour,close,fellow,june,july,october,september,understanding,sale,afford,darkness,closely,dog,afforded,wales,smooth,dogs,pole,fellows,affords,sheriff,poles,oct,curse,shepherd,haven,tuesday,winding,trout,affording,mamma,woollen,sales,smart,shepherds,closes,take,taken,took,according,service,taking,duke,etc,august,wrong,market,wonderful,aware,chancellor,knife,plot,beaten,aug,hit,webster,score,accord,tidings,chaplain,flank,purse,markets,pin,peel,wrongs,wonderfully,startled,startling,beseech,rider,dukes,links,accorded,knives,potter,kinsman,plots,every,also,both,therefore,left,large,court,manner,former,peace,council,crown,lower,grace,low,proved,prove,larger,serve,belief,courts,opening,spread,birth,mercy,blessed,jews,served,passion,fruit,spent,manners,oil,standard,skill,namely,teaching,nay,charity,gift,rent,passions,privileges,privilege,scattered,wings,fruits,largest,poverty,feeble,gifts,lowest,wing,start,hatred,dread,miracles,proves,spend,treasure,drinking,eating,lack,countess,councils,treasures,warmth,thrust,miracle,whoever,kindred,serves,jew,smell,lasting,thereupon,rents,needful,empress,proving,helpless,graces,crowns,dreaded,robbers,openings,standards,bidding,mercies,advent,teachings,spending,unbelief,boon,afar,churchyard,evangelists'.split(',')
fields['WordOrigin_1150-1700']='very,general,found,people,state,country,power,called,present,nature,order,form,during,government,because,public,sir,subject,several,just,use,number,point,course,certain,character,act,second,different,history,common,city,states,spirit,used,fact,person,necessary,account,cause,received,matter,reason,question,letter,persons,force,human,dr,effect,action,passed,age,purpose,family,system,natural,army,party,interest,river,view,money,period,united,sense,society,quite,generally,object,moment,possible,company,opinion,various,return,position,per,faith,influence,christian,face,chapter,formed,able,chief,authority,condition,important,ancient,carried,doubt,especially,probably,appear,round,office,duty,property,continued,respect,circumstances,language,considered,author,value,particular,change,appears,except,proper,attention,idea,prince,enemy,regard,fine,equal,immediately,forms,divine,degree,real,distance,appeared,pass,letters,moral,occasion,religious,principles,members,produced,result,march,sufficient,surface,considerable,obtained,required,evidence,principle,amount,beautiful,hour,due,vol,portion,political,mentioned,returned,favour,royal,judgement,original,similar,appearance,parliament,observed,employed,single,century,certainly,nation,captain,conduct,powers,officers,rule,entirely,class,labour,clear,note,future,supposed,battle,desire,existence,principal,hours,pleasure,troops,voice,direction,mere,command,jesus,entered,governor,private,practice,established,possession,published,constitution,paper,trade,majesty,merely,appointed,receive,notice,example,bill,foreign,past,species,really,simple,perfect,added,charge,experience,success,county,frequently,liberty,personal,extent,presence,learned,terms,suppose,doctrine,subjects,judge,numerous,complete,points,allowed,support,national,union,consequence,instance,sure,modern,passage,special,produce,acid,described,president,story,importance,measure,raised,science,ordinary,emperor,impossible,acts,colour,motion,remained,determined,college,progress,sound,military,remain,nations,usually,facts,danger,native,study,gentleman,arrived,attempt,capital,inhabitants,minister,difficulty,quantity,disease,peculiar,scarcely,beauty,plan,edition,sort,price,usual,applied,centre,consider,distinguished,difference,education,contrary,covered,noble,remains,necessity,civil,orders,intended,difficult,process,objects,advantage,plain,lived,causes,individual,relation,fixed,remarkable,equally,affairs,glory,increase,connexion,parties,method,duties,spiritual,presented,grand,direct,expression,connected,report,particularly,increased,section,front,consideration,visit,term,prepared,committee,replied,opposite,carry,virtue,removed,provided,volume,matters,passing,member,declared,congress,description,animal,expected,coast,series,page,results,stated,empire,proportion,effects,exercise,marriage,mountains,animals,conditions,style,engaged,size,piece,enter,attack,prevent,space,ideas,interesting,sum,pure,regarded,discovered,colonel,vessels,soldiers,scene,concerning,strange,proposed,village,obliged,superior,origin,information,aid,population,officer,district,expressed,proof,pain,entire,popular,situation,events,ordered,adopted,material,reign,views,operation,represented,forces,excellent,policy,division,remember,divided,supply,rendered,occupied,caused,opportunity,figure,married,prayer,introduced,numbers,mode,useful,practical,reference,notes,powerful,treatment,perfectly,square,theory,countries,date,memory,taste,relations,sacred,admitted,finally,save,measures,mountain,substance,succeeded,estate,application,advanced,defence,chiefly,claim,assembly,social,philosophy,enemies,confidence,purposes,article,articles,distinct,vast,directed,obtain,secret,soil,grant,interests,vessel,composed,possessed,rate,cent,pieces,trial,contained,separate,vain,pressure,questions,add,ministers,changes,gentlemen,accordingly,picture,previous,contains,honourable,patient,escape,opposition,entitled,universal,express,constant,classes,ages,simply,literature,killed,active,suddenly,fort,rock,founded,blue,opinions,attended,joy,journey,valuable,containing,stage,treaty,immediate,current,poet,solution,regular,variety,university,distant,genius,naturally,allow,granted,tried,rank,camp,conversation,local,province,fame,desired,administration,argument,secretary,moved,treated,season,christianity,gradually,fortune,resolved,catholic,cost,citizens,movement,instances,require,valley,external,accompanied,promise,professor,count,physical,directly,render,total,proceed,actual,advance,music,delivered,revolution,serious,suffered,capable,clock,extraordinary,evident,clearly,christians,execution,double,continue,design,secure,examination,temperature,pleased,attached,reduced,demand,protection,exist,derived,refused,spirits,affection,elements,actually,celebrated,servant,referred,trouble,changed,rules,throne,consequently,conclusion,observe,benefit,eternal,efforts,quarter,curious,absolute,base,mention,proceeded,destroyed,printed,absence,sufficiently,event,faithful,committed,addition,severe,guard,papers,joined,forced,services,consists,independent,extended,majority,rocks,distinction,address,remarks,garden,reasons,arrival,probable,branches,essential,extreme,affected,source,admit,issue,major,completely,decided,properly,error,plate,election,informed,sovereign,conscience,cities,minutes,central,reply,proceedings,assistance,happened,frequent,dangerous,performed,drawing,expect,occurred,palace,bible,farther,punishment,unknown,preserved,observation,relief,consent,difficulties,repeated,characters,nevertheless,destruction,expense,territory,bay,constantly,commerce,servants,exposed,preceding,evidently,exactly,slight,supreme,agreed,thinking,colony,extensive,credit,record,tone,energy,operations,pleasant,historical,provisions,quality,expedition,satisfied,level,mental,branch,intention,satisfaction,loved,successful,foundation,poetry,gained,observations,press,internal,actions,victory,compelled,quiet,legal,intelligence,establishment,region,fate,library,instrument,irish,minute,suffer,limited,circumstance,occasionally,doctor,lying,sudden,sentence,station,commission,favourable,surely,move,text,courage,pray,habits,whence,figures,payment,testimony,readily,resistance,literary,suffering,safety,disposition,families,recognised,impression,famous,furnished,possess,rapidly,clergy,effort,structure,belonging,sacrifice,created,introduction,acquired,dignity,apparently,silence,supported,try,parish,carrying,construction,list,commons,hitherto,weak,necessarily,news,chance,imagination,closed,commenced,addressed,inferior,occur,visited,yourself,fail,favourite,outside,contract,female,chamber,cried,intellectual,confined,control,resolution,boys,qualities,previously,tendency,advice,acquainted,forming,existing,thence,owing,dollars,promised,arts,choice,lieutenant,maintained,situated,dinner,disposed,parents,recent,devoted,moreover,safe,extremely,surrounded,copy,inquiry,solid,liable,approach,willing,absolutely,collection,please,familiar,holding,habit,anxious,permitted,individuals,dress,claims,violent,scripture,contain,accepted,slave,opposed,commanded,forest,striking,pointed,compared,senate,apparent,residence,praise,prisoners,correct,natives,writings,bar,immense,comfort,instruction,issued,midst,ministry,keeping,relative,commonly,est,collected,maintain,appeal,hearing,composition,european,visible,settlement,community,security,succession,singular,industry,assumed,colonies,possibly,failed,seized,advantages,honest,formation,consciousness,separated,final,symptoms,discovery,slaves,slavery,coloured,degrees,concerned,firm,gain,domestic,debt,rapid,subsequent,sign,apply,tender,tribes,institution,profession,elected,occasions,provinces,aside,fancy,princes,erected,avoid,including,limits,unable,pages,mission,fatal,knowing,charged,saved,remaining,buildings,accounts,colours,crime,ocean,interior,medical,accomplished,explained,acting,average,sympathy,delight,production,materials,turning,salvation,labours,communication,excited,consequences,glorious,convention,becoming,struggle,rivers,italian,judges,agent,passages,stay,accept,acted,vote,join,touch,features,requires,rare,parallel,improvement,regiment,solemn,terrible,institutions,prisoner,quarters,vice,movements,broke,window,reputation,supplied,range,zeal,noticed,sentiments,passes,scale,guide,train,exception,sitting,annual,presents,authorities,entrance,providence,enjoy,whenever,defendant,finished,attend,testament,independence,discharge,explain,image,request,merit,developed,executed,violence,imperial,straight,group,illustrations,ignorance,capacity,humble,infinite,doubtless,saviour,reality,commissioners,originally,statute,explanation,humanity,declaration,arranged,scientific,millions,ignorant,suit,journal,existed,legislature,permanent,soldier,attempted,remark,phenomena,characteristic,ecclesiastical,activity,doctrines,association,surprise,arc,decision,represent,suggested,secured,mixed,contents,oxford,centuries,silent,eminent,task,examples,calls,whereas,determine,discussion,restored,chapel,gas,induced,golden,guns,inclined,republic,creation,spite,agree,cry,calling,cattle,countenance,examined,proceeding,employment,abroad,details,effected,cover,conducted,cotton,voyage,conviction,assured,plaintiff,angle,imagine,additional,armed,agreeable,enabled,positive,destroy,completed,content,increasing,experiments,type,discipline,commander,pale,related,receiving,tube,respects,madame,depth,jurisdiction,occurs,remembered,gentle,area,affections,conception,confusion,element,creature,brief,commercial,signs,acquaintance,cavalry,authors,blow,exact,ruin,pictures,chair,directions,offices,reasonable,injury,volumes,reported,enjoyed,comparison,corner,diseases,scheme,departure,exists,remove,fashion,revenue,accustomed,stranger,retired,dutch,retreat,corresponding,mount,portions,brave,kindness,discover,plates,sentiment,official,remedy,tax,afraid,enable,apart,joint,establish,correspondence,instant,assembled,combined,sugar,delicate,protestant,dying,afternoon,appointment,search,mistake,concluded,scenes,uncle,brilliant,returning,attacked,methods,escaped,everywhere,prevented,districts,ie,experiment,recently,canal,ceased,endeavoured,endeavour,corps,remarked,creatures,atmosphere,session,admiration,prayers,fluid,diameter,exchange,sky,precious,records,objection,translation,caught,recorded,instructions,moving,continent,scriptures,teacher,career,companion,chinese,basis,perfection,contact,splendid,ambition,entering,utterly,owner,test,perform,ball,hospital,unity,sensible,depend,demanded,ease,pursued,obvious,esq,prosperity,specific,liberal,delay,ma,regions,fairly,confirmed,provision,climate,preserve,urged,surprised,considering,handsome,reward,exclaimed,notion,extend,unfortunate,intercourse,exhibited,pair,studies,farm,massachusetts,convinced,trace,representatives,continually,motives,multitude,aspect,yours,refer,proposition,attempts,examine,poem,experienced,smile,pressed,abandoned,pains,fault,agreement,princess,belongs,belonged,enterprise,arguments,quoted,counsel,review,perceive,adapted,deny,declare,painted,setting,conscious,profit,subsequently,rarely,band,narrative,included,suitable,metal,weakness,masses,justly,generation,condemned,management,consisting,conceive,translated,signed,conquest,ability,artillery,vision,involved,affair,calm,distinctly,plane,describe,nervous,furnish,cultivated,generous,wicked,attained,china,assigned,exercised,morrow,exceedingly,beings,rear,convenient,magnificent,reports,student,selected,universe,interested,failure,armies,prospect,commencement,painful,denied,preparation,petition,wherein,uniform,depends,assume,substances,sole,revelation,harmony,moderate,prevailed,pardon,neglect,tribe,mutual,instruments,primitive,male,column,nerve,admiral,discourse,outer,treat,liquid,performance,distress,determination,slightly,conceived,apt,ho,flat,manifest,windows,improved,estates,permit,cruel,reflection,representation,comparatively,cape,preached,functions,retained,divisions,virtues,errors,recommended,tea,expenses,muscles,legs,girls,students,publication,instantly,relating,excitement,reform,desert,representative,siege,wanting,charges,constructed,regards,provide,contempt,designed,combination,companions,behalf,senses,terror,constitutional,jury,using,misery,converted,entertained,perceived,gratitude,respective,plans,largely,particulars,plainly,savage,studied,supplies,dream,stomach,teachers,poems,reasoning,van,medicine,motive,miserable,approaching,garrison,rays,russian,preaching,monarch,marched,satisfactory,specimens,refuse,consisted,abundant,treasury,portrait,pronounced,pen,demands,poets,employ,properties,quantities,grief,criminal,bent,anxiety,prominent,removal,accident,companies,grain,assist,economy,balance,beloved,producing,consist,politics,faces,reformation,jewish,ruins,recovered,fond,associated,pour,excess,indicated,intelligent,enormous,chemical,uses,agents,loose,touched,claimed,melancholy,remainder,message,merchant,defend,confess,minor,charter,extending,bills,wrought,presently,compare,pity,piety,covenant,hero,strictly,tissue,precisely,rebellion,wages,campaign,engine,missionary,physician,infant,circulation,par,aged,seal,contrast,enjoyment,chain,gun,regret,patience,alliance,submit,aim,treatise,mixture,cultivation,remote,tradition,elevation,decree,sounds,dressed,intellect,sustained,displayed,useless,sphere,artist,undoubtedly,pursuit,surrender,pious,reception,machine,defeated,neglected,leader,sect,household,primary,dated,gravity,merchants,defeat,temporary,periods,taxes,vigour,culture,baptism,doubtful,pro,dispute,sources,invited,contest,lofty,lectures,villages,conclude,differences,communion,walking,produces,extremity,conference,represents,police,separation,trying,occupation,aforesaid,columns,educated,coat,alarm,simplicity,admirable,marshal,expressions,organised,proceeds,painting,desires,normal,rejected,chiefs,enthusiasm,conveyed,despair,cure,excuse,desirable,nerves,succeed,sensation,cease,faculties,mechanical,park,profound,spaniards,promises,suspicion,medium,definite,varied,resources,perpetual,meanwhile,apparatus,enlarged,obligation,genuine,protect,conflict,differ,languages,ideal,prepare,critical,estimate,arch,heir,revealed,distinguish,ascertained,axis,deprived,distribution,unhappy,meetings,occasioned,organic,route,membrane,utter,termed,corporation,analysis,curiosity,sufferings,noted,mistaken,travelling,illustration,approached,rude,surrounding,cutting,controversy,considerably,risk,angles,fee,touching,allies,deliver,citizen,constitute,artificial,antiquity,adds,infantry,occupy,channel,descent,gardens,bond,ranks,globe,detail,attachment,faculty,probability,retain,inquire,noise,possibility,sketch,merits,descended,venture,accused,restoration,honours,cathedral,authorised,dominion,esteem,designs,glance,theology,absent,specimen,congregation,leaders,submitted,persuaded,speaker,certainty,mines,rational,countrymen,host,flame,big,inflammation,suspended,recognise,investigation,strangers,mayor,resist,preparations,criticism,returns,strict,acute,imperfect,exhausted,fighting,varieties,powder,altered,lesson,departed,concealed,mounted,elevated,cabinet,mistress,approved,dreadful,lawful,navy,permission,morality,inspired,audience,editor,supper,purity,successive,mercury,commanding,copies,limit,manufacture,rival,dwelling,pleasures,renewed,admission,catholics,preparing,compound,extension,clause,obtaining,principally,delightful,processes,incident,irregular,obey,warrant,appropriate,anterior,travel,vigorous,folly,imposed,intervals,negative,procure,hotel,communicated,announced,orange,hut,adopt,hostile,concern,eloquence,discharged,respectable,velocity,intense,alleged,regarding,federal,affect,academy,ut,destined,pleasing,abundance,hi,tonnes,recover,check,continues,naval,reverence,moments,equivalent,executive,sections,confession,preferred,fundamental,signal,firmly,decide,imagined,governed,mortal,ample,dull,requisite,intimate,cousin,dozen,invention,procured,pupils,safely,circuit,debts,arrive,tomb,impressions,sincere,partial,justified,hist,inside,products,supposing,influences,impulse,totally,ascertain,nobles,formidable,attacks,martin,shores,quietly,fix,angry,parent,philosophical,notions,complaint,circular,bestowed,objections,vegetable,mystery,considerations,appendix,preach,ours,diminished,occurrence,counties,legislative,interpretation,baron,machinery,farmer,ceremony,display,invasion,site,elegant,legislation,americans,attributed,rode,fund,asserted,essay,healthy,plains,secondary,dependent,satisfy,diet,expressly,fragments,indicate,respectively,exclusively,crowded,interval,accurate,crimes,store,repeat,covering,favoured,civilised,occasional,cents,resemblance,conspicuous,scarce,prejudice,marry,republican,autumn,q,breadth,practically,reserved,based,flag,training,extract,societies,amidst,phrase,marquis,traced,map,absurd,poured,contributed,wooden,races,awful,successor,exhibit,judicial,summit,talking,prices,feature,regulations,reserve,indignation,saxon,duly,groups,exercises,debate,exceeding,museum,pretended,vicinity,border,adding,sciences,nobility,eager,constituted,desirous,representing,encouraged,problem,battery,gay,vanity,tract,purely,slightest,breakfast,governments,club,captured,exclusive,manifested,systems,capture,lover,commands,relieved,model,attorney,architecture,restore,disappeared,delighted,specially,preservation,pupil,summoned,distributed,improvements,uncertain,proofs,defined,nobody,conversion,preacher,mysterious,estimated,com,patent,income,excellence,accordance,chose,muscular,historian,stores,huge,illustrious,function,veins,monument,alas,forever,pursue,preface,engagement,magistrate,widely,product,heirs,describes,essentially,assurance,plenty,assertion,receives,posterior,dangers,unite,subjected,images,convey,refuge,happens,premises,suspected,agricultural,propose,induce,propriety,ac,resurrection,fearful,survey,mainly,magistrates,possesses,formal,desperate,attending,tend,responsible,perception,excite,abstract,unnecessary,creek,injured,inclination,universally,repose,protected,mostly,transferred,painter,bacon,duration,militia,vertical,dissolved,corruption,sorts,modes,uttered,superiority,unusual,egyptian,persecution,select,frank,obscure,enclosed,vital,insisted,solitary,pacific,urine,dominions,conquered,justify,assisted,lessons,voluntary,expectation,ancestors,extends,temporal,prize,productions,observing,project,measured,seriously,conferred,unfortunately,forests,arrest,declined,lecture,horizontal,deposited,crossing,gallant,welfare,dean,electricity,talked,fierce,traces,guards,cited,penalty,interview,comfortable,emotion,positions,definition,cannon,ambassador,recovery,steady,sex,splendour,grateful,pocket,shelter,tobacco,faint,deceased,crew,speedily,impressed,arrested,succeeding,franklin,prudence,interrupted,mischief,gross,agriculture,saving,advancing,disturbed,lastly,attracted,resting,jealousy,everlasting,treason,statutes,alluded,incapable,code,continuous,gulf,fortunate,coffee,shed,persian,fishing,promote,agency,submission,cavity,profits,inspiration,whatsoever,injustice,divinity,catch,documents,trustees,theological,proposal,discussed,funds,proclamation,aunt,shop,row,peculiarly,sublime,transactions,treasurer,void,equity,deity,severity,latitude,pastor,ultimate,exertions,charming,requested,alcohol,leg,exceed,sums,secondly,introduce,electric,revenge,grains,annually,realm,metals,horror,generals,oppose,navigation,monsieur,missionaries,roll,proportions,assert,regularly,inserted,possessions,apprehension,heated,decline,charm,entrusted,recommend,scotch,frontier,seasons,voices,funeral,league,competent,excessive,tenant,strain,omitted,votes,genus,commonwealth,indifferent,stages,endure,effective,prefer,manufactures,tenderness,motions,accuracy,affectionate,suited,consul,eternity,cruelty,inquired,voted,leisure,enters,vengeance,punished,shaped,defended,atlantic,jack,million,consistent,abbey,owners,esteemed,unjust,cord,repair,deserted,riches,relates,handed,prescribed,avoided,gracious,retire,instructed,abuse,trinity,protestants,quarrel,supposition,era,resolutions,index,acquire,gallery,beasts,enacted,coarse,amendment,attitude,managed,judged,bonds,disorder,dreams,immortal,deck,bodily,venerable,mineral,encouragement,ultimately,expedient,duchess,musical,quod,divide,assure,selection,generations,particles,surgeon,rage,varying,trials,visits,characteristics,maintenance,insects,rebels,indebted,creator,cylinder,brigade,dec,attain,dispatched,fortunes,suggestion,conclusions,deserve,poison,valleys,comparative,ascribed,engage,novel,monarchy,everybody,deposit,undertaking,prevail,inquiries,gently,collect,inscription,advised,arbitrary,diverse,morals,applicable,suggest,inform,sanction,stroke,superstition,attendance,labourers,contracted,cheerful,territories,professed,reduction,travelled,cum,interfere,inheritance,faults,essence,statue,effectually,patients,illness,couple,allied,invariably,unworthy,coin,hereditary,negro,mingled,fountain,vulgar,imprisonment,tyranny,crowned,temptation,create,discretion,landing,troubles,reduce,picturesque,happily,victim,imitation,relate,indies,commit,poetical,alteration,muscle,obligations,tired,virtuous,sept,wave,accomplish,perpendicular,costs,decay,hereby,destitute,fertile,foolish,adjoining,approbation,rocky,printing,fury,affecting,rejoice,misfortune,memoirs,qualified,ruined,attributes,margin,renders,sovereignty,consumption,fidelity,boundary,enlightened,parliamentary,recognition,repentance,bands,continuance,boiling,nice,resident,indispensable,magnitude,compensation,ray,indifference,ceremonies,changing,version,taxation,severely,pushed,portuguese,accompany,entry,regiments,adoption,provincial,reverse,successfully,disappointed,personally,daring,tent,amusement,lifted,dimensions,apartment,posterity,emotions,exceptions,discoveries,advocate,resumed,hypothesis,driving,excellency,monuments,magazine,preceded,ordained,outline,sacrifices,meantime,prey,lodge,depart,tales,resort,cries,rushed,vary,endeavours,realised,prudent,avail,grammar,delivery,include,swept,declaring,animated,fur,graceful,haste,turks,tribute,dance,beast,doubts,fortress,brick,innumerable,track,furniture,disgrace,assistant,settlements,invested,troubled,reflected,chambers,communicate,prophecy,caution,successors,settlers,solely,gradual,observes,reaction,ingenious,eminence,descendants,chase,declares,benevolence,effectual,intent,pressing,modified,layer,directors,decisive,mud,ratio,relieve,embrace,precise,assent,convent,recall,intensity,improve,partially,erect,accession,prejudices,seize,dispatch,embraced,inevitable,administered,exhibition,instinct,transfer,sentences,prose,heroic,isle,propositions,abandon,alarmed,integrity,exertion,hastened,resembling,classical,remembrance,govern,essays,infinitely,commissioner,doubted,intentions,indulgence,foregoing,abode,preference,legitimate,appearing,expressing,complained,grandeur,ventured,farewell,signifies,identical,episcopal,opportunities,excluded,lawyer,currency,allowing,perish,associations,dealing,carries,luxury,chronic,redemption,amiable,quit,benefits,horizon,deposits,rates,efficient,alter,judgments,strata,pause,trunk,equality,acceptance,liberties,implied,increases,hanging,descend,papal,injurious,rendering,dramatic,amounted,disappointment,assault,deceived,annals,sustain,exalted,astonishment,adequate,beneficial,ardent,farmers,travels,lateral,norman,gloomy,proclaimed,resided,loaded,recollection,revised,deserves,bottle,exquisite,hurried,productive,astonished,illustrate,rice,passengers,toil,cottage,transmitted,appoint,possessing,trifling,rapidity,founder,scots,sensations,pulpit,implies,subordinate,messenger,artists,proprietors,literally,females,concerns,incidents,admired,crisis,car,substantial,borders,compact,reflections,consequent,trained,invisible,democratic,romance,fraud,requiring,complaints,curve,register,fool,laboured,surfaces,ornaments,metallic,presenting,reproach,corrupt,uncommon,slender,continual,plea,invitation,accounted,viewed,excepting,stations,execute,rush,attraction,chapters,dismissed,magnetic,ordinance,nephew,municipal,mucous,jersey,access,headed,steadily,conceal,complain,evolution,currents,regulated,allowance,logic,remedies,resigned,confessed,influenced,document,presume,drama,traffic,zealous,marine,contemporary,withdrawn,unexpected,endowed,mentions,agitation,commence,negroes,usage,clergyman,selfish,theories,block,oppression,savages,distances,baptised,traditions,tragedy,substitute,consumed,shock,operate,analogy,perished,descriptions,innocence,newspaper,publicly,pretend,pursuits,chains,pitch,marvellous,inconsistent,landscape,guardian,exerted,rolled,elaborate,tubes,sulphur,accommodation,suspect,engines,recollect,memorial,privy,intermediate,surprising,supremacy,endeavouring,barbarous,rebel,peaceful,damage,vague,entertain,opponents,invented,foreigners,keeper,refers,leagues,petty,bulk,causing,compass,appetite,contribute,reflect,foundations,faithfully,inflicted,comprehend,victims,offensive,newspapers,phenomenon,lightning,maker,closing,tonne,contracts,ambitious,sinners,prevailing,pulse,protest,contented,cleared,formula,trusted,serving,african,lift,romantic,establishing,attempting,resulting,sultan,briefly,anticipated,rulers,enforced,murdered,lease,convenience,trembling,hastily,coach,appearances,hunter,yourselves,destiny,transaction,repeatedly,superficial,insurrection,consecrated,chemistry,utility,securing,expresses,eagerly,refusal,profitable,eve,ha,envy,width,portraits,inhabited,print,memorable,laughing,observer,jealous,pretensions,heartily,butler,limestone,adorned,adjacent,damages,tour,tracts,dissolution,attractive,dedicated,deserved,sincerity,manuscript,assumption,justification,prayed,dancing,deputy,consulted,precipitate,available,inspection,tones,benjamin,manual,extracts,characterised,suffice,defects,receipt,termination,patron,withdraw,physicians,constitutes,willingly,elementary,oriental,competition,disturbance,descending,visiting,grasp,removing,proprietor,encourage,consented,statesman,refined,vein,candidate,heroes,extremities,irritation,vices,battles,pointing,opera,visitors,grandfather,persuade,historians,flames,job,complex,maintaining,repaired,respected,derive,binding,monastery,sinner,varies,negotiations,shares,allegiance,rolling,consolation,celestial,warriors,tends,conceptions,attribute,basin,import,exposure,destructive,remarkably,filling,wealthy,imported,separately,censure,mail,floating,immortality,advances,horrible,mysteries,prosecution,palm,valour,supporting,ornament,eminently,speculation,conqueror,engineer,conspiracy,valued,intend,squadron,contrived,abilities,depths,youthful,sport,augustus,catalogue,loyal,zinc,ancients,realise,resignation,soda,referring,locality,regent,differs,defect,arches,ridiculous,furnace,exile,complicated,infancy,confirmation,presbyterian,preachers,examining,allusion,includes,regulation,continuing,guided,dates,gothic,revolt,uterus,epoch,annum,germans,assuming,sleeping,treating,warfare,confirm,resemble,match,breathing,certificate,ascended,colouring,lad,applying,barren,readiness,aided,khan,admire,senator,contemplation,turkey,variation,disagreeable,encounter,variations,summary,peninsula,instituted,veil,delicacy,recourse,verdict,prolonged,punish,strengthened,presumption,hearted,madam,tempted,granite,confident,charms,latest,applause,governors,hostility,release,conformity,humility,pace,depended,peculiarities,patronage,neutral,parted,mansion,recalled,guarded,tranquillity,applies,cave,sober,consult,facility,fatigue,maps,mining,enforce,spectacle,stating,diligence,annexed,creditors,types,terminated,nurse,popularity,smiling,demonstration,admits,sailors,oppressed,mortgage,surrendered,ruler,tenants,violation,martial,representations,accurately,finish,conjunction,schemes,breeze,describing,detected,acids,eloquent,expansion,combat,determining,logical,solitude,supernatural,rites,preliminary,gaze,disposal,strife,tissues,fortunately,dense,subdued,peril,compliment,legend,stout,destroying,offended,moves,boast,inference,fitting,bars,colonists,benevolent,artery,pledge,sensitive,estimation,obeyed,sterling,betrayed,feudal,joints,reformed,summons,beach,professors,plus,dependence,minded,transparent,platform,elect,deputies,confused,hesitation,modest,valve,dispersed,victorious,judicious,treaties,affirmed,sovereigns,porter,resembles,loyalty,incurred,mirror,contraction,adventures,revived,expended,homage,labouring,ruled,glands,customary,shared,indications,appeals,presumed,affirm,amusing,firing,wont,exposition,tribunal,positively,sensibility,conveyance,additions,contend,concentrated,li,restraint,festival,convince,mediterranean,scope,cabin,eventually,circumference,partner,condemnation,objected,entertainment,spinal,addresses,contemplated,adventure,insult,blows,manor,salary,develop,attendant,comprehensive,abolished,erroneous,prospects,fails,beauties,secular,amounts,balls,colleges,promoted,arriving,contradiction,denial,identity,providing,notices,dictionary,communications,originated,counted,transverse,slope,unit,unequal,vegetation,falsehood,statues,compel,odd,injuries,flour,yielding,imprisoned,fuel,firmness,roused,grants,vacant,inconvenience,rigid,attendants,vested,replaced,disputes,clever,polished,specified,agony,sincerely,tyrant,equation,orator,bag,depression,verb,hopeless,secretly,prussian,valid,morbid,hospitality,apology,reject,distinctions,exhibits,geography,gloom,ben,apartments,correspond,analogous,pillars,rural,conjecture,abundantly,prosperous,withdrew,tumour,insist,ascending,malice,mob,beautifully,heresy,solemnly,compromise,burial,successively,repetition,soluble,coins,friction,smiled,resentment,famine,contributions,refusing,hearty,anatomy,uniformly,significance,wickedness,incorporated,editions,practicable,suggestions,lens,appropriated,encountered,diocese,hint,moderation,subtle,miller,synod,exterior,peers,paint,combinations,expectations,bestow,indicates,payable,wilt,convert,forts,critics,ditto,advancement,confounded,insurance,stress,abolition,reminded,deliverance,collections,ardour,failing,sacrificed,edifice,values,accidental,industrial,moisture,fortified,courtesy,obviously,anywhere,multiplied,compounds,compassion,extravagant,courses,arabs,es,seeming,abdomen,statesmen,peasant,shops,evidences,boundaries,journals,manufacturing,manage,elections,banished,covers,diminish,chairman,suppressed,scales,resulted,geographical,imply,unfavourable,rounded,sketches,russians,signified,lodged,applications,parting,revenues,pas,coasts,threatening,apprehended,magic,prohibited,paragraph,impatient,vivid,resisted,apprehend,accompanying,hostilities,vegetables,imposing,diminution,expose,memoir,quote,materially,neat,decisions,issues,afflicted,assumes,released,pile,tended,pursuing,substituted,fiction,ambassadors,quitted,judging,theme,monthly,detachment,obedient,tremendous,noblest,cherished,designated,spectator,institute,structures,apostolic,acceptable,lonely,clement,eagle,warlike,resolve,rows,calculation,grecian,alps,confederate,regulate,occupies,approve,strengthen,hurry,experiences,alterations,mason,negotiation,advise,despised,retirement,plague,overthrow,rector,relics,revival,plunder,intelligible,patriotic,victor,predecessors,poetic,counter,pa,crying,pamphlet,uncertainty,stationed,defiance,dispose,assemblies,arid,establishments,gases,marching,constituents,hardy,ingenuity,slaughter,questioned,barons,progressive,salmon,engagements,topics,disputed,generosity,heretofore,shallow,bounded,serpent,metropolis,scarlet,maxim,constable,appreciate,preserving,japan,fees,aloud,especial,killing,jesuits,critic,procedure,commodities,tents,fiery,insufficient,expelled,imaginary,casting,favours,preventing,ensued,approaches,completion,ruling,situations,hesitate,athenians,connexions,lawyers,abuses,agitated,obstinate,penalties,associates,publish,infer,manhood,pattern,resorted,praying,exchequer,egyptians,potash,admirably,connecting,fashionable,henceforth,comedy,missions,stable,gigantic,guess,ascend,machines,admitting,europeans,erection,cunning,prompt,elapsed,weekly,picked,whispered,dislike,corrected,demonstrated,washing,engraved,awe,captivity,honesty,subsistence,cubic,habitual,converse,restrain,volunteers,scanty,accidents,obstacles,pencil,deficient,gilt,concert,deficiency,profess,checked,layers,orthodox,garments,communities,planet,misfortunes,sigh,cash,custody,forthwith,occurring,dividing,pecuniary,data,aristocracy,energies,advantageous,faction,promptly,collecting,lovers,similarly,gratification,elastic,solar,joining,rolls,pronounce,blocks,footing,baggage,condemn,stately,inferred,acquisition,superintendent,maritime,reside,ordinances,decrees,embarked,promising,candidates,defective,hasty,tune,precepts,limitation,blank,confinement,confine,sealed,crushed,advocates,wasted,whig,calamity,visitor,card,doctors,transport,hazard,technical,treachery,fancied,radius,endured,styled,pastoral,badly,mathematical,enjoying,considers,boiled,symbol,item,necessities,dynasty,signify,irresistible,disturb,closer,detained,impress,cards,bowels,opium,relatives,universities,passive,prerogative,dragged,indulge,acknowledgment,performing,lesser,combine,reigned,rail,problems,juice,convicted,improper,exert,addressing,pagan,cured,publications,tending,engravings,canvas,hers,temperance,ammunition,perseverance,exclude,delegates,boiler,tolerably,warrior,bat,diagnosis,contemporaries,precision,exchanged,harsh,enthusiastic,cargo,correctly,mexican,industrious,gardener,energetic,canals,localities,respiration,numbered,tedious,dignified,plead,occupying,directing,humbly,suits,captains,furious,costly,ridicule,horseback,phrases,density,obscurity,miraculous,lustre,deliberate,appealed,repeal,prevalent,vicar,extinct,officials,speculations,warren,gentiles,unnatural,sculpture,flew,backward,reverend,emperors,anybody,multitudes,breathe,indulged,peasants,modifications,involve,passionate,debtor,persuasion,opposing,copied,triangle,foliage,screw,peculiarity,struggling,utterance,charitable,reasonably,aggregate,jean,labourer,skull,restrained,dispositions,projects,radical,urge,mathematics,madness,disappear,giant,protector,furnishes,sway,inscriptions,counsels,trading,drunk,exceeded,wrapped,accumulated,stayed,organism,elegance,instructive,unpleasant,granting,creditor,approval,inter,improbable,asylum,idolatry,dirty,ounces,speedy,discontent,expensive,transition,contended,significant,occupations,ascent,exaggerated,democracy,rely,reconciled,walker,sunset,authentic,discourses,rescue,wagon,suspension,incredible,atonement,merchandise,gazette,historic,dressing,abound,squire,revolutions,nobleman,uniformity,studying,captive,confederacy,damp,engineers,equilibrium,singularly,gaining,dollar,planets,confer,associate,folio,genera,laborious,conducting,magnificence,complexion,choir,discuss,rotation,deceive,objective,argued,prevails,expired,channels,amounting,farms,imitate,beating,infants,affliction,liberality,vine,promotion,avenue,devote,usefulness,manifestation,francs,finely,clearness,educational,qualifications,delayed,remind,drawings,whereupon,disc,systematic,independently,involves,indication,obstacle,projected,ally,absorption,baby,correction,projecting,disgust,inscribed,modesty,cheer,tastes,tempest,victories,perceiving,cultivate,comparing,penetrated,superstitious,decomposition,modification,attract,lighter,slate,cordial,adverse,guides,signification,venus,extinguished,daylight,chronicle,torture,deprive,boots,introducing,distinguishing,graduated,hints,director,tendencies,converts,flourishing,spectators,bargain,reconciliation,insanity,atoms,ai,rewarded,massacre,traversed,defendants,chancery,manufactured,ceases,troublesome,gratified,admiralty,levied,dispensation,evangelical,concluding,barbarians,sanctuary,sessions,laden,plantations,japanese,mi,expressive,forcibly,devout,swiss,paintings,adult,anguish,whigs,palaces,copious,conquests,elector,suggests,amused,pick,chimney,insect,tombs,spirited,observance,fragment,dwellings,triumphant,pomp,requirements,reducing,conclusive,conductor,avowed,goddess,desolate,redeemer,spaces,dose,potatoes,efficiency,subscription,alternative,overlooked,displeasure,recommendation,aimed,surplus,convictions,comforts,ordinarily,tropical,electors,exclusion,mistakes,tension,investigations,doses,survived,arteries,commencing,temptations,deserving,correspondent,fay,minority,odious,extensively,charcoal,beef,arrange,accomplishment,fairy,susceptible,located,factory,slip,permanently,achieved,wagons,astonishing,supposes,sympathies,traders,unreasonable,constitutions,geology,gazed,payments,sects,asiatic,penetrate,identified,expecting,accumulation,literal,impatience,compositions,figs,transported,sympathetic,petitions,bleeding,struggles,merciful,actor,grandson,manager,transportation,appreciated,subjection,indefinite,frightful,diffused,methodist,dined,vicious,gravel,disastrous,besieged,praises,magnet,adherents,piston,temperate,astronomy,drowned,assuredly,howe,solutions,strangely,novelty,commissioned,domain,cane,reformers,attach,chances,secrets,disguise,males,retaining,sounded,meridian,cloak,papa,restricted,pack,negligence,rewards,advisable,variable,invaded,denote,violently,vanished,sandstone,topic,architect,joys,thermometer,enables,emancipation,onward,absurdity,testator,electrical,accusation,actors,parishes,divorce,pension,sexes,induction,reconcile,curved,spontaneous,visions,interruption,buffalo,allows,apprehensions,alternately,push,fence,destination,log,contribution,biography,mechanics,notorious,veneration,asserts,insignificant,anecdote,resume,hind,gratify,aroused,perpetually,nous,link,comment,collector,hamlet,unsuccessful,finer,temperament,pump,cautious,jewels,typical,attentive,secretion,recollections,connect,physiology,regretted,canoe,encamped,unfit,disorders,reveal,massive,diary,indictment,painters,physiological,tutor,impartial,murmur,supplying,conquer,despise,mortality,voluntarily,discern,persians,corporations,sweep,plunged,polite,violet,unlimited,expiration,prohibition,extracted,discussions,informs,natures,corresponds,settling,proposals,motor,jaw,speculative,impose,breeding,contemplate,bounty,corpse,senior,blown,pretext,employing,disaster,senators,touches,formally,efficacy,intimately,turner,coats,renew,exhibiting,reigns,corpus,fun,arab,presidency,cuts,athenian,personality,improving,stained,carpenter,confederation,precedent,luck,impossibility,validity,quarterly,marriages,gale,dorsal,repent,ahead,compose,instruct,shipping,delicious,differing,promoting,comrades,lapse,backwards,dame,horrors,searching,barrier,elephant,manuscripts,governing,ar,divines,rushing,indirectly,adversary,brush,suppression,trustee,illegal,resign,compliance,battalion,histories,cheerfully,cork,trip,nominated,moist,unquestionably,isles,timid,residue,array,heretics,fashioned,depending,professions,embracing,simplest,pleaded,patriot,inherent,corners,dreamed,commanders,memories,polish,commentary,pernicious,cooper,procuring,antiquities,differently,iniquity,distrust,urgent,inherited,extant,fruitful,majestic,shaking,fixing,manure,signature,calamities,villa,lodging,encouraging,tight,rugged,omit,seminary,hasten,denounced,explains,appreciation,particle,freight,monstrous,notable,committees,reflecting,pairs,solemnity,rash,denying,haughty,sagacity,degraded,indignant,mate,drift,plantation,manifestly,comply,paused,commences,intervention,enumerated,basket,sailor,tariff,ranges,saxons,issuing,perilous,praised,diseased,monster,exempt,seconds,punishments,desiring,injure,crimson,slopes,interpreted,suspicions,reversed,ounce,prodigious,insensible,selfishness,depressed,orbit,invite,enlargement,inflamed,redeemed,grade,robes,references,adversaries,dissolve,germ,challenge,smiles,executors,presentation,peoples,spacious,delusion,fortifications,aims,cone,mechanism,superfluous,file,malignant,bondage,stuff,precaution,stupid,nomination,lists,terminate,julian,desolation,intrigues,prelates,regularity,pouring,diversity,relied,continuation,pond,dine,obstruction,patriarch,pig,habitation,legends,felicity,skies,periodical,banner,inspire,nominal,vile,impressive,tumult,troop,impracticable,tory,champion,cough,maxims,penal,concurrence,lever,stripped,tolerable,movable,somebody,israelites,enmity,models,receiver,affinity,retiring,rupture,communicating,silently,amusements,measuring,ink,provoked,delights,novels,bravery,imports,os,unanimously,incessant,harmless,hospitals,assailed,restrictions,degradation,miscellaneous,partnership,digestion,supports,prevents,trades,entertaining,intolerable,indirect,trains,distressed,toleration,spoil,embassy,barrel,needless,gains,monopoly,diamond,ensuing,tenure,opponent,refinement,calmly,populace,exports,facilities,dispatches,flung,flourished,assign,faced,imperfectly,mature,alarming,precautions,enjoined,betray,prophetic,wreck,mercantile,outlines,suspicious,libraries,coke,uneasiness,abused,peer,profane,forwards,partake,yonder,telescope,omission,fugitive,appellation,comprehended,conquerors,symbols,contending,tenor,weakened,alternate,pierced,finite,inquisition,antique,accepting,eagerness,surpassed,luminous,implements,oval,arrives,partition,rejoicing,gifted,screen,assemble,rob,printer,decent,simultaneously,uneasy,ornamental,arabic,aversion,microscope,despite,operated,inevitably,pillar,dug,laud,chairs,grievances,anecdotes,transformed,dick,weaker,confederates,registered,violated,mix,ethics,testify,shout,verbal,horrid,satire,equator,tip,limitations,necessaries,persecuted,uniting,bases,treats,classic,rejoiced,fossil,pleases,factor,edict,hesitated,chorus,vale,nonsense,perceptible,dining,guarantee,robe,expeditions,worthless,favourably,repeating,las,clearing,robbed,viceroy,indicating,combustion,injunction,nave,metaphysical,argue,impulses,ministerial,gazing,affects,claiming,impart,diagram,involving,distinctive,cars,predecessor,aspects,debates,pasture,strip,suppress,comprising,renowned,voyages,publishing,inability,declining,disturbances,torrent,acquiring,crest,coffin,device,controlled,maturity,administer,ca,echo,bearer,creating,dashed,candour,perusal,equitable,minerals,protracted,prophecies,colleagues,trumpet,messengers,strive,patches,dialogue,nobler,salutary,sage,ordination,excepted,compressed,popery,repealed,unanimous,truce,oblique,springing,eruption,augmented,zone,substantially,construct,crystalline,symptom,skeleton,intestine,oppressive,reliance,stimulus,hosts,virtually,softened,explaining,cholera,experimental,brandy,denomination,purest,receipts,spell,genial,jordan,define,precipitated,intimacy,embodied,powerfully,brute,drum,canadian,personage,idol,fabric,deposition,yellowish,caroline,nucleus,accessible,happier,evinced,somehow,diligent,hindus,inadequate,influential,consistency,concession,longitudinal,resolute,manifestations,gauge,discovering,adjusted,traitor,detect,flattering,agreeably,canton,convulsions,parson,conservative,escort,retreated,excursion,agrees,engraving,scandal,differed,halt,tests,verily,concrete,explanations,fits,contrasted,canoes,presided,proportional,resembled,imparted,decreed,relatively,unbroken,fluids,satisfactorily,huts,commandments,seals,lamented,adhered,honestly,commodore,furnishing,declarations,concludes,posture,fable,twilight,venetian,alkaline,enlarge,urging,defending,allotted,seizing,driver,comprised,stature,attacking,boundless,texture,outset,deliberately,exercising,units,enduring,vomiting,fancies,fertility,forfeited,protecting,enclosure,chivalry,dome,disadvantage,herald,territorial,mosaic,velvet,pledged,economic,shouted,adopting,workers,inquiring,default,medicines,expulsion,italians,paternal,areas,tavern,texts,anxiously,frigate,unfrequently,sanctioned,mischievous,corrupted,miseries,interpreter,squares,qualification,melody,unlawful,soils,prompted,politicians,humiliation,halted,prelate,fol,translations,cement,assigns,respectfully,privately,avarice,concord,oblige,divides,lodgings,awaited,pilot,grievous,constituting,scriptural,instincts,royalty,frequency,panic,overwhelming,aperture,ethical,niece,voting,infallible,repairs,intimated,levy,beforehand,pavement,minimum,ashore,reluctance,curtain,valves,preferable,patriots,survive,horsemen,vigilance,glimpse,renewal,forcing,overwhelmed,sweeping,monasteries,employments,infamous,regulating,partisans,nova,regeneration,dissenters,convex,battalions,gown,separating,apex,swedish,pistol,jest,successes,amply,imputed,glittering,footsteps,ross,intending,masonry,inventions,momentary,subscribed,resource,inconsiderable,compliments,garment,captives,cortes,marches,uterine,allude,renounce,amended,dresses,mortification,split,calculations,angular,moneys,economical,fracture,lifetime,tested,precept,defensive,escaping,terrors,explosion,vow,seemingly,republicans,holder,pence,harmonious,curves,ivory,flash,arctic,tumours,fibrous,adjustment,adhere,extravagance,tremble,behaved,estimates,enriched,condensed,idols,response,constituent,riot,adduced,impaired,conveying,appropriation,complied,mademoiselle,orderly,hardships,provides,variance,restoring,pleading,personages,clergymen,expanded,genesis,offender,correctness,trembled,shouts,arsenic,constrained,insoluble,preparatory,caste,gland,fountains,pilgrimage,location,decrease,supra,sire,citadel,patiently,coronation,disregard,fir,conscientious,denominated,pus,forcible,appointments,superseded,hardened,vibrations,passenger,retains,subjective,stability,flora,confidently,hideous,reddish,quotation,solicitor,commissions,clerical,arabian,pathetic,clauses,missing,cavities,census,serene,presbytery,brazil,crust,headquarters,focus,frontiers,amuse,emigration,surgery,bricks,emphasis,refuses,abbe,pearl,fools,attainment,abruptly,anticipate,facing,lifting,evaporation,surround,quaint,desk,vent,pontiff,adieu,burr,puritans,sheltered,potential,engaging,gum,makers,winning,philosophic,examinations,sanguine,confidential,interfered,habitually,proclaim,consuls,rivals,remorse,conform,invalid,importation,rhetoric,commandment,enlisted,collision,congregations,latent,escapes,glories,hunters,legally,conceded,jar,embraces,candid,awkward,remnant,extremes,deliberation,circulated,commend,starch,dialect,calendar,polar,sanctity,plentiful,similarity,peak,consecration,lace,tranquil,insure,malady,plundered,repulsed,trick,separates,pet,concessions,deduced,composing,mythology,classed,popish,clubs,outrage,doubled,fruitless,await,exported,hemisphere,exceeds,vault,export,perils,protested,filed,surmounted,theoretical,trunks,persisted,planes,consciences,vines,pulmonary,decease,observers,unusually,counting,supporters,tories,interposed,pitched,sounding,undue,catastrophe,courtiers,ascribe,galleries,doings,imitated,hindu,resisting,meditation,criminals,prairie,avoiding,descends,sunny,prizes,comprehension,dismal,combining,illuminated,excused,jaws,specie,pink,puritan,deserts,guardians,pierce,dictated,usages,replies,concur,papists,muse,fortitude,gout,earthquake,nourishment,quarrels,discord,ascertaining,deference,securities,scarcity,gratifying,murderer,striving,longitude,dealings,performances,suffrage,familiarity,traits,ordering,adventurers,shah,projection,maintains,whip,sports,respectful,populous,availed,ceiling,revive,conveniently,arithmetic,vainly,partners,terrace,incumbent,magnetism,intervening,mutually,compression,joyous,demonstrate,dragon,joins,charters,denotes,accidentally,politic,asserting,planned,stealing,dam,eyed,recovering,anarchy,trivial,coil,enterprises,suffers,ancestor,surroundings,prosecuted,drainage,romish,refrain,contingent,cursed,corpuscles,committing,couch,veteran,troy,taxed,terrified,mound,contradictory,joyful,pleas,altitude,trusting,corporate,undisturbed,vitality,extinction,diminishing,blush,rails,introductory,proposes,terrestrial,executor,devices,undoubted,memorials,contrivance,amendments,elasticity,fervent,fantastic,penetrating,clearer,waving,gentry,curiously,transmission,catching,inclinations,botany,insert,celebration,spectrum,stain,blended,viscount,banquet,scruple,incomplete,renown,revision,styles,trifle,initial,remonstrance,syllable,infidelity,adaptation,insured,alien,cream,solicitude,colleague,lament,emergency,infirmities,generated,exhaustion,inspector,silly,mournful,embarrassment,abounds,exploits,effecting,attractions,delivering,motto,countless,cylindrical,chamberlain,tiny,chariot,rejoined,monastic,reforms,oratory,ind,unique,pasha,rescued,chevalier,pike,rejection,anticipation,archives,remembering,quest,surgeons,figured,crush,grapes,monarchs,sexual,occurrences,discussing,pyramid,grape,tribune,surviving,assistants,serviceable,brigadier,invaders,lieu,cod,barrels,obnoxious,commended,bottles,cordially,facilitate,cemetery,festivals,publishers,anyone,gorgeous,antecedent,legions,vocal,poisonous,plausible,legacy,cancer,pockets,geometry,dedication,stimulated,succeeds,treatises,stored,vigorously,partiality,constancy,longed,incline,handkerchief,demanding,peasantry,organisms,background,duchy,capitals,bankruptcy,deities,vehicle,plaintiffs,strains,strove,membranes,profusion,fishermen,faded,intensely,pushing,libel,mutiny,hiding,tout,flatter,flourish,amend,trusts,stationary,cherish,endurance,disinterested,equations,coral,bushels,winged,annoyance,skilled,spoils,courteous,formations,treacherous,construed,quotations,lien,ruinous,farming,piles,salaries,parcel,perplexed,announce,embryo,forefathers,entreated,potent,jurisprudence,dishonour,diligently,relish,discouraged,liquids,twisted,protestantism,managers,jesuit,frankly,effusion,distilled,tragic,inflict,recompense,achievements,hooker,comic,elephants,brutal,collar,inseparable,adjourned,maybe,conversations,sizes,fines,closet,aggravated,visitation,insolence,finance,vows,succour,palmer,responded,cable,parade,staircase,prints,bags,dig,diamonds,aristocratic,thanksgiving,servitude,displays,wondering,joke,participation,fugitives,boil,vs,disobedience,expediency,enjoyments,affirmative,shattered,physics,grieved,adoration,climates,obstinacy,compiled,dilute,assessment,propagation,secrecy,contention,volunteer,presbyterians,instrumental,tenderly,vanquished,determines,pint,cheerfulness,foreigner,rubber,pursuance,gunpowder,oral,individuality,rubbed,rectum,intrigue,imminent,rite,gaiety,intrinsic,wooded,capacities,conical,peaceable,metropolitan,unhappily,guise,excesses,redeem,pregnancy,inhabitant,denies,devoid,deposed,incessantly,empowered,rooted,sine,diffusion,pirates,presiding,gentleness,embarrassed,awaiting,slipped,politeness,despotic,allusions,offend,wondrous,reigning,factories,casts,disgraceful,authorise,reinforcements,arduous,helping,controversies,bankrupt,fervour,mountainous,whosoever,gesture,vascular,fog,spine,consultation,tit,perceptions,epic,operative,emblem,superstitions,transformation,inhabit,dictates,psychology,externally,zero,paces,offenders,grandmother,kidney,hospitable,neutrality,dash,boasted,decoration,pox,antagonist,conspirators,publisher,corporal,pelvis,calcareous,dominant,valiant,gallantry,envoy,blockade,incompatible,revelations,mentioning,cessation,aspirations,proportioned,bail,frequented,infection,derives,counsellor,flattered,tyrants,banker,waist,coincidence,fitness,equipped,investigate,adorn,friars,factions,morally,crude,mask,junior,counts,traditional,segments,radiant,agencies,noun,pleasantly,inviting,refreshing,retina,flattery,colourless,oration,transient,subsist,concave,qua,possessor,obtains,disappears,observable,ballot,rumour,legion,founders,persistent,bravely,entreat,mystic,acquitted,transmit,verge,dirt,anil,insulted,variously,tracing,forfeiture,lance,assisting,mall,parable,watered,drying,celtic,exaggeration,vexed,ratified,directs,contemptible,actuated,balanced,fraternity,politician,edifices,exemption,enabling,assignment,inmates,pastors,explicit,burgesses,interpret,magnesia,privileged,malicious,subdue,ravages,astronomical,poisoned,assertions,plural,stipulated,premature,caustic,charging,schoolmaster,attainments,proverb,sufferer,ranged,incidental,saturated,rhyme,destroys,reminds,award,restriction,rebellious,congenial,drag,holders,navigable,banishment,vindication,disturbing,contra,equals,spiral,summits,forbearance,rum,journeys,nobly,tempered,bundle,conferences,involuntary,bile,auxiliary,convincing,recollected,ownership,oblong,illustrative,quakers,ensure,disclosed,rouse,factors,anniversary,ticket,ravine,admirers,britons,eccentric,charmed,insolent,suggestive,demeanour,conflicting,dragoons,liturgy,abscess,professing,unexpectedly,substitution,secession,terra,lingering,scotia,oblivion,replace,advocated,traverse,sensual,parliaments,penance,peaks,impunity,fondness,pardoned,atom,dissatisfied,impelled,outlet,staying,suicide,flags,outbreak,dispense,bearings,distracted,defenders,hypocrisy,reckoning,happiest,sentenced,disregarded,bud,chord,reproduction,exposing,lessen,counsellors,rebuke,prosecute,gallons,attaining,reformer,tabernacle,ceremonial,mortals,memorandum,intestinal,lucky,reputed,beneficent,evolved,epidemic,noisy,fletcher,supervision,reflex,orifice,nutrition,enterprising,descendant,imputation,solicited,invaluable,phosphorus,kidneys,workmanship,certificates,superiors,porcelain,salute,scruples,assurances,capitol,temperatures,accomplishments,distressing,struggled,ludicrous,concealment,oracle,prostrate,resides,cheers,severally,denominations,creative,benefactor,peopled,verified,cartilage,packet,fortresses,circulating,beggar,feminine,demon,iris,attested,nursery,irritated,suspend,displeased,pearls,soever,unbounded,octavo,alphabet,inconvenient,executing,finances,motionless,conventional,spherical,accent,illustrating,impending,jay,seizure,buds,consternation,tartar,loch,lawfully,follies,illusion,accuse,luxurious,assemblage,affirms,terminal,darkened,like,glen,conceivable,case,once,war,cooking,perhaps,already,eclipse,cases,confided,immersed,poor,incense,fossils,schism,transit,paramount,optic,orphan,commodity,believe,indefatigable,superb,industries,glanced,instinctive,fays,seems,lining,reprinted,carpet,trespass,delirium,air,controlling,sided,predicted,glorified,cautiously,hinted,contracting,religion,designation,miners,purified,fermentation,prominence,prefixed,puzzled,cake,lessened,graciously,inconveniences,burnet,woodcuts,drug,reliable,vie,hopeful,seemed,pathological,incision,imperative,get,syphilis,perfected,neither,pigs,repairing,died,spy,abstain,reveals,honour,deputation,cottages,swamp,amazement,conceit,deviation,encampment,ritual,want,revolving,turbulent,unsatisfactory,spinning,prone,residents,expert,tossed,drunkenness,bias,enclose,discrimination,poisoning,disguised,vagina,brotherhood,decayed,justice,fatty,triangular,directory,seem,designate,crushing,unchanged,got,pay,recognising,paid,dismay,maternal,easily,clamour,abandonment,incarnation,excursions,gap,distinguishes,haunted,happy,grasped,ill,convocation,surveyor,disgusted,appointing,caves,indolence,laity,unavoidable,somewhat,volatile,caprice,weighty,deluge,perspective,attentions,vacancy,built,spoiled,patrons,remission,sequel,celebrate,impeachment,forfeit,verbs,easy,locomotive,deception,carrier,supplement,degeneration,palms,franchise,ablest,strokes,spleen,dispensed,quotes,interposition,building,transports,april,sustaining,happiness,shocked,tricks,sensibly,likely,theatrical,abrupt,favourites,connective,scaffold,marking,invariable,believed,spot,cellar,casual,bullet,ineffectual,flowers,presses,abstraction,demons,primarily,palpable,shaded,precarious,seat,superintendence,sup,vehement,tincture,deplorable,recited,refraction,admiring,dismiss,custom,averse,skin,intestines,rising,running,raise,torment,moderately,wanted,electoral,tasted,twice,getting,ingredients,recess,pending,obedience,chapels,modify,cells,raging,prison,purchase,plunge,affectionately,wait,illumination,arched,congestion,sanctified,acquaintances,wants,cheering,medal,spontaneously,incur,infidel,finishing,root,carriage,waiting,testified,started,unites,discharging,minutely,customs,innocent,obliging,waste,flower,burton,discharges,cardinal,schedule,numberless,pestilence,nasal,perplexity,commandant,seventy,rumours,imposition,trail,demonstrations,ane,habitations,conversed,dissatisfaction,editors,strikingly,rupees,devotion,prophet,component,luxuriant,metaphysics,wars,summon,overthrown,happen,ingratitude,magazines,contradict,doubly,seated,affectation,ragged,drill,dilatation,backed,curate,therefrom,super,roots,withered,sculptor,calculate,princely,reciprocal,conflicts,grades,airy,pamphlets,gastric,mounting,refreshment,purchased,invade,marble,reservoir,penitent,orators,persecutions,repel,carnal,arranging,raising,uttering,awarded,magnified,collateral,employer,renounced,ligament,albumen,arrayed,ordnance,tenets,build,cell,relaxation,recesses,propagated,paying,rustic,subsided,suburbs,rod,friar,bull,mock,heroism,nationality,ode,anger,wretched,requesting,virgin,aqueous,uninterrupted,watchful,liberally,shrewd,investment,feast,invincible,prophets,bologna,searched,precipice,informing,inverted,median,realities,detection,sermons,murmured,whereby,diverted,celebrity,assyrian,waited,enjoys,papacy,contiguous,dock,baronet,inventor,contests,triple,lights,unimportant,tributary,deduction,affidavit,lantern,fables,shouting,marrying,rheumatism,provisional,miniature,suspense,believing,musket,dilated,wilderness,unlike,patch,concentration,mixing,haunts,vibration,ensue,relaxed,spheres,illustrates,infirmity,theologians,definitions,frail,porch,quitting,profoundly,consume,purer,endowments,blame,sayings,depravity,tenth,piercing,drake,premier,clothing,vastly,barracks,cornea,chronicles,coinage,scandalous,lawn,franks,unjustly,chaos,spots,employs,spectacles,confound,pregnant,servile,perverted,button,hurt,commune,sample,adherence,actively,peas,assailants,acknowledging,dexterity,anatomical,complaining,inconsistency,bays,advertisement,durable,opaque,undertake,draught,procession,forum,diminishes,disabled,impure,infected,insertion,naturalist,vindicate,lamp,recording,reproaches,campaigns,fragrant,laboratory,ballad,investigated,unacquainted,volition,chestnut,seats,pretending,capsule,czar,lutheran,enclosing,remonstrances,rid,palatine,frivolous,decency,maid,rosy,paste,diarrhoea,whereof,computed,undertaken,intricate,pyramids,ungrateful,erecting,adultery,disasters,flattened,christmas,terminates,neatly,pie,disadvantages,definitely,repay,talks,painfully,rigorous,messages,liked,starting,dies,disdain,wanton,translator,digging,vassals,temporarily,incapacity,vivacity,appetites,discontented,fraudulent,stamp,terrific,jumped,assassination,clan,liberated,slumber,catechism,tertiary,accommodate,parochial,justices,suite,ladyship,proximity,banners,ascension,solve,abnormal,rajah,sally,pronunciation,prism,dissensions,scrupulous,auspices,permits,rivalry,denoting,urethra,elongated,subscribe,lingered,pathology,paved,mingle,cylinders,mortgagee,theirs,explored,dawn,sans,annoyed,assented,crooked,provoke,epithet,goal,wherefore,piled,gets,strained,continuity,usurpation,sunrise,graphic,awakening,prediction,narratives,baptist,amazing,diversion,postponed,vocation,inst,drunken,enveloped,sergeant,savings,divert,segment,barbarism,tempt,bass,furnaces,noblemen,momentous,loan,regency,fisheries,felony,filter,undertook,equalled,pathos,dogma,voters,tries,spreading,solved,deferred,displaced,warden,purposely,carriages,toils,deportment,rushes,restitution,obsolete,larynx,sacrament,believes,rebuilt,robin,irrigation,royalists,intuition,dickens,employers,foretold,prosper,congregational,mediation,excelled,consummate,stockings,pneumonia,convict,luxuries,recommending,unsettled,regardless,excessively,impious,scorn,powerless,vicissitudes,internally,devised,uncomfortable,mute,acquires,alludes,founding,acquaint,exemplified,fictitious,vinegar,purchaser,sufferers,amazed,subscribers,keepers,tribunals,ferocious,classics,intercession,somewhere,profuse,infernal,bankers,easier,conductors,multiplication,crusade,digestive,adjutant,masterly,lecturer,pot,sufficiency,believers,tolerated,strengthening,skins,olive,fellowship,atlas,bard,tracks,vatican,faintly,terminating,originate,prop,administrator,ay,rods,practitioner,commentaries,excites,predominant,traitors,sanguinary,signifying,sullen,conversing,antimony,reproduced,trader,downwards,assuring,apparel,permitting,benefited,bewildered,pays,paralysis,religions,wholesome,announcing,individually,dire,hail,garrisons,turk,tangent,thoughtful,stamped,confines,soup,affixed,herbs,avert,polity,flask,imaginative,torrents,planters,republics,coincide,reptiles,downward,recommends,sturdy,stove,bachelor,fro,items,sighed,innate,comfortably,bloom,gravitation,cakes,spelling,fanciful,danced,restraints,curtains,prevention,muttered,warranted,sue,powdered,lamps,dreaming,inherit,irregularity,excluding,rubbish,murderers,discerned,unseen,pilgrims,insight,devise,ceded,frenchman,censured,logs,ferry,infidels,exportation,enumeration,juncture,clasped,mediator,cruelly,singly,reminiscences,unfinished,seconded,desertion,barometer,bordering,testimonies,gentile,airs,symmetry,proposing,generality,stripes,tails,mastery,pendulum,strait,duel,pastures,digest,flushed,devoured,vexation,impetuous,straits,comprises,inaccessible,stratum,unlikely,complains,leases,mutton,fraction,persevering,sierra,attends,extraction,exemplary,freezing,creates,conjectures,basil,believer,abhorrence,vanish,wasting,recognises,bonnet,rejecting,kindled,relinquish,stormy,intemperance,thrice,courier,prescribe,lowered,checks,boarding,boilers,adversity,embroidered,sacraments,loans,civility,withal,congratulate,jet,tub,garb,administrators,prisons,preserves,regal,diffuse,foresight,filial,kirk,reasonings,afflictions,fishery,camps,displacement,pilgrim,robbery,flashed,palate,entreaties,meagre,sequence,dove,comments,bullion,abounding,converting,lowly,sparkling,shirt,subservient,diluted,premium,raises,abolish,deceit,enraged,abiding,routed,fusion,consummation,testing,merited,reluctantly,vendor,marvel,dullness,relieving,retreating,propensity,assessed,neglecting,ably,robust,completing,registration,barbarian,pots,wherewith,contemplating,ballads,quaker,ransom,deciding,bait,anonymous,germs,bust,poorer,selecting,expand,cardinals,punch,plumage,unaffected,customers,departing,fascinating,sympathise,membership,beware,speakers,exclamation,decorations,handwriting,abominable,frenchmen,fiercely,accompanies,meek,animosity,stimulating,parental,fondly,resultant,withheld,mystical,fallacy,embark,compulsory,confessions,lumber,eligible,homely,tact,detain,baffled,gang,lazy,lesions,status,likes,hermit,conquering,excel,hailed,humorous,cheered,spaniard,ribbon,sued,comet,spreads,subscriptions,violate,sized,galleys,chart,purchasing,spies,endowment,chronology,scent,secretaries,dower,deformity,echoed,mechanic,blamed,recreation,surveyed,repugnant,enrolled,exhortation,mosque,jacket,pier,insomuch,valuation,achievement,consistently,sheer,rubbing,fearless,commendation,adults,mustered,wee,trifles,exactness,corporeal,teutonic,unintelligible,feasts,geometrical,sighted,vases,highlanders,licensed,comforted,lawless,abyss,weaken,invest,excommunication,dusky,dictate,piers,credulity,composure,physic,nuisance,bundles,regain,sublimity,perfections,dagger,solidity,wholesale,filaments,blunt,concurred,booth,accusations,scholarship,forlorn,rags,unlucky,dazzling,handsomely,purchases,proportionate,terribly'.split(',')
fields['Locke_SimpleIdeas']='time, name, light, length, taste, weight, figure, presence, space, shape, motion, absence, heat, union, volume, size, warmth, succession, void, depth, duration'.split(', ')
fields['Locke_SimpleModes']='day, life, nothing, place, end, state, rest, death, none, sense, something, night, hour, point, sight, form, view, age, occasion, number, force, design, year, birth, sound, change, distance, fact, stage, evening, season, period, month, source, instant, week, noise, spot, mark, image, height, article, prospect, minute, winter, sign, conclusion, morrow, term, darkness, addition, mode, frame, stream, circle, list, opening, multitude, corner, level, date, flow, class, match, midnight, yesterday, infancy, pace, antiquity, abundance, interval, scale, collection, aspect, mass, dozen, couple, yard, remainder, series, edge, vision, noon, amount, century, pitch, oblivion, obscurity, score, descent, feature, faculty, repetition, o\'clock, mile, sphere, pound, process, structure, specimen'.split(', ')
#fields['Locke_ComplexIdeasOfSubstance_orig']='man, mind, nature, heart, body, head, person, world, art, friend, face, author, blood, eye, air, soul, self, fire, home, ground, room, house, child, breast, bed, wife, object, earth, money, gentleman, family, court, town, master, letter, door, ear, father, nation, field, company, daughter, creature, mother, son, mankind, foot, seat, husband, enemy, brother, tongue, servant, fellow, wind, reader, gold, bosom, throne, stranger, hair, breath, mouth, sun, arm, paper, picture, witness, sea, street, estate, smile, crown, storm, bread, party, mistress, city, sword, parent, table, boy, dress, food, road, slave, kingdom, sister, page, lover, countenance, stone, glass, soldier, girl, writer, flame, flesh, silver, treasure, prey, ship, chair, tear, substance, shore, tree, school, brain, infant, thunder, iron, skin, queen, neck, brow, wood, chain, board, garden, window, fool, dream, animal, weather, poison, chamber, tyrant, dust, dinner, dog, wound, fruit, pen, monster, load, ring, prison, moon, feast, villain, soil, offspring, army, village, sky, devil, hero, tide, river, rival, rain, coat, fever, widow, branch, hat, store, coast, disease, foe, gate, supper, victim, meat, shadow, rock, lordship, prize, band, angel, hall, nobody, wall, member, neighbourhood, prisoner, market, box, building, cheek, pocket, physician, posterity, cloud, stomach, maker, bit, guardian, ocean, mountain, philosopher, plate, merchant, minister, publication, cap, material, nose, darling, surface, roof, shelter, tea, ball, officer, residence, host, publick, captain, ornament, flower, carriage, port, province, palace, grain, flood, poetry, copy, ray, chapter, corn, crowd, veil, abode, cup, forest, lightning, priest, individual, island, plant, bank, bell, compass, parish, tomb, floor, altar, furniture, star, meal, assembly, coach, uncle, plague, habitation, patron, throat, blush, sage, shoulder, tribe, guest, castle, bride, sheep, poem, crew, drink, marble, grove, shop, finger, steel, nurse, county, oak, pipe, fountain, traveller, stuff, knee, lion, cottage, continent, tract, liquor, scripture, attendant'.split(', ')
#fields['Locke_MixedModes_orig']='care, reason, power, truth, order, word, purpose, work, case, love, use, pleasure, manner, opinion, peace, respect, fear, vain, matter, doubt, thought, danger, knowledge, strength, account, passion, loss, virtue, duty, regard, attention, interest, service, question, advantage, proof, fate, price, character, scene, language, ease, example, sake, appearance, honour, beauty, opportunity, law, assistance, pity, action, favour, justice, history, friendship, praise, glory, trouble, judgment, liberty, title, rage, genius, confusion, faith, difference, situation, merit, practice, satisfaction, memory, health, wonder, grace, shame, value, war, charge, notice, fault, passage, freedom, fortune, chance, difficulty, joy, pain, consideration, experience, rule, temper, meaning, necessity, idea, patience, confidence, circumstance, relief, story, aid, ignorance, defence, crime, conscience, conduct, description, sex, violence, deal, event, command, safety, intention, plan, concern, delight, rank, conversation, distinction, disposition, task, ambition, importance, consent, expression, relation, labour, claim, misery, speech, observation, resolution, wisdom, variety, fame, grief, ruin, haste, goodness, perfection, understanding, weakness, note, government, protection, method, property, application, punishment, authority, argument, help, courage, marriage, profession, sleep, reputation, revenge, progress, promise, zeal, expence, esteem, wit, triumph, pardon, reflection, inclination, folly, terror, request, suspicion, custom, security, leisure, impression, information, gratitude, opposition, sentence, vengeance, approbation, objection, resentment, humanity, exercise, compassion, affair, caution, flight, mischief, evil, contempt, anger, skill, imagination, accident, approach, performance, reign, education, pretence, charity, fury, evidence, trial, industry, fashion, system, habit, debt, obedience, sacrifice, employment, vice, trust, mistake, applause, study, suit, excuse, murder, censure, pursuit, benefit, admiration, remembrance, writing, meeting, music, battle, tenderness, complaint, guilt, search, discourse, victory, stroke, propriety, retreat, song, injury, fancy, probability, direction, capacity, demand, kindness, excess, tale, science, reception, harmony, foundation, delay, entrance, journey, prudence, reproach, generosity, station, creation, empire, scheme, instruction, benevolence, gift, examination, entertainment, discovery, malice, sentiment, report, lot, intelligence, alteration, indulgence, instrument, comparison, resemblance, alarm, management, path, motive, deed, charm, disappointment, neglect, greatness, feeling, jealousy, eloquence, reverence, issue, testimony, conviction, arrival, elegance, anguish, indignation, notion, remedy, representation, vigour, behaviour, compliment, doctrine, amusement, submission, repose, hint, discretion, game, proposal, attachment, sin, lustre, anxiety, mirth, remark, treatment, prosperity, audience, severity, astonishment, profit, enquiry, prejudice, bounty, insolence, envy, imitation, rapture, reasoning, apology, sport, sickness, waste, obligation, introduction, tone, departure, policy, operation, oppression, reply, conquest, luxury, apprehension, madness, corruption, worship, certainty, attack, woe, preference, verse, sincerity, affliction, commerce, ceremony, indifference, expedient, fondness, oath, encouragement, presumption, modesty, dance, style, belief, integrity, injustice, absurdity, disguise, partiality, restraint, consolation, hurry, phrase, tribute, voyage, reformation, surprize, toil, supposition, transport, awe, complexion, effort, provision, behalf, aversion, assurance, invention, sweetness, explanation, resistance, piety, hatred, insult, intercourse, hazard, preservation, curse, connection, calm, tyranny, eminence, excellence, devotion, separation, abuse, commission, possibility, vessel, scorn, assertion, breach, continuance, reserve, privilege, debate, persuasion, advance, intent, decay, posture, gloom, mystery, fraud, growth, defect, tumult, regret, tendency, impatience, lesson, scandal, enjoyment, incident, pomp, parting, bargain, temptation, veneration, shock, compliance, quarrel, sale, extremity, establishment, refuge, appetite, sanction, repentance, hunger, signal, undertaking, fatigue, climate, confession, test, grandeur, standard, surprise, theme, communication, ardour, correspondence, counsel, declaration, interview, sensibility, recovery, stile, determination, glance, fidelity, warning, artifice, plot, blaze, energy, bliss, deity, solitude, constancy, impulse, softness, bloom, dominion, civility, felicity, diligence, confinement, tempest, honesty, persecution, council, maxim, cost, rent, displeasure, thirst, convenience, experiment, preparation, production, homage, conflict, superstition, contest, defiance, balance, talent, sagacity, conjecture, access, engagement, discipline, indolence, affectation, miracle, contradiction, basis, trifle, alliance, gravity, disadvantage, perusal, decency, torture, agony, welfare, activity, retirement, tranquillity, politeness, decision, model, decline, bond, jest, abhorrence, dancing, strife, agitation, enmity, moderation, readiness, settlement, position, flattery, nonsense, breeding, mortification, sway, contemplation, extravagance, disgust, refusal, interruption, accusation, exit, humility, institution, divinity, vein, draught, construction'.split(', ')


#fields['Locke_ComplexIdeasOfSubstance_word_pos']=[u'god_n', u'man_n', u'lord_n', u'men_n', u'king_n', u'people_n', u'sir_n', u'world_n', u'nature_n', u'church_n', u'body_n', u'person_n', u'son_n', u'house_n', u'christ_n', u'court_n', u'father_n', u'country_n', u'persons_n', u'hand_n', u'spirit_n', u'heart_n', u'water_n', u'earth_n', u'mind_n', u'hall_n', u'head_n', u'soul_n', u'england_n', u'city_n', u'blood_n', u'heaven_n', u'prince_n', u'kingdom_n', u'hands_n', u'land_n', u'john_n', u'money_n', u'children_n', u'parliament_n', u'town_n', u'letter_n', u'duke_n', u'fire_n', u'air_n', u'eyes_n', u'wife_n', u'lady_n', u'ground_n', u'army_n', u'france_n', u'nation_n', u'earl_n', u'friends_n', u'sea_n', u'arms_n', u'author_n', u'company_n', u'rome_n', u'friend_n', u'queen_n', u'crown_n', u'family_n', u'jesus_n', u'book_n', u'london_n', u'lands_n', u'brother_n', u'lords_n', u'party_n', u'feet_n', u'enemy_n', u'foot_n', u'enemies_n', u'woman_n', u'face_n', u'art_n', u'letters_n', u'mankind_n', u'county_n', u'eye_n', u'goods_n', u'river_n', u'room_n', u'mother_n', u'daughter_n', u'gentleman_n', u'scripture_n', u'kings_n', u'officers_n', u'child_n', u'hearts_n', u'henry_n', u'souls_n', u'william_n', u'bishop_n', u'troops_n', u'pope_n', u'bill_n', u'gold_n', u'emperor_n', u'gentlemen_n', u'charles_n', u'sons_n', u'inhabitants_n', u'estate_n', u'women_n', u'captain_n', u'wine_n', u'horse_n', u'princes_n', u'servant_n', u'master_n', u'thomas_n', u'ships_n', u'servants_n', u'houses_n', u'husband_n', u'scotland_n', u'mouth_n', u'ship_n', u'books_n', u'throne_n', u'minds_n', u'spirits_n', u'spain_n', u'jews_n', u'corn_n', u'britain_n', u'paul_n', u'tree_n', u'governor_n', u'pieces_n', u'heirs_n', u'creatures_n', u'trees_n', u'bread_n', u'wood_n', u'heir_n', u'stone_n', u'james_n', u'gods_n', u'ministers_n', u'david_n', u'commons_n', u'flood_n', u'island_n', u'ireland_n', u'field_n', u'temple_n', u'countries_n', u'brethren_n', u'sun_n', u'edward_n', u'bed_n', u'george_n', u'tears_n', u'board_n', u'door_n', u'text_n', u'wind_n', u'defendant_n', u'table_n', u'creature_n', u'sword_n', u'heads_n', u'churches_n', u'madam_n', u'seed_n', u'salt_n', u'reader_n', u'soldiers_n', u'peter_n', u'fellow_n', u'oil_n', u'saints_n', u'leaves_n', u'tongue_n', u'fruit_n', u'lines_n', u'tome_n', u'europe_n', u'paper_n', u'silver_n', u'apostle_n', u'angels_n', u'clergy_n', u'robert_n', u'israel_n', u'plaintiff_n', u'ear_n', u'devil_n', u'minister_n', u'judges_n', u'richard_n', u'flowers_n', u'hat_n', u'iron_n', u'parents_n', u'province_n', u'scriptures_n', u'waters_n', u'courts_n', u'plants_n', u'chapter_n', u'fathers_n', u'authors_n', u'flesh_n', u'page_n', u'writers_n', u'food_n', u'muff_n', u'apostles_n', u'officer_n', u'vessels_n', u'prophet_n', u'seat_n', u'port_n', u'horses_n', u'parish_n', u'hell_n', u'root_n', u'prisoner_n', u'paris_n', u'self_n', u'wall_n', u'fleet_n', u'lordship_n', u'disciples_n', u'glass_n', u'cap_n', u'belt_n', u'breast_n', u'college_n', u'romans_n', u'cities_n', u'moses_n', u'mountains_n', u'ears_n', u'towns_n', u'road_n', u'hill_n', u'writings_n', u'egypt_n', u'offices_n', u'witness_n', u'moon_n', u'surface_n', u'sinners_n', u'substance_n', u'writer_n', u'guard_n', u'priest_n', u'bishops_n', u'ladies_n', u'disease_n', u'italy_n', u'animal_n', u'walls_n', u'fruits_n', u'families_n', u'garden_n', u'weather_n', u'grave_n', u'prophets_n', u'cattle_n', u'palace_n', u'assembly_n', u'chamber_n', u'stall_n', u'prison_n', u'ghost_n', u'clerk_n', u'marks_n', u'provinces_n', u'grounds_n', u'race_n', u'building_n', u'bath_n', u'isle_n', u'stones_n', u'copy_n', u'angel_n', u'member_n', u'committee_n', u'luke_n', u'dominions_n', u'fever_n', u'priests_n', u'arm_n', u'prisoners_n', u'provisions_n', u'gate_n', u'shore_n', u'heavens_n', u'estates_n', u'colonel_n', u'castle_n', u'countenance_n', u'neck_n', u'jury_n', u'coast_n', u'clothes_n', u'vessel_n', u'elizabeth_n', u'altar_n', u'kingdoms_n', u'soil_n', u'home_n', u'flower_n', u'philip_n', u'ice_n', u'sister_n', u'branch_n', u'witnesses_n', u'hair_n', u'mayor_n', u'neighbours_n', u'animals_n', u'market_n', u'saint_n', u'archbishop_n', u'liquor_n', u'rock_n', u'sheep_n', u'fish_n', u'roots_n', u'merchants_n', u'seeds_n', u'tail_n', u'meat_n', u'plate_n', u'breath_n', u'sand_n', u'rivers_n', u'papers_n', u'manor_n', u'commissioners_n', u'citizens_n', u'plant_n', u'sovereign_n', u'chancellor_n', u'poet_n', u'scots_n', u'milk_n', u'signs_n', u'object_n', u'alexander_n', u'knight_n', u'bones_n', u'powder_n', u'justices_n', u'maid_n', u'senses_n', u'cardinal_n', u'admiral_n', u'slaves_n', u'cloth_n', u'lion_n', u'daughters_n', u'bell_n', u'spaniards_n', u'adam_n', u'boy_n', u'princess_n', u'holland_n', u'wound_n', u'flame_n', u'gates_n', u'realm_n', u'jerusalem_n', u'francis_n', u'protestant_n', u'deity_n', u'virgin_n', u'fields_n', u'allies_n', u'chris_n']

#fields['Locke_MixedModes_word_pos']=[u'power_n', u'law_n', u'reason_n', u'manner_n', u'order_n', u'love_n', u'words_n', u'word_n', u'account_n', u'religion_n', u'faith_n', u'truth_n', u'matter_n', u'means_n', u'use_n', u'peace_n', u'honour_n', u'grace_n', u'care_n', u'war_n', u'authority_n', u'laws_n', u'opinion_n', u'government_n', u'work_n', u'sin_n', u'virtue_n', u'duty_n', u'glory_n', u'judgement_n', u'service_n', u'purpose_n', u'knowledge_n', u'liberty_n', u'favour_n', u'council_n', u'interest_n', u'justice_n', u'pleasure_n', u'condition_n', u'regard_n', u'works_n', u'office_n', u'happiness_n', u'danger_n', u'advantage_n', u'answer_n', u'sins_n', u'notice_n', u'mercy_n', u'lie_n', u'trade_n', u'reign_n', u'fear_n', u'joy_n', u'title_n', u'thoughts_n', u'principles_n', u'orders_n', u'method_n', u'business_n', u'strength_n', u'evidence_n', u'question_n', u'history_n', u'rule_n', u'wisdom_n', u'pain_n', u'charge_n', u'terms_n', u'affairs_n', u'ways_n', u'example_n', u'proof_n', u'return_n', u'fortune_n', u'value_n', u'powers_n', u'hope_n', u'command_n', u'matters_n', u'oath_n', u'salvation_n', u'difference_n', u'consideration_n', u'wit_n', u'appearance_n', u'prayer_n', u'hopes_n', u'aid_n', u'voice_n', u'execution_n', u'language_n', u'success_n', u'goodness_n', u'covenant_n', u'obedience_n', u'proportion_n', u'marriage_n', u'loss_n', u'members_n', u'society_n', u'necessity_n', u'opportunity_n', u'circumstances_n', u'reasons_n', u'names_n', u'relation_n', u'argument_n', u'providence_n', u'defence_n', u'beauty_n', u'duties_n', u'action_n', u'pains_n', u'desire_n', u'empire_n', u'advice_n', u'experience_n', u'foundation_n', u'conscience_n', u'pride_n', u'practice_n', u'situation_n', u'sake_n', u'thought_n', u'battle_n', u'fate_n', u'zeal_n', u'principle_n', u'treaty_n', u'character_n', u'labour_n', u'trial_n', u'sorts_n', u'security_n', u'price_n', u'benefit_n', u'charity_n', u'influence_n', u'rules_n', u'doubt_n', u'health_n', u'error_n', u'difficulty_n', u'dignity_n', u'expense_n', u'forces_n', u'resolution_n', u'promise_n', u'discourse_n', u'sentence_n', u'rights_n', u'credit_n', u'merit_n', u'disposition_n', u'species_n', u'worship_n', u'debt_n', u'passion_n', u'temper_n', u'vice_n', u'deal_n', u'quality_n', u'pretence_n', u'idea_n', u'occasions_n', u'conversation_n', u'learning_n', u'doctrine_n', u'praise_n', u'sentiments_n', u'statute_n', u'prayers_n', u'choice_n', u'opposition_n', u'custom_n', u'virtues_n', u'measures_n', u'possession_n', u'memory_n', u'trouble_n', u'behaviour_n', u'courage_n', u'attention_n', u'advantages_n', u'notion_n', u'constitution_n', u'declaration_n', u'crime_n', u'confidence_n', u'kinds_n', u'intention_n', u'property_n', u'ruin_n', u'payment_n', u'exercise_n', u'application_n', u'ideas_n', u'conduct_n', u'note_n', u'importance_n', u'righteousness_n', u'offence_n', u'speech_n', u'violence_n', u'safety_n', u'riches_n', u'sacrifice_n', u'reputation_n', u'grief_n', u'arts_n', u'verse_n', u'turn_n', u'repentance_n', u'guilt_n', u'observation_n', u'cure_n', u'relief_n', u'accounts_n', u'opinions_n', u'arguments_n', u'pardon_n', u'support_n', u'affair_n', u'purposes_n', u'freedom_n', u'reward_n', u'concern_n', u'fault_n', u'piety_n', u'punishment_n', u'designs_n', u'privileges_n', u'contempt_n', u'ignorance_n', u'actions_n', u'issue_n', u'variety_n', u'proceedings_n', u'genius_n', u'comfort_n', u'assistance_n', u'rent_n', u'deed_n', u'respect_n', u'consequences_n', u'delight_n', u'evil_n', u'consent_n', u'ease_n', u'endeavours_n', u'rage_n', u'need_n', u'scene_n', u'passions_n', u'sorrow_n', u'patience_n', u'help_n', u'manners_n', u'folly_n', u'gift_n', u'story_n', u'understanding_n', u'commission_n', u'testimony_n', u'holiness_n', u'christianity_n', u'news_n', u'inquiry_n', u'fame_n', u'miracles_n', u'commerce_n', u'crimes_n', u'share_n', u'act_n', u'wrath_n', u'rank_n', u'misery_n', u'confusion_n', u'meaning_n', u'observations_n', u'event_n', u'creation_n', u'obligation_n', u'education_n', u'journey_n', u'methods_n', u'wealth_n', u'pleasures_n', u'revelation_n', u'remedy_n', u'operation_n', u'report_n', u'attempt_n', u'friendship_n', u'plea_n', u'bond_n', u'counsel_n', u'fears_n', u'trust_n', u'promises_n', u'petition_n', u'prejudice_n', u'errors_n', u'thanks_n', u'blessing_n', u'scheme_n', u'progress_n', u'liberties_n', u'notions_n', u'ambition_n', u'debts_n', u'vanity_n', u'belief_n', u'difficulties_n', u'corruption_n', u'communion_n', u'weakness_n', u'information_n', u'privilege_n', u'forms_n', u'expression_n', u'suit_n', u'description_n', u'claim_n', u'humour_n', u'wonder_n', u'qualities_n', u'examination_n', u'prudence_n', u'views_n', u'shame_n', u'nobility_n', u'charms_n', u'style_n', u'malice_n', u'acquaintance_n', u'arrival_n', u'reformation_n', u'agreement_n', u'murder_n', u'satisfaction_n', u'capacity_n', u'innocence_n', u'disorder_n', u'services_n', u'eternity_n', u'inclination_n', u'siege_n', u'division_n', u'gratitude_n', u'blessings_n', u'gifts_n', u'discovery_n', u'behalf_n', u'profession_n', u'desires_n', u'expedition_n', u'dominion_n', u'esteem_n', u'honours_n', u'lot_n', u'address_n', u'plan_n', u'song_n', u'testament_n', u'distress_n', u'philosophy_n', u'pity_n', u'system_n', u'wars_n', u'imagination_n', u'decree_n', u'fury_n', u'play_n', u'silence_n', u'management_n', u'breach_n', u'devotion_n', u'conditions_n', u'truths_n', u'judgments_n', u'anger_n', u'discharge_n', u'surprise_n', u'skill_n', u'motives_n', u'inheritance_n', u'examples_n', u'rebellion_n', u'burden_n', u'sermon_n', u'habit_n', u'request_n', u'sufferings_n']

fields['Locke_ComplexIdeasOfSubstance']=[u'god', u'man', u'lord', u'men', u'king', u'people', u'sir', u'world', u'nature', u'church', u'body', u'person', u'son', u'house', u'christ', u'court', u'father', u'country', u'persons', u'hand', u'spirit', u'heart', u'water', u'earth', u'mind', u'hall', u'head', u'soul', u'england', u'city', u'blood', u'heaven', u'prince', u'kingdom', u'hands', u'land', u'john', u'money', u'children', u'parliament', u'town', u'letter', u'duke', u'fire', u'air', u'eyes', u'wife', u'lady', u'ground', u'army', u'france', u'nation', u'earl', u'friends', u'sea', u'arms', u'author', u'company', u'rome', u'friend', u'queen', u'crown', u'family', u'jesus', u'book', u'london', u'lands', u'brother', u'lords', u'party', u'feet', u'enemy', u'foot', u'enemies', u'woman', u'face', u'art', u'letters', u'mankind', u'county', u'eye', u'goods', u'river', u'room', u'mother', u'daughter', u'gentleman', u'scripture', u'kings', u'officers', u'child', u'hearts', u'henry', u'souls', u'william', u'bishop', u'troops', u'pope', u'bill', u'gold', u'emperor', u'gentlemen', u'charles', u'sons', u'inhabitants', u'estate', u'women', u'captain', u'wine', u'horse', u'princes', u'servant', u'master', u'thomas', u'ships', u'servants', u'houses', u'husband', u'scotland', u'mouth', u'ship', u'books', u'throne', u'minds', u'spirits', u'spain', u'jews', u'corn', u'britain', u'paul', u'tree', u'governor', u'pieces', u'heirs', u'creatures', u'trees', u'bread', u'wood', u'heir', u'stone', u'james', u'gods', u'ministers', u'david', u'commons', u'flood', u'island', u'ireland', u'field', u'temple', u'countries', u'brethren', u'sun', u'edward', u'bed', u'george', u'tears', u'board', u'door', u'text', u'wind', u'defendant', u'table', u'creature', u'sword', u'heads', u'churches', u'madam', u'seed', u'salt', u'reader', u'soldiers', u'peter', u'fellow', u'oil', u'saints', u'leaves', u'tongue', u'fruit', u'lines', u'tome', u'europe', u'paper', u'silver', u'apostle', u'angels', u'clergy', u'robert', u'israel', u'plaintiff', u'ear', u'devil', u'minister', u'judges', u'richard', u'flowers', u'hat', u'iron', u'parents', u'province', u'scriptures', u'waters', u'courts', u'plants', u'chapter', u'fathers', u'authors', u'flesh', u'page', u'writers', u'food', u'muff', u'apostles', u'officer', u'vessels', u'prophet', u'seat', u'port', u'horses', u'parish', u'hell', u'root', u'prisoner', u'paris', u'self', u'wall', u'fleet', u'lordship', u'disciples', u'glass', u'cap', u'belt', u'breast', u'college', u'romans', u'cities', u'moses', u'mountains', u'ears', u'towns', u'road', u'hill', u'writings', u'egypt', u'offices', u'witness', u'moon', u'surface', u'sinners', u'substance', u'writer', u'guard', u'priest', u'bishops', u'ladies', u'disease', u'italy', u'animal', u'walls', u'fruits', u'families', u'garden', u'weather', u'grave', u'prophets', u'cattle', u'palace', u'assembly', u'chamber', u'stall', u'prison', u'ghost', u'clerk', u'marks', u'provinces', u'grounds', u'race', u'building', u'bath', u'isle', u'stones', u'copy', u'angel', u'member', u'committee', u'luke', u'dominions', u'fever', u'priests', u'arm', u'prisoners', u'provisions', u'gate', u'shore', u'heavens', u'estates', u'colonel', u'castle', u'countenance', u'neck', u'jury', u'coast', u'clothes', u'vessel', u'elizabeth', u'altar', u'kingdoms', u'soil', u'home', u'flower', u'philip', u'ice', u'sister', u'branch', u'witnesses', u'hair', u'mayor', u'neighbours', u'animals', u'market', u'saint', u'archbishop', u'liquor', u'rock', u'sheep', u'fish', u'roots', u'merchants', u'seeds', u'tail', u'meat', u'plate', u'breath', u'sand', u'rivers', u'papers', u'manor', u'commissioners', u'citizens', u'plant', u'sovereign', u'chancellor', u'poet', u'scots', u'milk', u'signs', u'object', u'alexander', u'knight', u'bones', u'powder', u'justices', u'maid', u'senses', u'cardinal', u'admiral', u'slaves', u'cloth', u'lion', u'daughters', u'bell', u'spaniards', u'adam', u'boy', u'princess', u'holland', u'wound', u'flame', u'gates', u'realm', u'jerusalem', u'francis', u'protestant', u'deity', u'virgin', u'fields', u'allies', u'chris']

fields['Locke_MixedModes']=[u'power', u'law', u'reason', u'manner', u'order', u'love', u'words', u'word', u'account', u'religion', u'faith', u'truth', u'matter', u'means', u'use', u'peace', u'honour', u'grace', u'care', u'war', u'authority', u'laws', u'opinion', u'government', u'work', u'sin', u'virtue', u'duty', u'glory', u'judgement', u'service', u'purpose', u'knowledge', u'liberty', u'favour', u'council', u'interest', u'justice', u'pleasure', u'condition', u'regard', u'works', u'office', u'happiness', u'danger', u'advantage', u'answer', u'sins', u'notice', u'mercy', u'lie', u'trade', u'reign', u'fear', u'joy', u'title', u'thoughts', u'principles', u'orders', u'method', u'business', u'strength', u'evidence', u'question', u'history', u'rule', u'wisdom', u'pain', u'charge', u'terms', u'affairs', u'ways', u'example', u'proof', u'return', u'fortune', u'value', u'powers', u'hope', u'command', u'matters', u'oath', u'salvation', u'difference', u'consideration', u'wit', u'appearance', u'prayer', u'hopes', u'aid', u'voice', u'execution', u'language', u'success', u'goodness', u'covenant', u'obedience', u'proportion', u'marriage', u'loss', u'members', u'society', u'necessity', u'opportunity', u'circumstances', u'reasons', u'names', u'relation', u'argument', u'providence', u'defence', u'beauty', u'duties', u'action', u'pains', u'desire', u'empire', u'advice', u'experience', u'foundation', u'conscience', u'pride', u'practice', u'situation', u'sake', u'thought', u'battle', u'fate', u'zeal', u'principle', u'treaty', u'character', u'labour', u'trial', u'sorts', u'security', u'price', u'benefit', u'charity', u'influence', u'rules', u'doubt', u'health', u'error', u'difficulty', u'dignity', u'expense', u'forces', u'resolution', u'promise', u'discourse', u'sentence', u'rights', u'credit', u'merit', u'disposition', u'species', u'worship', u'debt', u'passion', u'temper', u'vice', u'deal', u'quality', u'pretence', u'idea', u'occasions', u'conversation', u'learning', u'doctrine', u'praise', u'sentiments', u'statute', u'prayers', u'choice', u'opposition', u'custom', u'virtues', u'measures', u'possession', u'memory', u'trouble', u'behaviour', u'courage', u'attention', u'advantages', u'notion', u'constitution', u'declaration', u'crime', u'confidence', u'kinds', u'intention', u'property', u'ruin', u'payment', u'exercise', u'application', u'ideas', u'conduct', u'note', u'importance', u'righteousness', u'offence', u'speech', u'violence', u'safety', u'riches', u'sacrifice', u'reputation', u'grief', u'arts', u'verse', u'turn', u'repentance', u'guilt', u'observation', u'cure', u'relief', u'accounts', u'opinions', u'arguments', u'pardon', u'support', u'affair', u'purposes', u'freedom', u'reward', u'concern', u'fault', u'piety', u'punishment', u'designs', u'privileges', u'contempt', u'ignorance', u'actions', u'issue', u'variety', u'proceedings', u'genius', u'comfort', u'assistance', u'rent', u'deed', u'respect', u'consequences', u'delight', u'evil', u'consent', u'ease', u'endeavours', u'rage', u'need', u'scene', u'passions', u'sorrow', u'patience', u'help', u'manners', u'folly', u'gift', u'story', u'understanding', u'commission', u'testimony', u'holiness', u'christianity', u'news', u'inquiry', u'fame', u'miracles', u'commerce', u'crimes', u'share', u'act', u'wrath', u'rank', u'misery', u'confusion', u'meaning', u'observations', u'event', u'creation', u'obligation', u'education', u'journey', u'methods', u'wealth', u'pleasures', u'revelation', u'remedy', u'operation', u'report', u'attempt', u'friendship', u'plea', u'bond', u'counsel', u'fears', u'trust', u'promises', u'petition', u'prejudice', u'errors', u'thanks', u'blessing', u'scheme', u'progress', u'liberties', u'notions', u'ambition', u'debts', u'vanity', u'belief', u'difficulties', u'corruption', u'communion', u'weakness', u'information', u'privilege', u'forms', u'expression', u'suit', u'description', u'claim', u'humour', u'wonder', u'qualities', u'examination', u'prudence', u'views', u'shame', u'nobility', u'charms', u'style', u'malice', u'acquaintance', u'arrival', u'reformation', u'agreement', u'murder', u'satisfaction', u'capacity', u'innocence', u'disorder', u'services', u'eternity', u'inclination', u'siege', u'division', u'gratitude', u'blessings', u'gifts', u'discovery', u'behalf', u'profession', u'desires', u'expedition', u'dominion', u'esteem', u'honours', u'lot', u'address', u'plan', u'song', u'testament', u'distress', u'philosophy', u'pity', u'system', u'wars', u'imagination', u'decree', u'fury', u'play', u'silence', u'management', u'breach', u'devotion', u'conditions', u'truths', u'judgments', u'anger', u'discharge', u'surprise', u'skill', u'motives', u'inheritance', u'examples', u'rebellion', u'burden', u'sermon', u'habit', u'request', u'sufferings']





fields['Locke_ComplexIdeasOfRelation']='part, kind, side, cause, degree, piece, measure, consequence, sort, youth, spring, contrary, beginning, quarter, sum, pair, bottom, middle, plenty, mixture, top, result, origin, reverse, front, yonder'.split(', ')

fields['WordOrigin_NN_0850-1150']='time, day, part, man, life, nothing, name, mind, care, place, heart, kind, end, body, rest, death, none, side, head, light, truth, world, friend, word, work, love, something, night, blood, manner, eye, length, soul, sight, peace, fear, self, fire, home, thought, knowledge, strength, ground, passion, loss, room, house, child, breast, service, year, bed, wife, sake, earth, law, birth, youth, friendship, court, town, title, master, door, ear, father, health, wonder, field, grace, daughter, shame, weight, mother, son, freedom, mankind, foot, spring, husband, brother, tongue, temper, meaning, beginning, fellow, deal, wind, reader, gold, evening, bosom, hair, breath, mouth, sun, month, speech, wisdom, arm, goodness, understanding, witness, sea, help, week, street, sleep, crown, storm, bread, mark, wit, triumph, shape, sword, table, boy, height, food, road, flight, kingdom, sister, evil, skill, charity, winter, bottom, stone, glass, trust, girl, writer, murder, writing, meeting, guilt, darkness, heat, song, flesh, silver, treasure, ship, tale, tear, shore, tree, school, brain, frame, thunder, middle, iron, queen, neck, gift, brow, wood, board, warmth, lot, stream, circle, top, path, deed, greatness, feeling, weather, opening, dust, dog, game, wound, fruit, sin, mirth, load, ring, moon, offspring, sickness, devil, tide, flow, worship, rain, fever, widow, woe, verse, hat, midnight, foe, gate, oath, yesterday, meat, shadow, lordship, angel, belief, hall, wall, market, box, cheek, sweetness, hatred, curse, cloud, bit, philosopher, breach, privilege, cap, yard, nose, darling, growth, roof, port, edge, flood, corn, sale, crowd, noon, cup, hunger, priest, island, standard, plant, bell, warning, plot, blaze, bliss, floor, altar, star, meal, council, rent, thirst, throat, score, shoulder, mile, talent, guest, castle, bride, sheep, miracle, drink, pound, grove, finger, steel, oak, pipe, knee, lion, bark, smoke, writ, cloth, dish, harvest, oil, grass, yoke, hole, laughter, thread, silk, spark, shrine, nearer, boat, tail, clay, penny, sail, straw, beam, glow, thief, hue, stead, wedding, shilling, boldness, coldness, west, wheel, fortnight, thyself, shot, ghost, cock, knife, horn, shield, dost, forehead, seed, pre, doom, belly, heap, whisper, sand, morn, weapon, forgiveness, foul, cat, brass, pit, bone, stake, fold, leaf, throng, lake, blast, fare, inn, elder, idleness, worm, hunting, shepherd, burden, smell, pole, lap, linen, shell, beard, lip, candle, bridge, herd, aught, beer, bitterness, shower, pine, feather, ice, limb, mourning, kitchen, bowl, arrow, frost, epistle, offering, stick, chest, longing, copper, lust, healing, cradle, hedge, errand, cow, shift, lamb, knot, ass, breed, accord, wax, tie, landlord, sheet, fan, slain, wheat, spear, thereto, ore, ale, chin, turf, staff, childhood, liking, earnestness, token, net, blade, gleam, organ, clerk, leather, fowl, tooth, deer, bee, mill, mantle, goose, steward, moss, needle, cheese, betwixt, sweat, mood, den, timber, cart, mist, elbow, wolf, wool, craft, bush, womb, ward, cook, ditch, blindness, brook, oxen, swimming, boon, beneath, hen, toe, heed, harp, didst, bend, egg, shaft, ox, spur, thigh, apostle, pin, thrust, crystal, mare, web, martyr, fain, apple, richness, hardness, bridegroom, thickness, highway, calf, meadow, chaplain, nut, creed, barley, oar, pool, weed, hearth, bath, stem, kin, heathen, kinsman, sack, reed, player, din, witch, trap, hook, hymn, ladder, threat, ridge, cliff, lime, saddle, thumb, chill, holiness, bind, rider, carelessness, disciple, spake, swan, goat, holiday, lore, lily, crow, hood, snake, nail, span, mortar, threshold, ash, fasting, shalt, ware, liver, wire, archbishop, righteousness'.split(', ')
fields['WordOrigin_NN_1150-1700']="""nature, reason, power, state, sense, person, order, face, purpose, case, use, pleasure, author, hour, opinion, point, air, respect, form, vain, matter, doubt, view, age, occasion, danger, cause, account, number, virtue, degree, duty, piece, regard, attention, force, interest, design, question, advantage, proof, fate, price, object, character, scene, language, ease, measure, example, appearance, honour, money, beauty, opportunity, consequence, assistance, pity, action, gentleman, favour, family, sort, justice, history, sound, praise, change, taste, glory, trouble, liberty, rage, genius, confusion, letter, faith, difference, situation, merit, practice, satisfaction, memory, nation, distance, company, value, war, fact, charge, creature, figure, notice, fault, passage, fortune, seat, chance, difficulty, joy, contrary, pain, consideration, experience, enemy, rule, necessity, idea, patience, confidence, stage, circumstance, relief, story, aid, servant, ignorance, defence, crime, conscience, conduct, description, sex, violence, event, command, safety, intention, presence, plan, concern, delight, rank, conversation, distinction, throne, disposition, task, stranger, ambition, importance, season, consent, expression, relation, labour, claim, period, misery, observation, resolution, variety, fame, source, grief, ruin, haste, perfection, weakness, instant, note, government, protection, paper, picture, method, property, application, punishment, authority, argument, quarter, courage, marriage, noise, profession, reputation, revenge, estate, progress, promise, smile, zeal, spot, party, mistress, space, esteem, pardon, reflection, inclination, city, folly, terror, request, suspicion, custom, security, motion, principle, leisure, impression, parent, information, gratitude, opposition, image, sentence, dress, vengeance, approbation, objection, resentment, humanity, exercise, slave, absence, compassion, affair, caution, article, sum, mischief, contempt, anger, pair, imagination, accident, prospect, page, approach, performance, lover, reign, education, countenance, fury, evidence, trial, minute, industry, fashion, system, habit, debt, sign, obedience, sacrifice, conclusion, employment, vice, mistake, applause, study, morrow, suit, soldier, excuse, censure, pursuit, benefit, admiration, remembrance, music, battle, term, tenderness, complaint, addition, search, discourse, victory, stroke, propriety, retreat, flame, injury, fancy, probability, direction, capacity, prey, demand, kindness, excess, science, reception, harmony, foundation, delay, chair, entrance, substance, journey, union, mode, prudence, volume, reproach, generosity, size, station, infant, creation, empire, scheme, skin, plenty, instruction, mixture, benevolence, examination, entertainment, discovery, malice, chain, sentiment, report, intelligence, alteration, indulgence, instrument, comparison, resemblance, alarm, management, motive, charm, disappointment, garden, window, neglect, jealousy, eloquence, reverence, issue, testimony, conviction, arrival, elegance, anguish, fool, list, dream, animal, indignation, notion, poison, chamber, succession, remedy, representation, multitude, vigour, corner, tyrant, compliment, doctrine, amusement, submission, repose, hint, result, dinner, discretion, proposal, attachment, lustre, pen, anxiety, remark, treatment, prosperity, monster, audience, severity, astonishment, profit, prejudice, bounty, prison, insolence, envy, feast, imitation, reasoning, level, apology, soil, sport, army, waste, obligation, introduction, tone, departure, policy, operation, village, oppression, sky, date, reply, conquest, hero, luxury, apprehension, river, madness, corruption, rival, certainty, coat, attack, branch, class, preference, match, sincerity, affliction, store, commerce, coast, disease, ceremony, indifference, expedient, fondness, supper, infancy, encouragement, presumption, victim, rock, modesty, dance, prize, band, style, pace, integrity, injustice, nobody, antiquity, absurdity, abundance, disguise, interval, partiality, member, restraint, consolation, hurry, phrase, tribute, voyage, prisoner, reformation, toil, building, supposition, transport, awe, complexion, effort, provision, behalf, scale, pocket, aversion, assurance, invention, physician, collection, explanation, resistance, piety, insult, intercourse, void, hazard, posterity, preservation, aspect, calm, tyranny, eminence, stomach, origin, maker, guardian, excellence, ocean, mountain, devotion, separation, abuse, dozen, commission, possibility, depth, vessel, couple, scorn, plate, assertion, continuance, merchant, minister, publication, reserve, debate, persuasion, advance, intent, decay, material, posture, gloom, mystery, fraud, defect, surface, tumult, shelter, remainder, tea, regret, tendency, ball, impatience, lesson, officer, scandal, residence, host, enjoyment, captain, incident, pomp, parting, ornament, flower, carriage, bargain, series, temptation, veneration, province, shock, palace, grain, compliance, poetry, quarrel, copy, ray, chapter, vision, extremity, establishment, veil, refuge, appetite, sanction, abode, repentance, signal, forest, lightning, undertaking, fatigue, climate, confession, individual, test, grandeur, surprise, theme, communication, ardour, correspondence, counsel, reverse, declaration, interview, sensibility, recovery, determination, compass, glance, fidelity, amount, century, energy, deity, parish, tomb, solitude, constancy, impulse, front, yonder, bloom, dominion, furniture, civility, felicity, pitch, diligence, confinement, tempest, honesty, persecution, maxim, assembly, cost, coach, displeasure, uncle, plague, oblivion, habitation, patron, obscurity, blush, convenience, sage, experiment, preparation, production, homage, descent, feature, conflict, faculty, superstition, repetition, contest, defiance, balance, tribe, sagacity, conjecture, access, sphere, engagement, discipline, indolence, affectation, poem, crew, contradiction, basis, marble, trifle, process, structure, alliance, gravity, shop, disadvantage, perusal, decency, torture, agony, welfare, activity, nurse, county, retirement, fountain, tranquillity, politeness, decision, model, stuff, decline, bond, jest, abhorrence, dancing, strife, agitation, enmity, moderation, duration, readiness, settlement, cottage, position, flattery, nonsense, breeding, mortification, continent, specimen, sway, contemplation, extravagance, disgust, refusal, tract, interruption, accusation, scripture, humility, institution, divinity, vein, draught, attendant, construction, purchase, consciousness, region, wickedness, decree, medicine, dawn, sympathy, allowance, essence, exception, ingenuity, satire, uncertainty, beast, keeper, machine, arch, amazement, novelty, pause, leader, obstinacy, monument, detail, resignation, firmness, pence, conception, tune, roll, resource, caprice, generality, inconvenience, division, confirmation, appeal, agreement, narrative, plea, morality, messenger, diversion, solemnity, vale, powder, leg, nobleman, appellation, profusion, career, mansion, vow, concert, channel, pulse, pencil, transaction, administration, credulity, regularity, benefactor, agent, vivacity, sensation, robe, exchange, contention, nobility, recommendation, contract, centre, inspection, extreme, slavery, valour, scruple, struggle, enthusiasm, fable, artist, closet, somebody, estimation, strait, discontent, contrast, deceit, magnificence, famine, treason, survey, record, square, gown, afternoon, treachery, falsehood, seal, perseverance, trumpet, slender, relish, deference, summit, cousin, combat, partner, spleen, community, subsistence, breathing, rebellion, concurrence, fiction, farm, removal, prosecution, revolution, disturbance, native, rod, clock, magic, owner, reverend, speculation, equality, acceptance, imputation, bulk, magnitude, embrace, converse, precaution, murmur, intimacy, income, propensity, isle, sect, slaughter, fund, eagerness, dislike, theory, controversy, assent, utility, rapidity, camp, element, advocate, combination, demonstration, gale, deliverance, hospitality, dependence, tragedy, riot, dissolution, accuracy, funeral, inheritance, footing, hesitation, dialogue, painter, patriot, breeze, criticism, pressure, cell, breakfast, treaty, gaiety, spectacle, print, handkerchief, efficacy, dexterity, denial, spectator, tax, gallantry, track, preface, precipice, acknowledgment, cultivation, wave, dirt, familiarity, crisis, row, bard, league, conqueror, liberality, destiny, inhabitant, regulation, warrant, eternity, perplexity, stature, discord, violation, car, cunning, equity, syllable, critic, circulation, tradition, majority, accommodation, citizen, deficiency, scope, invasion, pledge, vigilance, occupation, proverb, outside, coffee, facility, couch, quest, captivity, torment, correction, lad, bag, sire, partake, vote, review, deformity, ink, palm, founder, protector, subjection, serpent, deportment, library, catalogue, adoration, resort, failure, consternation, pot, diet, incense, nephew, traitor, consumption, dame, plunder, joke, curtain, eve, portrait, sketch, vent, splendour, historian, independence, pattern, intrigue, verge, elevation, observance, obstacle, impossibility, wherefore, autumn, rebel, magistrate, soever, conformity, infidelity, successor, damage, gaze, beef, tenor, patronage, treatise, lamp, idol, reference, conjunction, restoration, composure, conference, outrage, deliberation, refinement, faction, troop, dwelling, stain, fabric, salvation, dagger, baggage, inscription, conceit, valley, gentleness, commencement, imprisonment, card, prerogative, dishonour, maintenance, cabinet, interpretation, pension, sterling, precept, attitude, exactness, ladyship, pupil, melody, desolation, topic, range, manhood, bay, vindication, assault, pamphlet, realm, surgeon, parade, comment, immortality, import, delivery, jury, destitute, cash, tribunal, denomination, custody, distribution, receipt, horseback, observer, prophet, solicitude, dispatch, brute, connexion, stamp, gallery, asylum, tour, exile, champion, precision, fee, beggar, cloak, loyalty, variation, vegetable, reconciliation, nourishment, coin, meditation, prophecy, hospital, par, tent, gout, commander, mob, murderer, disobedience, imposition, park, niece, regiment, array, fort, dove, orator, monarchy, metropolis, interposition, comedy, bull, sufferer, cave, renown, sanctity, temperance, colouring, investigation, lawn, recess, completion, mate, prose, warrior, manuscript, inside, allusion, impunity, diversity, symptom, parson, compound, latitude, scent, quod, rear, cannon, campaign, competition, catastrophe, echo, iniquity, maturity, tincture, metal, offender, paragraph, scarcity, disaster, tutor, aunt, inspiration, edifice, depravity, diamond, slumber, trading, stress, wreck, dose, risk, drama, subscription, animosity, pack, margin, peer, ministry, circuit, sovereignty, waist, riches, senate, fro, solution, banishment, excursion, product, tenant, infirmity, accent, succour, club, chapel, physic, entry, trip, counsellor, advancement, concealment, pestilence, convent, pox, uniformity, illustration, pint, distrust, deception, gilt, inquiry, column, adversary, robbery, confederacy, nerve, workmanship, engine, utterance, justification, porter, chorus, juncture, emblem, dismay, condemnation, representative, comprehension, navy, aforesaid, analogy, relaxation, sine, formation, prohibition, omission, travelling, garb, motto, treasury, anecdote, statesman, juice, vault, judgement, definition, victor, republic, artillery, malady, cent, congregation, suggestion, drum, velvet, grasp, dusky, preacher, border, decease, sinner, legislature, crest, lace, secretary, landing, pistol, irregularity, detection, district, apparel, tobacco, usurpation, fishing, chaos, abyss, bearer, fertility, forbearance, patent, continuation, allegiance, basket, vehicle, deck, mud, pulpit, garrison, redemption, populace, conversion, salary, horizon, ruler, memorial, shirt, atonement, exhibition, popularity, sailor, mirror, adherence, jurisdiction, garment, circumference, commendation, clause, teacher, inquisition, intemperance, idolatry, pavement, hunter, substitute, whip, politician, dining, statute, twilight, territory, cane, bud, navigation, shout, exclamation, legacy, festival, household, landscape, foreigner, deviation, compensation, ambassador, alternative, inconsistency, unity, fortress, group, desertion, peculiarity, inter, insect, expert, corpse, incapacity, debtor, moisture, slip, quotation, consultation, sublimity, culture, indication, function, ode, spy, oration, pilot, penance, sequel, variance, manufacture, excellency, brandy, colony, barrier, fairy, rhetoric, manager, register, avenue, predecessor, consistency, traffic, passenger, intercession, promotion, extension, resurrection, route, agriculture, effusion, commodity, vacancy, corps, divorce, meagre, earthquake, atmosphere, embassy, symmetry, sufficiency, heroism, lance, charter, estimate, tip, amendment, surrender, communion, brick, conveyance, flag, block, dedication, drunkenness, exclusion, truce, fragment, movement, assemblage, concession, ally, beach, anarchy, vegetation, pond, gesture, reckoning, pearl, ivory, coffin, foresight, candidate, reduction, fence, hypothesis, fraternity, judging, mutton, pasture, nursery, digestion, similarity, lieu, folio, ounce, gunpowder, expulsion, extraction, palate, zone, sculpture, translator, scotch, counter, circular, remnant, celebration, remonstrance, sanctuary, stability, rhyme, magazine, interpreter, advertisement, hermit, goal, cellar, session, lecture, visitor, hut, spell, monastery, illusion, intervention, selfishness, tavern, canal, prelate, association, mail, meridian, domain, foliage, suppression, suspension, olive, revolt, duel, verdict, insurrection, annum, mechanic, propagation, compact, hostility, clearness, resident, rubbish, choir, tenure, veteran, massacre, deduction, cord, burial, exemption, exhortation, canvas, pilgrimage, expiration, planet, dragon, dialect, extinction, link, creditor, assassination, oracle, rumour, velocity, reliance, spaniard, pursuance, gang, militia, ancestor, rebuke, printer, collar, humiliation, diameter, banker, enterprise, fervour, fallacy, architect, emphasis, panic, enumeration, grammar, papa, herald, grandson, enlargement, packet, mediation, carpenter, depression, fluid, apparatus, warfare, suicide, vicinity, emergency, director, restriction, printing, partition, banner, fur, fun, section, carpet, toleration, angle, sacrament, adoption, exterior, cheerfulness, fellowship, washing, curate, bias, cathedral, cream, drift, rupture, destination, limitation, pig, rejoicing, loan, accumulation, dissatisfaction, oratory, forfeiture, cake, version, type, trespass, scaffold, contact, rite, logic, violet, attainment, opponent, population, citadel, restitution, hind, mayor, duchess, mission, abbey, boundary, edict, remission, bait, cement, cavity, prevention, prediction, selection, area, index, chivalry, cavalry, constable, code, machinery, trader, renewal, covenant, baptism, concord, brush, usefulness, cotton, skull, battery, jacket, rice, guise, miniature, importation, mineral, gravel, libel, concave, inventor, summary, temperature, firing, punch""".split(', ')
fields['AbstractValues_NN']='heart, passion, virtue, character, shame, temper, conduct, bosom, reputation, zeal, esteem, principle, caution, vice, admiration, propriety, excess, sentiment, indulgence, feeling, discretion, sin, prejudice, presumption, modesty, partiality, restraint, excellence, reserve, ardour, sensibility, softness, decency, politeness, moderation, extravagance, humility, morality, nobility, deference, infamy, arrogance, conformity, decorum, conceit, gentleness, ostentation, depravity, mildness, insensibility, pang, ardor, sobriety, impartiality, licentiousness, debauchery, haughtiness, fervour, prodigality, bias, prepossession, indecency'.split(', ')
fields['HardSeed_NN']='head, face, work, blood, eye, ground, breast, ear, foot, tongue, gold, hair, mouth, arm, pair, flesh, silver, skin, neck, brow, opening, cheek, stomach, dozen, couple, nose, front, throat, shoulder, finger, knee, vein, roll, leg, spleen, forehead, belly, bone, outside, smell, lap, palm, beard, verdure, lip, limb, chest, inside, chin, waist, trip, tooth, nerve, elbow, dusky, womb, toe, slip, bend, thigh, heel, azure, dropt, thumb, fist, quiver, skull, liver'.split(', ')

fields['Locke_ComplexIdeasOfSubstance_General']='man, mind, nature, body, person, world, art, friend, author, self, home, room, house, child, wife, object, earth, gentleman, family, court, town, master, father, nation, field, company, daughter, creature, mother, son, mankind, husband, enemy, brother, servant, fellow, reader, stranger, paper, picture, witness, party, mistress, parent, boy, dress, sister, page, lover, countenance, girl, writer, prey, substance, infant, queen, fool, animal, weather, tyrant, monster, load, villain, offspring, army, village, devil, hero, widow, store, coast, foe, victim, lordship, band, nobody, member, neighbourhood, physician, posterity, maker, guardian, philosopher, merchant, minister, darling, officer, residence, host, publick, captain, province, poetry, copy, crowd, abode, priest, individual, bank, parish, assembly, uncle, habitation, patron, sage, tribe, guest, bride, crew, nurse, county, traveller, stuff, continent, tract, material'.split(', ')
fields['Locke_ComplexIdeasOfSubstance_Specific']='heart, head, face, blood, eye, air, soul, fire, ground, breast, bed, money, letter, door, ear, foot, seat, tongue, wind, gold, bosom, throne, hair, breath, mouth, sun, arm, sea, street, estate, smile, crown, storm, bread, city, sword, table, food, road, slave, kingdom, stone, glass, soldier, flame, flesh, silver, treasure, ship, chair, tear, shore, tree, school, brain, thunder, iron, skin, neck, brow, wood, chain, board, garden, window, dream, poison, chamber, dust, dinner, dog, wound, fruit, pen, ring, prison, moon, feast, soil, sky, tide, river, rival, rain, coat, fever, branch, hat, disease, gate, supper, meat, shadow, rock, prize, angel, hall, wall, prisoner, market, box, building, cheek, pocket, cloud, stomach, bit, ocean, mountain, plate, publication, cap, nose, surface, roof, tea, ball, ornament, flower, carriage, port, palace, grain, flood, ray, chapter, corn, veil, cup, forest, lightning, island, plant, bell, compass, tomb, floor, altar, furniture, star, meal, coach, plague, throat, blush, shoulder, castle, sheep, poem, drink, marble, grove, shop, finger, steel, oak, pipe, fountain, knee, lion, cottage, liquor, scripture, attendant'.split(', ')

#combos=[('HardSeed','AbstractValues'), ('WordOrigin_1150-1700','WordOrigin_0850-1150')]
#for c1 in combos[0]:
#	for c2 in combos[1]:
#		words = set(fields[c1]) & set(fields[c2])
#		fields[c1+'_and_'+c2] = list(words)

#fields['Abstract']="""disinterested, candour, modesty, politeness, levity, haughtiness, moderation, sincerity, licentiousness, affectation, probity, ingenuous, vicious, generosity, indulgence, decency, magnanimity, licentious, partiality, unjustifiable, sentiment, illiberal, unbecoming, sensibility, immoral, affability, virtuous, inconsistent, amiable, injudicious, heroism, sobriety, ungenerous, selfish, enthusiasm, imprudent, respectful, humanity, frankness, tenderness, unaffected, refinement, unwarrantable, humility, diffidence, prudence, indifference, benevolence, extravagance, inconsistency, gentleness, generous, impartiality, absurd, intemperate, modest, unreasonable, indecent, deportment, ostentatious, passionate, consistent, simplicity, blameable, piety, sincere, sentiments, fondness, virtue, behaviour, integrity, humiliating, arrogant, indecency, extravagant, temerity, vanity, liberality, arrogance, affection, incompatible, profligate, deference, inexcusable, complacency, character, exemplary, discreet, gallantry, presumption, unpardonable, justifiable, severity, insolence, meekness, seriousness, elegance, esteem, submission, polite, propriety, attachment, imprudence, impiety, avowed, meanness, considerate, injustice, ostentation, friendship, depravity, manners, conduct, insolent, singularity, rudeness, injurious, disgusting, rectitude, passion, austerity, infidelity, supercilious, merit, depraved, presumptuous, absurdity, prudent, sanctity, bigotry, civility, equitable, disposition, obstinacy, impropriety, ridiculous, scrupulous, morality, unrestrained, refined, innate, justly, unmanly, inattention, constancy, discernment, adulation, detestation, insensibility, folly, pursuits, dissipation, consistently, abilities, gaiety, rashness, dignified, fashionable, reasonable, acquiescence, culpable, judicious, temper, lenity, resentment, excesses, inferiors, serious, selfishness, excellence, capricious, gratitude, steadiness, abhorrence, compliance, zeal, affecting, vivacity, ingenuity, respect, petulance, credulity, dispositions, professions, rational, prodigality, ardour, beneficent, contempt, brutality, temperance, independence, debauchery, endearing, excusable, decorum, feelings, boldness, propensities, kindness, timidity, strictness, parental, tyrannical, fortitude, justify, wisdom, actuated, inordinate, irrational, expressions, confidence, prepossession, engaging, chastity, disapprobation, ferocity, regularity, indelicate, interference, vehemence, sensual, precision, indolent, susceptible, impertinence, mildness, unaccountable, neglect, unexceptionable, obliging, unmerited, baseness, flattery, compassionate, unsuitable, deserving, calmness, discretion, inexperience, forbearance, accomplishments, dignity, impetuosity, patriotism, inspired, brutal, partial, affectionate, excess, prejudices, servility, zealous, benignity, dictates, ignorance, attachments, interested, veneration, familiarity, harshness, preposterous, magnanimous, censure, energy, consistency, ungovernable, freedom, honesty, goodness, atrocious, censured, approbation, propensity, scandalous, assiduity, prejudice, exalted, asperity, ingratitude, indolence, indulged, justness, malevolence, firmness, correct, disdained, disgust, observance, austere, impartial, ambition, tranquillity, characteristic, affect, beneficence, detestable, worthy, dictate, shameless, regard, contemptible, despotic, gratifying, address, misguided, warmth, affected, passions, qualifications, humane, resent, conscientious, revere, candid, uniformity, rash, distrust, affable, grossest, superiors, reserve, uncommon, contradiction, caution, malignity, clemency, filial, cruelty, fidelity, authorize, consciousness, actions, desirable, heroic, merited, principle, prejudiced, becoming, adopt, pleasantry, formality, interesting, approve, admiration, fervour, liberal, reject, justified, hospitality, inculcate, excellencies, excessive, constitutional, resignation, greatness, superstitious, enlightened, cowardice, assuming, universally, practices, moderate, cheerfulness, temperate, criminal, moral, religious, unfeeling, pride, softness, illegal, eloquence, habitual, applause, commendation, seducing, respectable, untainted, earnestness, infamous, motive, endowments, education, weakness, sublimity, unnatural, ignorant, avarice, abject, prevalence, convincing, blameless, impertinent, rigorous, innocence, plainness, pernicious, affections, admirers, exaggerated, repugnant, sensible, happiness, disgusted, shameful, suitable, exhortations, morals, perfidy, inclinations, highly, extravagancies, mortifying, tranquility, superstition, professed, invective, restraint, foolish, barbarity, rationally, unalterable, outrage, unfeigned, adviseable, haughty, perseverance, insignificant, consummate, intemperance, insinuation, dishonour, refinements, hypocritical, reproach, pathetic, naturally, magnificence, despotism, negligent, novelty, understanding, enjoyments, maxims, attentions, veracity, consequence, prostituted, gratified, exorbitant, violation, sincerely, ambitious, chearfulness, specious, noble, warmest, connexions, attractions, unfavourable, sanction, irresistible, learning, devotion, applauded, groundless, avidity, headstrong, forgiving, falsehood, eloquent, accuracy, solicitude, allurements, dislike, declamation, explicit, conversation, imputation, contradictory, conformity, inconstant, raillery, transcendent, suggestions, reproof, plausible, impious, assertions, munificence, expression, wickedness, chaste, agreeable, prevalent, essential, language, reverence, vehement, pious, reason, speculative, popularity, flattering, luxury, imposture, strictest, honourable, prosperity, spirited, concession, surely, adopting, refusal, diffident, despised, sense, misled, grandeur, style, assurance, rapturous, conformable, unbounded, incorrigible, invidious, debase, pretensions, sufficiency, weaknesses, unjustly, trifling, indispensable, conviction, dictated, abuse, importance, disdain, impunity, artless, felicity, truth, barbarous, favourable, inhuman, acquiesce, irregularity, assertion, condemn, philosophic, erudition, ought, aversion, notions, illiterate, indulging, mankind, unparalleled, persuasions, instructive, ungrateful, inherent, restraints, prevailing, audacious, concise, communicative, resentments, exertions, admonition, adequate, degeneracy, reasonably, rigour, submissive, attainment, equality, encourage, suspicious, iniquitous, sweetness, hatred, indignity, devout, displeasing, importunate, drunkenness, beauty, intrepidity, invariable, eligible, authorised, deserve, caprice, deemed, authority, authorized, indignation, courtship, charity, connections, conceited, solicitations, intercourse, circumstanced, professedly, degree, coldness, resented, compassion, unkind, preference, suggest, avow, abstracted, instruction, friendships, studiously, emotions, despondency, distinction, patriotic, indulge, inattentive, indulgent, dissatisfied, tolerated, distressing, interests, enjoyment, implacable, abused, encouraged, wisest, intimacy, equity, sophistry, respected, sober, dutiful, treatment, presuming, prompted, characteristics, tyranny, inferiority, composure, reasoning, fondest, professing, inducement, fallacy, encouragement, failings, admonitions, inspire, decisions, perfections, convince, mistaken, immoderate, concessions, unsuspecting, however, calumny, utility, deservedly, understandings, perspicuity, expressive, repress, confident, perverted, tendency, credulous, emulation, zealously, remonstrance, honour, insinuations, equivocal, despise, avaricious, invectives, opinions, moralist, unquestionably, constraint, idleness, recommend, outrageous, personal, tranquil, chimerical, unquestionable, loyalty, predominant, acknowledgement, peremptory, gratify, shocked, result, cautious, inadequate, lawless, possessed, accomplished, opulence, serenity, foolishly, remonstrances, imputed, decent, holiness, truest, doctrines, insupportable, disregarded, displeased, ridicule, indirect, insufferable, openly, inviolable, admirer, suppress, obstinate, servile, profusion, truly, wildness, possessing, desires, excited, modes, insinuating, reputable, christianity, institution, countenanced, retirement, encouraging, illustrious, incomprehensible, pretend, advantages, frailty, peculiarities, impotence, ashamed, endearments, impudence, enforced, fiction, fanciful, ludicrous, reproaches, bashful, metaphysical, carelessness, attention, animated, modish, jealousy, genius, conscious, gesture, applaud, adherence, genteel, acceptable, unthinking, perfidious, policy, impute, principles, juster, severest, induce, seditious, rigid, acknowledgment, whimsical, gravity, precise, sensations, national, classical, courteous, profane, lavish, profession, circumstances, conciliate, pleasures, blamed, performances, offended, innovation, equally, dishonest, enamoured, amusements, solicitation, childish, unfair, slavish, justifies, solitude, convinced, unnecessary, ardent, seasonable, appetites, nicety, seduced, vices, implicit, sinful, comprehensive, averse, wantonness, persuasion, fallacious, influence, importunity, penetration, maternal, undeniable, encourages, wilful, demonstrations, tenets, positive, instances, desire, irreconcileable, deem, provoked, artful, virtues, regarded, animosity, activity, undertaking, insult, selection, knowledge, powerful, declarations, necessary, strict, ridiculed, ferocious, courtly, imperfections, precept, indigence, obedience, cherished, acquire, effectual, tempers, extraordinary, innocent, judgement, detest, charitable, disdainful, enormity, partially, peculiarly, erroneous, restrained, contradictions, estimation, recommendation, impatience, refractory, solemnity, persist, establishments, artifice, insipid, degrade, festivity, important, patronage, addicted, obscurity, acknowledge, entertaining, ridiculously, literary, flattered, reproached, conceptions, qualities, inclination, brilliancy, unacquainted, useful, attach, impudent, resolute, facility, meek, sagacity, suppressing, insatiable, representations, exasperated, admire, favourably, excite, foibles, vice, treated, institutions, rancour, invariably, inspires, intrepid, deliberate, humiliation, professes, luxurious, rashly, regretted, abstruse, interest, noblest, decisive, perverse, refin'd, undoubtedly, exactness, justification, mildest, writer, mode, deserved, profess, successful, frugal, abominable, ardently, graceful, consideration, precepts, talents, behave, pleasing, tacitly, fond, eagerness, regulation, contemporaries, oppressive, usurpation, lewd, minds, advantageous, disagreeable, rapacity, reputation, penitence, connection, sublime, mislead, observer, forcible, induced, generality, peculiarity, trivial, provoking, adoration, congenial, acknowledgments, apologize, welfare, entertain, establish, opulent, belief, industry, strictly, insensible, conducive, despising, independency, imposing, benign, imaginable, criticism, insults, indifferent, unconnected, precaution, toleration, attribute, prudently, grateful, honest, madness, vigilance, vague, seeming, modestly, dissemble, whatever, seriously, possess, alarming, deserves, opprobrious, warmly, passionately, teachers, spirit, timed, instance, peaceable, urged, intelligible, expensive, brevity, inflexible, declaration, opinion, protestations, entitle, acuteness, dissatisfaction, addresses, contrition, highest, governments, reciprocal, sympathy, qualify, deviation, behaved, fierceness, intrinsic, unanswerable, prejudicial, customs, reasonings, inoffensive, valour, nature, relinquish, blame, youth, chastisement, conjugal, notwithstanding, philosopher, entertainments, praise, society, sentimental, assemblies, disgraced, therefore, elevated, prostitute, merits, characters, indications, reserved, accustomed, excels, softened, exemption, testimonies, administration, comply, love, reflection, restrictions, enormities, reformed, disappointment, formal, necessarily, captivate, idolatry, intellectual, malice, such, merely, daring, legislature, boasting, inexplicable, acknowledged, hardship, amusement, real, despises, publicly, generously, forgiveness, impossibility, embellishments, bitterness, probability, condemns, approving, limitations, excuse, hypocrite, wishes, justice, imply, association, disadvantageous, undertakings, appearances, implies, community, grossly, sufficiently, philosophy, offending, obligations, superior, vindictive, righteousness, possibly, acrimony, obligation, effectually, steady, prerogatives, intentions, public, general, similarity, honoured, argues, taste, elaborate, unguarded, neatness, ignoble, obvious, depreciate, constrained, liberties, opposition, indirectly, nevertheless, solicitous, civilities, peculiar, defects, persisting, disorderly, lover, models, faults, impartially, incurring, invincible, constant, allegorical, investigation, independent, imitation, suicide, natured, egregious, doctrine, unmixed, assent, stupid, revengeful, enforce, courage, perfection, unhappy, inquisitive, timorous, splendour, dangerous, most, infirmity, practise, suspicions, neglecting, imposition, heinous, poverty, soothing, inactivity, wicked, fictions, determination, dissipated, excites, qualified, accuse, seduce, ingenious, excellency, injured, hazardous, incapable, distressed, personally, satisfied, studied, gross, efficacious, suspicion, even, mysterious, unremitting, accomplish, blindness, proposal, morally, christian, investigate, subordination, ascribe, advocates, pleased, offensive, imaginations, imitate, favour, patriot, frigid, natural, manner, testified, precarious, humour, notion, apparently, suspected, prodigal, wise, regulate, peevish, assume, deception, favoured, contradicted, universal, condemning, unfortunate, omission, uneasiness, accusation, cultivate, ambiguous, ideas, detested, happy, expectations, unsullied, punish, fashion, scorn, reigning, punishing, effect, infamy, constitution, errors, assuredly, satisfaction, supreme, dejection, certainty, powerfully, intrusion, coldly, judgment, ministry, legislators, heightened, mirth, injunctions, argument, incensed, pleasure, inflame, action, parents, sovereigns, alacrity, exultation, stile, sensibly, effects, palliate, amorous, ultimately, performance, tinctured, delusive, cultivation, subjection, assumed, concern, exacted, practice, solicited, terms, ardor, blinded, limitation, deity, opportunities, supposition, restriction, treachery, philosophical, benefactors, definition, contradict, truths, advocate, divested, blasphemy, certainly, speaker, extensive, compulsion, praises, hostility, duty, tender, necessity, imbibed, restraining, clandestine, compositions, suffering, proceedings, abhor, professors, glory, humblest, voluntary, patrons, cherish, affront, religiously, temptations, really, nobly, impulse, improvement, savage, merriment, sex, improbable, captivated, abandon, controul, impotent, villainy, confessedly, deceitful, private, thankful, obdurate, indication, rejecting, caresses, examples, betrayed, assiduous, solicit, satire, inseparable, deplorable, less, demonstration, impure, wealth, mutual, lively, relish, blemishes, deformity, astonishing, attentive, testimony, diabolical, domestic, emphasis, assert, disobedience, entreaties, suggested, anxiety, not, profuse, humorous, crime, violated, harsh, acquired, uncertainty, regret, endeavour, suited, amicable, adopted, endued, compliment, unsuccessful, temptation, capable, misery, requisites, vulgar, competition, gestures, enmity, mildly, awkward, elegant, that, government, manly, owning, pathetically, sallies, employments, fictitious, honor, object, corrupt, discipline, considered, unanimous, reconciliation, deluded, offence, orators, guilt, politely, insulted, assurances, politics, incomparable, punishments, acknowledging, comparison, tempered, conclusive, convinces, aspiring, capacity, votaries, commercial, exquisite, jealous, formally, desertion, entertained, exact, corruption, unshaken, interfere, punishment, disguise, submit, suspect, amity, aggravation, corruptions, unavoidable, pecuniary, favouring, foreigners, quality, resolution, conspicuous, altogether, sanctified, recommending, delicate, extinction, gallant, destructive, tenderest, extinguish, admirable, contrast, diversity, oppression, proofs, compensate, diligence, familiar, picturesque, arguments, proof, chastise, particular, simplest, privacy, insinuate, counteract, objections, affliction, condescended, arduous, acting, unhappily, godlike, pertinent, careless, individual, counsels, reclaim, honourably, acquiring, topics, imagination, extreme, reconcile, expose, restrain, strenuously, efficacy, intense, distinguishing, suppressed, associate, rewards, dulness, vindicated, poignant, man, accusations, anger, charmed, gentlest, subservient, intimate, sufficient, incapacity, transports, meanly, gen'rous, persuade, abuses, learned, resisted, writings, whoever, inconsiderable, application, limited, remorse, retract, benefactor, disdaining, beings, of, this, presumed, duties, accountable, delighted, cordially, sufferer, apprehensions, awe, falsely, wit, demonstrate, amendment, expressing, decision, courageous, slighted, personages, humble, judging, circumstance, cruel, reflections, excel, forbidding, ineffectual, neglects, deceit, conceived, envy, endowed, designing, favours, intention, magistrate, enterprise, erring, indigent, acceptance, deficient, illustration, favor, positions, earnest, forgive, beneficial, conform, ourselves, mortified, strongly, frailties, approves, panegyric, parent, animation, inspiring, acquit, adventurer, legitimate, shame, apt, considering, pedant, fervent, figurative, disadvantage, outrages, ignominy, pity, officious, discontent, prescribe, always, allowed, obviate, peremptorily, accept, philosophers, perfect, fraud, nonsense, injure, loyal, legal, feeling, illustrate, embarrassed, persisted, mortify, transaction, calculated, confessed, displeasure, penitent, concealment, checked, disdains, exhortation, societies, renders, barbarian, sanctify, gracious, experience, nobleman, to, wretchedness, treat, abstinence, moderns, sufferings, mind, thoroughly, inferences, exerted, punctual, difficult, reform, industriously, exception, converse, rebuke, liberty, usefulness, insulting, unpleasant, vanities, situations, antiquated, winning, instructing, anxieties, apparent, tiresome, abstract, prompts, inequality, flatterers, disadvantages, animate, induces, insuperable, bewitching, alliance, disobey, simple, necessities, palpable, obnoxious, illusion, example, flights, awaken, majority, promote, facetious, nicest, confession, admit, success, nation, modern, extorted, materially, elevation, assured, renounce, indefatigable, imagery, think, boisterous, undisturbed, yourself, engages, undoubted, guidance, consequently, it, manifest, prompt, endeavours, anarchy, slavery, meanest, vindicate, asserted, falshood, estimate, predecessors, facts, apology, mixture, vain, strenuous, fortunate, maxim, witty, atonement, sexes, clearness, dispensed, respectfully, emolument, unsettled, correspondence, phrenzy, suggestion, paternal, engagements, experienced, exalt, splendid, objects, comparatively, involuntary, offend, bountiful, insinuated, influences, suggests, render, proposals, exciting, answerable, hitherto, oppressed, avoiding, obstinately, crimes, benefits, historian, deceived, ignominious, denial, regards, divine, subject, oblige, overlooked, presume, exert, reforming, but, aukward, asserting, shamefully, calm, wickedly, nobler, rapture, drama, impose, descriptive, lust, embellish, unsupported, authentic, neglected, contentment, allegory, possessors, intrigue, instructed, though, supernatural, behalf, distraction, desperation, parade, reconciled, productions, resolutions, acquisitions, author, indulgences, derision, undeceive, avoided, utterly, intimately, establishment, admitting, acted, claims, rejects, impositions, power, populace, critics, distinctions, distinguished, natures, totally, preferable, opposes, responsible, placid, attached, inveterate, renouncing, consider, human, evident, sovereign, flatter, scandal, mental, destitute, uninterrupted, glaring, uncommonly, appearance, critic, science, repentance, impatient, renounced, lamented, needless, accusing, evidence, happiest, dependence, correctly, likely, incurred, politicks, contented, unavailing, defective, profitable, betraying, connexion, natur'd, delightful, commend, riotous, prerogative, prosecution, wretched, disgrace, inexorable, statesman, repute, grounded, governing, christians, immortality, valuable, correction, contrary, esteemed, frantic, respects, unwelcome, blush, requires, abusing, preacher, wiser, corrected, spotless, disconcerted, be, promoting, sparing, interview, proud, oppressions, posterity, punished, injury, nobility, impracticable, congratulations, frequently, faction, sordid, possible, accepting, require, foresight, apprehension, sensation, equals, exists, surpassing, dissenters, gentle, greatest, expect, lucrative, inflamed, constitutes, submitted, sovereignty, soever, superfluous, transitory, judiciously, attain, would, symmetry, violate, idea, attributes, avoid, excelling, persecution, indelible, frankly, willing, friendly, publickly, garb, momentary, wildest, opponents, distresses, preferring, proper, pretended, wish, amend, superficial, concerns, fatally, utmost, satisfying, good, inclined, ignorantly, religions, topic, invaluable, tumultuous, amusing, prevails, attaining, representation, temporary, protesting, boundless, objection, conclusions, shewn, persons, attempt, entreat, laborious, alleviate, slander, excused, security, mortification, required, constitute, pretenders, improving, abandoning, easily, charms, situation, glorious, establishing, frenzy, observation, remind, more, misfortunes, pitied, constitutions, milder, preachers, accommodations, heathen, readily, transient, unusual, fluctuating, strongest, replete, contentions, injuries, habits, assented, infinitely, plead, simply, reception, rewarded, inconveniences, concurrence, weak, unexpected, appeal, inconceivable, anticipate, accommodation, impetuous, vexatious, authenticity, solemn, enslave, methods, can, allied, learn, prelate, custom, appellations, claim, whatsoever, persecuting, florid, uncultivated, humbler, detail, blended, much, rendering, adored, adapted, unskilful, reader, literal, dismission, stranger, conversion, advancement, patron, satisfy, distress, testify, harmony, lessons, indiscriminately, chearful, promises, surpassed, easy, lawful, commended, severer, defect, menaces, speculation, violence, considerations, wilfully, wayward, oratory, pardoned, violent, cultivated, discouraged, rules, kindest, reasons, protection, dispense, entitled, wrong, ruling, unity, offences, solely, judged, serene, poetry, condemnation, studies, renewal, enchanting, comedy, enterprize, permit, consequent, alliances, forgiven, appellation, union, laboured, ruinous, permitting, climate, purposes, ministerial, obtain, cheerful, discordant, riches, enjoined, express, active, irreparable, refine, yet, orator, slightest, hateful, presence, maintained, heroine, dissuade, pursuit, voluntarily, aspire, forgetful, probable, raptures, warranted, abundant, transport, favour'd, undaunted, legislative, idle, blindly, you, digression, if, schools, generally, minister, ascribed, incidents, anxious, magistrates, epithets, pastoral, advantage, associates, innocently, regular, clergy, seemingly, purest, exhibition, reliance, narrowness, madman, willingly, protested, mortifications, betray, considers, bestowing, combination, publick, awful, improperly, conduce, accepted, detection, calmly, lasting, exempt, fashions, consolation, sloth, vigorous, circumstantial, aristotle, world, liking, mistresses, dispensations, sacrifice, unprofitable, excessively, iniquity, reformation, disappoint, dependance, delights, comprehension, successes, beloved, abhors, impossible, antiquity, temperature, afflictions, englishman, submitting, must, liberally, hopeless, vile, erroneously, altercation, accurate, ministers, limits, knowing, privilege, denying, foreigner, tasteless, successfully, refusing, entertainment, youthful, pomp, improve, determined, themselves, least, praising, politic, graciously, excuses, unkindly, regulations, honestly, forbids, their, practical, suppression, representatives, talent, corporeal, multiplicity, may, confess, owing, audience, atone, adventitious, inactive, attributed, betrays, observant, treating, incoherent, correspondents, malignant, fault, services, concerned, mean, boast, disappointments, conclusion, artifices, coxcomb, prescribes, ensure, defined, similitude, rights, complicated, descriptions, clamorous, writers, apologies, breeding, unwearied, seem, infirmities, credited, patience, chargeable, subjects, concurred, princely, pretence, seldom, argue, entrusted, espouse, deliberation, venal, hoped, will, engross, unconcerned, recital, vilest, emoluments, orderly, particularly, conceit, applicable, barbarians, bounty, confessing, monitor, encreased, sprightly, owned, deliberately, is, relations, inestimable, perjury, chivalry, herein, obsequious, politician, arguing, hasty, declaring, loving, astonishment, obstacles, hereafter, inquiries, exerting, disappointed, exposes, wealthy, regulating, amply, exemplified, dedication, foreseeing, precedents, hate, miserable, prefer, melancholy, cowardly, strangers, attained, creator, never, quickness, antipathy, enjoy, combinations, unruly, who, source, only, marriage, exclusion, offer, introduce, reference, thought, wife, want, protected, matrimony, and, favourite, afflicted, jealousies, population, habit, singular, annihilate, inconveniencies, judge, what, reputed, greatly, alledged, downright, doubted, hypothesis, occasional, confidently, assure, person, reflect, scrutiny, admits, evil, workings, unutterable, too, purely, governors, obliges, phrases, sacrilegious, theirs, favourites, inexhaustible, suffer, demonstrated, confesses, recommended, casual, station, abundantly, pacific, publications, actual, brave, magnificent, rightly, perplexed, know, poetical, affirm, ordinances, pure, free, obtaining, adoption, thinking, assembly, extinguished, maintaining, model, abstain, objected, diligent, life, humbled, distribution, prove, bitterest, saucy, astonished, rather, external, athenians, significant, repent, villainous, whim, persecuted, assemblage, positively, govern, practised, translator, occasion, for, spite, believing, insensibly, afflict, repining, publishing, indisposition, ideal, happily, connected, infallible, propose, fruitless, vindication, apostle, greater, so, resist, vainly, governed, offender, previous, should, teach, adversaries, civil, obscure, pattern, engrossed, desperate, exclamations, majesty, venerable, function, abolish, irregular, dissipate, possibility, remark, sequestered, exist, excellent, apprehend, luxuries, revelation, contributed, tyrant, repaid, husbands, fact, sorrow, prosecuted, arrangement, rectify, noisy, repetitions, providence, tenacious, allow'd, remarked, encreasing, perusal, common, anguish, scriptures, unkindness, gaming, narration, celebrated, discontented, unceasing, prince, content, proposing, grievance, variety, manhood, system, charming, compelled, delight, as, plea, yourselves, accompanies, faculties, socrates, dissent, mistake, faithful, dared, distinguish, execrable, praised, interpretation, contemplation, persuading, divines, acquainted, childhood, allowing, nourish, engage, rendered, courted, information, manifestly, inconvenience, implanted, serviceable, prowess, rude, acts, dazzled, obscured, mysteries, hardships, import, thoughtful, earthly, intirely, imitating, brilliant, accusers, supplications, expected, disciple, bounties, dreaded, exclaim, wished, perfectly, brutes, senseless, eager, divinely, understood, animates, utter, enquiries, lenient, oftener, subversion, monarchy, absence, fullest, hurtful, competent, gifts, regarding, demands, visible, comparisons, prevail, interposition, afflicting, instrumental, prohibition, request, lessened, surmount, husband, concur, visibly, exclamation, defiance, evils, any, attitudes, acquitted, aweful, hope, contribute, usual, gradation, propagating, interpreted, researches, inevitable, enjoying, deny, subsequent, magnify, than, pretending, disguised, fundamental, contradicts, impressed, perpetuate, divest, remarkable, subordinate, conceiving, arts, expence, protestants, doubt, honours, repented, protector, anecdote, heedless, richness, licence, forgetting, premise, happier, insidious, redress, envious, those, earnestly, nice, dignities, wedlock, beauties, soliciting, negative, almighty, convenience, essex, faculty, resulting, citizen, cicero, continuance, cunning, anxiously, inclin'd, dress, frank, remarkably, republic, conquer, contention, introduction, disputes, forbear, stations, imagining, branded, myself, repetition, consent, beautiful, employment, section, forgotten, have, proving, gay, warned, secular, fortune, gospel, credit, management, confirm, proved, affirming, speculations, occasions, artfully, withstand, conscience, circumscribed, laughter, vehemently, controversy, mockery, valiant, calamitous, monarch, precipitate, suspecting, allusions, demonstrates, troublesome, needful, encroachments, formidable, reverse, bold, reward, noticed, insurrection, silence, impressions, tempt, affair, sciences, dependants, awakened, expences, fools, occurrences, judgments, dealings, attracting, retribution, catholic, advising, epistle, explanation, guided, unanimously, choice, injur'd, discourses, oppose, trusting, propositions, vigour, subdued, believe, conditions, former, consciences, clearly, perplexity, predecessor, tacit, ease, consenting, subsists, addison, accomplice, questioned, availed, hesitate, emulate, grief, henceforth, presumes, blessing, entertains, direct, promising, subduing, pagan, complaints, tamely, instructions, exercise, weakly, comforts, tully, deduction, pursuance, fully, changeable, teacher, unwilling, amuse, impaired, admitted, historians, heighten, advise, treacherous, lawfully, enliven, persecutions, easiest, basely, preferred, opposed, deviate, unlikely, rage, publication, them, traditions, promotion, marvellous, accurately, constituted, compliments, banished, authorities, antients, own, suffers, hostilities, condition, scipio, assign, imposed, base, acquaint, discover, united, material, countenance, event, concurring, delude, acknowledges, lordship, complain, study, instantaneous, joy, exhibit, reluctantly, painful, countrymen, impregnable, publish, inquiry, vicissitudes, revealed, addressed, exclude, imperceptibly, maintain, devoted, conclude, incur, intitled, reproaching, woman, diminish, question, credible, refuse, blest, entirely, permitted, mutually, create, due, stability, loved, observe, bliss, penalties, hereditary, endeavoured, purport, ascertain, contributes, difference, indignant, denied, obeying, beware, useless, design, state, comparative, concern'd, remembrance, aggravate, utterance, displaying, administer, events, wonderful, pardon, in, voltaire, god, mystery, female, instructs, convicted, prevailed, homage, merciless, asserts, fickle, deserting, deficiency, impart, graces, commonwealth, scorns, oppressors, assumes, communications, trust, antient, obliterated, sneer, sacrificing, attract, reciprocally, speech, disturbance, thoughts, neither, surprised, recal, tempted, managers, cultivating, weaken, exceeded, confirmation, whining, softening, effusion, adapt, able, ceremony, method, class, visitor, ascertained, favorite, wonder, propitious, comprehend, visitors, decided, conquerors, recommends, urge, endured, repulsed, yours, speeches, kind, parliaments, conception, seconded, curb, failure, hesitation, rebellious, trifles, requests, allegiance, guise, incurable, analogy, penury, great, confusion, effecting, otherwise, production, although, candidate, alteration, exigencies, condemned, cautiously, conversant, kindly, scenery, wisely, obstacle, appetite, comfortable, fiercest, property, mahometan, countenances, sedition, resistance, worship, cause, contributing, symptom, conversing, matters, could, operated, exempted, pleaded, deceiving, powers, fondly, maliciously, enquiry, disposed, preservation, comic, trespass, substitute, lovers, sinners, interruption, rank, perplexing, destination, transported, reminded, wishing, promoted, indebted, imperfect, effort, appeased, loudest, expedient, horace, offices, frustrated, compel, universities, verbal, respective, relaxation, possessions, correspond, infancy, whenever, applauding, disposes, prodigy, he, compared, destroys, intreat, admiring, apprized, extraction, prosecute, roughness, ornaments, colouring, sway, accounted, riot, display, construction, discussed, adore, shown, scenes, confused, might, gentler, subscribe, acquaintances, ceremonies, enumeration, excelled, precisely, readiness, occupation, blessings, disengaged, perplexities, pupils, odium, tragedy, deficiencies, teaching, shun, employing, future, blushes, evidences, confirming, proves, multitude, aspect, resigned, equipage, inaccessible, cooler, console, observable, gradations, amiss, mindful, stedfast, rally, enable, brutus, dominion, silly, warm, mere, emergency, earlier, sufferers, conceal, by, representative, shew, inhospitable, asham'd, prescribed, boasts, owe, invaders, intercession, howard, majestic, devote, equal, confinement, narrative, different, desist, bias, perpetual, all, tends, offenders, commit, immediate, dispute, retain, wedded, others, planned, inexpressible, wanton, protecting, means, sacred, authors, resolutely, nor, commerce, revenge, quiet, ascertaining, matchless, marrying, inanimate, leisure, spectator, enabled, softest, congreve, harangue, unsuspected, intrusted, dejected, safety, diverting, eternal, triumph, celebrating, tormenting, civilly, teaches, affectionately, uneasy, alarmed, diverted, partake, foreseen, instantaneously, councils, her, customary, accuser, augustus, the, hardened, prevention, attracted, improvements, ingeniously, engaged, faithless, frolic, courtier, fancy, which, forego, men, malady, dissolution, discourse, dreading, distinguish'd, duration, contest, rival, tyrants, suppose, asylum, am, numerous, rustic, herself, enjoyed, wondered, undertaken, position, renown, lament, languor, doing, corresponds, accent, benefit, culture, overpowered, vogue, exquisitely, enables, persecute, statesmen, unwary, reconciling, rhetoric, warn, patient, dismissed, essence, incline, mothers, withheld, inform, rulers, exercised, imitated, recollect, locke, feverish, sacrificed, suffered, propagate, thirdly, confute, nothing, politicians, hearts, poetic, strange, dispensing, beginnings, duly, claimed, unheard, widely, inflict, his, deprives, often, historical, existing, deduced, domestics, uttered, immortal, friend, repay, secondly, disorders, misfortune, these, seems, people, managed, magnified, they, obtained, durable, expeditious, forget, vanquished, practising, inspiration, fulfilled, confederacy, doubting, urging, mortals, gentry, absorbed, profound, abounds, scheme, proudest, fund, relative, abolished, accommodate, elevate, supplication, check, my, existed, total, hurry, disturbed, expectation, i, earliest, very, committing, protestant, expedients, novels, features, early, knew, tainted, affords, agree, aspires, aware, affairs, mien, furious, incident, detected, unobserved, embellished, humbly, inferior, trader, beyond, diversified, machinations, transporting, invent, posthumous, warfare, extort, fraternity, disturb, tenor, employ, demand, populous, depression, prosecuting, allows, unrelenting, opportunity, movements, auspices, meditated, operating, surprise, oblig'd, concludes, taught, amused, guiltless, fears, safest, hated, afford, grace, traders, contemplate, calamities, purer, speak, meant, originals, countess, itself, particulars, offspring, patterns, became, introduced, endangered, me, share, wretch, languishing, mercies, tinsel, exclusive, adventurers, strangely, alarm, retained, illustrated, brute, plainer, original, strengthened, needed, secretly, rivals, giddy, imperceptible, appealing, melody, reveal, temporal, conferred, dialect, secret, forfeited, overwhelm, beget, hopes, decorations, latter, nuptials, feature, we, none, estimated, judges, consider'd, applications, introduces, exposing, danger, ancient, sacrifices, ratified, abated, bounds, bestow, avail, combined, confide, complaint, ornament, conquest, faith, bestowed, harmless, worst, contests, fury, family, nicer, fulness, histories, without, scope, imitations, brightest, profusely, defender, consented, heartily, fearless, passive, languages, complex, physical, communicate, carefully, elude, alternative, declares, completion, decline, stated, expresses, concealing, imagined, controversies, wiles, infected, rapacious, homely, weakened, ladyship, charm, diversion, counsel, overcome, intreated, memorable, obeyed, conferring, warmer, title, admirably, nominal, compensation, warmed, understand, cool, coward, diffuse, complained, incense, prescription, proportion, airs, designs, repeatedly, offers, hardy, contagion, supposing, blessed, progress, weakest, appease, despairing, confounding, provincial, hates, refuses, minded, pleads, assistance, sects, cordial, prefers, walpole, declared, discovers, futurity, season, concerted, interrupted, sparingly, communication, him, omitted, improved, are, unmindful, answer, ambassador, informing, females, void, preferment, recollected, grievous, petitions, timely, studying, madly, meditation, rarely, tokens, opposing, fertility, likelihood, mature, distinguishes, rely, protest, briton, masquerade, attempts, withhold, scripture, interrupting, stubborn, involved, conflict, no, pronounces, increased, enjoys, vows, unless, acquainting, miscarriage, insist, rightful, anonymous, extirpate, foul, primary, expresly, affirmed, tumult, clamours, lustre, secondary, excellently, abate, grant, attested, princess, deed, art, themes, refrain, foreign, mitigate, matrimonial, preface, exceedingly, preliminary, treats, handsome, barb'rous, assuring, oration, access, milton, estates, thirst, admires, eminence, catastrophe, infallibly, clement, termed, mechanic, boundaries, exterior, exploits, episcopal, determine, observing, certain, enduring, remembered, mediation, fabulous, scholars, destiny, deeds, juncture, states, vested, tumults, hero, every, logic, spared, teeming, enhance, vassals, rapidity, specific, conflicts, wonderfully, forefathers, period, consist, stronger, familiarly, chagrin, menace, maturity, breach, impelled, prohibit, selves, celebration, debate, observed, preserve, frauds, women, exercises, produce, loveliest, assures, present, delirious, senate, conjecture, scarcely, suffice, meer, depress, dread, successor, scruple, ill, terrors, attitude, versed, inclines, discarded, desert, classic, properties, grievances, absent, expatiate, stifled, promise, gentleman, manager, unwillingly, conquered, exposed, tedious, carriage, comfort, astray, claiming, permits, abound, poem, amended, comprehended, expressly, dealing, moderately, endless, roused, preside, mission, scoundrel, ascribes, bad, relying, nations, moments, impediment, include, inevitably, behoves, mistress, affirmative, discovery, appeals, reparation, internal, dearly, masculine, advances, proprietors, enough, accommodated, answered, meaning, alterations, reply, disputed, excluded, ancients, overwhelmed, military, incredible, obedient, fearful, show, americans, undertake, clearest, betters, pursue, poorest, symptoms, profit, throughout, hardly, proposes, hearty, gen'ral, struggle, traffic, dedicate, transacted, relation, matron, 2dly, alone, confined, bravest, conversed, homer, better, penetrating, cruelly, anecdotes, engagement, adhered, resumed, popery, concerning, deceive, genial, inviting, dispel, comedies, needs, rule, mingled, uncouth, performer, titles, cease, thyself, force, gladly, defeat, cowley, befriend, masters, department, amours, conveniences, dearest, retrieve, factions, decently, propagation, disturbances, heart, foresee, obliged, assigned, undertakes, concord, remarking, doubly, heed, commanding, agreeing, climates, appealed, conjectures, accompanied, divorce, attended, please, englishmen, inglorious, latitude, tincture, talked, ecclesiastical, invention, informed, owed, extent, shewed, crisis, sage, unison, support, discern, constantly, assassination, evade, finally, skilful, darling, seemed, unknown, rebel, party, liked, unspeakable, destroy, explain, proceeding, sternly, lamentable, aught, thanked, explaining, cowards, intended, agitated, commotion, with, principally, editor, miraculous, unnoticed, miracle, represent, summary, gain, words, unerring, citizens, fool, especially, assigning, obsolete, lowly, occasionally, accounting, threats, proceeds, empress, auspicious, remedied, advised, need, whether, eternally, strain, fortunately, pitiful, conjunction, enlargement, fear, remission, foundation, subtle, tacitus, adversary, higher, lovely, continue, decree, skill, dubious, insure, trifle, hearken, well, whose, hinted, craft, dull, there, consternation, mention, climes, slave, premises, formation, memory, easier, mercy, best, preserved, fraught, surpass, forgery, delightfully, fulfilling, assassin, intelligence, few, struggles, copious, attacking, infirm, softens, bodily, joys, hereby, enacted, either, hesitated, exhibiting, heat, unable, witness, confest, denominations, presage, transactions, mechanical, feel, fellowship, music, produces, penal, princes, ours, depressed, endeavouring, enlarged, sentences, furnishes, contended, distracted, impatiently, monarchs, traitor, embassy, outward, attracts, antidote, still, drudgery, contract, indisposed, numberless, forcibly, distinct, resolved, dear, director, beseech, tend, avert, ages, juvenal, prospects, madam, mortal, debating, dexterity, interpose, proceed, resource, agitation, surly, bounteous, prediction, resolve, age, descendants, infectious, apprehending, advises, using, consists, procure, gust, proposed, originally, version, birth, unequal, giver, accompany, country, completely, relentless, tame, conquests, mixed, examination, whereas, secure, gloomy, preaching, nuptial, pow'rful, shewing, or, classed, noxious, gaining, notice, spoke, reluctant, choosing, discord, bustle, interrupt, friends, successors, warning, instituted, suspend, notable, cibber, directors, nearly, tended, sidney, endure, transcribe, bosoms, obey, communicating, enchantment, attest, sages, miracles, celestial, disordered, vastly, sources, quaint, sympathetic, founded, equivalent, repeated, nero, enemies, listening, communicated, promised, has, questions, subscription, debated, latent, supposes, value, pleas'd, machinery, speaking, wantonly, doubts, guardians, wretches, business, bestows, foregoing, jest, inspection, ferment, performers, clime, sooth, robust, lectures, deprive, worth, determin'd, sinister, designed, amazed, shews, resemblance, inward, embraces, impression, heiress, exceed, artists, everlasting, spectators, imports, stress, victim, tending, stuart, conceives, prose, created, concluded, firmest, us, fables, derived, soul, preserves, cato, agreement, administered, preserving, operate, ranked, income, boldly, subsisted, bitterly, tradesman, blind, spaniard, deceased, wherever, imperfectly, discovering, substituted, grieved, select, writing, romish, vowed, elsewhere, allowance, fittest, session, dazzling, celebrate, maiden, implored, redeemer, protect, recite, sharing, baffled, fancies, martial, doings, demanding, edict, banishment, regimen, suits, account, appropriated, eternity, metropolis, terrify, classes, variance, silent, feeble, riots, competitors, gravely, tradition, blend, determining, periods, code, tribute, spectacle, keenest, commission, consult, verified, inflicted, thereby, known, terrified, infer, assisting, prov'd, durst, resign, goddess, conference, preventing, just, attire, mentioning, allude, pamphlet, lessens, heights, glances, altered, undertook, saviour, hellish, shuns, penance, deaf, effected, considerable, task, lecture, sequel, creature, craving, ordinarily, muse, bitter, wou'd, intricate, amidst, connect, con, convert, idol, warlike, governs, pert, ask, weaker, agrees, enumerate, smile, find, nicely, a-days, ratify, repeat, visits, hostile, revolt, extinct, care, do, respectively, submits, conducted, ladies, diminished, disastrous, robbery, hardness, languid, nay, dependant, lov'd, detriment, married, depredations, thanking, she, laments, prophetic, granted, fatal, plutarch, libel, rebellion, offered, addition, shows, determines, fierce, jarring, communion, acclamations, styled, noblemen, obtains, sentence, advertisement, prone, mathematics, native, contributions, salvation, limit, heretofore, translations, strength, comical, joyous, fortunes, agent, monstrous, possession, pow'r, difficulty, university, perceptible, empire, tone, variations, prohibited, affording, further, menial, fairest, since, accomplishing, roughly, graver, converts, exhibited, embraced, disturbs, specimens, implore, humours, helpless, jovial, against, torments, remedies, describe, depriving, origin, opposite, subsist, spoken, remembering, ancestor, costly, defence, comely, pleases, britons, jurisdiction, voted, proportioned, alarms, committed, causes, explication, disturbing, company, abundance, demanded, nourished, comforted, diligently, omitting, commodious, how, continual, imprisonment, doubtful, lamentations, nephew, european, cou'd, prying, sakes, shou'd, was, separation, recited, conqueror, strengthens, spending, recount, restoration, courts, wives, oblivion, fathers, avails, consonant, marquis, deities, believed, playful, festivals, promotes, pleasant, wavering, adventures, intrude, disciples, additional, once, phrase, grieve, marks, sinner, tastes, peace, uniting, reduce, ceases, hermit, pronounce, appear, strengthen, doomed, succeed, progression, insisting, produced, strife, perform, lesson, heaven, lastly, believes, heroes, adhere, aiming, kindled, invasion, flourishing, antagonists, raphael, triumphant, contracted, contend, ruined, perused, expects, kindle, vigorously, provoke, agents, theory, peaceful, paragraph, embrace, betray'd, uttering, gloriously, directing, apprehended, jesus, laughed, improves, risk, frustrate, preach, forbid, nobody, referring, untimely, commander, accumulation, endanger, allies, other, superb, commotions, intend, functions, mightily, tutor, complaining, declined, performing, grants, aright, skilled, silently, oppress, consulted, revived, purpose, challenge, drift, delays, ambassadors, comparing, repel, comprehends, reign, commits, gods, artificial, professor, urg'd, dangers, repose, conveniency, includes, in-law, kinsman, chiefly, delay, reverend, fame, knows, discoveries, complexions, prussia, victories, gradual, accumulated, progeny, deformed, inherit, niece, younger, witnesses, tragedies, pupil, confound, promiscuous, plan, scarcity, commenced, importation, compare, plaintive, a, frequented, complains, resume, expert, proportionable, lowest, denies, pardons, sure, remember, solomon, pensive, companion, insisted, neighbourhood, peruse, reviving, gaudy, highness, briefly, agriculture, solitary, execution, begged, murderer, sluggish, terrestrial, furnish, renewed, whate'er, generals, occasioned, afraid, thanks, counsellor, likewise, confining, treaty, frail, give, callous, strive, variously, pensions, stratagems, article, living, conjure, suspension, relax, patrimony, grave, watchful, moreover, fair, confine, paintings, diminishing, expulsion, confer, derive, attend, edmund, quotation, recollecting, tempting, health, blackest, trained, farewell, allusion, varied, procuring, decrees, fewer, undergo, speaks, safer, reposed, shortness, owes, jews, artist, distinctly, redemption, pitying, resistless, differing, snares, revenged, blunders, commissions, victory, noted, regal, showing, observes, horror, festival, academy, voice, penetrated, keener, oracles, gothic, denote, dangerously, singly, helen, blockhead, talking, commands, verdict, spark, lady, supposed, disasters, farce, usually, intire, intimating, suspects, attends, guardian, surmounted, exerts, horrible, creed, strains, desolate, subsistence, congratulate, received, terror, punctually, lamentation, sketch, victorious, defending, awaits, rejoice, holy, omit, sports, admittance, rochester, attacks, devised, daughter, complexion, relate, intent, athens, sin, curiosities, things, commonly, multitudes, guest, recorded, remarks, vivid, managing, province, involves, intervention, commons, emblem, distemper, chosen, weighty, smiled, theme, devise, been, observ'd, attentively, listen, torpid, sanguine, paltry, districts, felt, unmarried, arises, subscribed, variation, proclaims, languish, reading, differ, fatigue, captivity, devotions, nobles, emperor, chusing, confederates, monuments, poet, bourbon, companies, cromwell, vassal, assisted, longer, pursued, extension, venture, incessant, perceive, fancied, estate, treason, exceeding, ventured, survive, vengeance, trouble, beforehand, subscriptions, voluminous, agonies, self, enlarge, additions, acquires, archbishop, grievously, apostles, shocks, companions, aimed, conspire, deserts, inform'd, enemy, re, subside, votes, secrets, hear, jests, sphere, withdraw, avenge, deferred, wanted, ravages, some, entire, swearing, something, daily, ceased, fitter, pursuing, raged, fatigues, afforded, retreat, feared, veteran, chearfully, interpret, necessaries, air, tradesmen, methinks, undone, dispose, students, supply, governor, creation, use, measure, progressive, savages, lucy, memorial, adds, accord, selected, oracle, colonies, abortive, frenchman, because, hazard, retinue, beg, resemble, footing, frowns, invested, dialogue, disputing, images, dismay, spontaneously, occur, repine, chide, review, potent, trading, father, test, understands, convenient, quitting, advanced, abounding, serve, poets, costs, popish, pension, persians, hers, execute, perishing, flame, opponent, varieties, make, exhausted, debates, adverse, president, perpetually, pray'r, mad, course, workmanship, celerity, creditors, conquering, proportionably, surprize, owns, sorry, joyful, aims, settlements, surprizing, its, replies, spleen, school, related, sect, safely, criminals, raving, involve, warms, displays, hourly, pre, devoutly, examine, obstruct, sanctuary, referred, handsomely, dearer, memorials, prayer, plentiful, stupendous, answers, receive, many, weal, nervous, solid, desolation, besides, virgin, threatened, rhyme, leader, grammar, martyr, offering, jewish, dependent, gift, far, feebly, mischiefs, resort, motions, specimen, dare, conducting, youths, rejoiced, forbidden, blooming, monster, senses, marries, decide, strong, delivering, added, household, britain, subsisting, increase, interpreter, fate, immoveable, unforeseen, maintenance, smiles, deal, spight, deepest, restoring, virgil, greece, solution, terrible, pulpit, horrors, learnt, alluded, contrivances, succeeded, mr, cognizance, cheap, urges, canst, stain, tho, brighter, caesar, tax, fop, unfinished, europeans, species, manufacturers, rejoicing, paths, surrender, copied, loves, gentlewoman, purchasing, discerned, deciding, change, dares, torment, perusing, deliverance, odd, adieu, forfeit, warrior, restitution, command, alter'd, barbarously, girl, preached, hurried, reflected, denomination, reigns, prodigiously, marlborough, hearted, consort, aunt, immersed, transferred, history, pastime, being, acute, mimic, divinity, debts, loss, exactly, given, countries, reflects, advance, forsaken, digested, assist, lofty, height, woe, choose, accordingly, tribunal, madame, dresses, service, mischief, physician, movement, prevented, prevents, brief, inheritance, medical, recent, heav'n, annals, procured, colleges, heated, stamp, presides, british, thy, prospect, inquire, ferdinand, charge, musician, theatres, both, bordering, inspect, vote, remedy, trappings, inhabitant, firmly, stating, joyfully, un, glories, forms, defended, reducing, memoirs, sins, trusts, settlement, heav'nly, booksellers, phantom, nameless, inhabitants, adjust, innumerable, gracefully, secures, blissful, eastern, defy, alexander, contending, lives, direction, commencement, sake, essays, directed, combine, abruptly, manufactures, vehicle, struggled, seeks, treasure, shakespeare, enquire, preparation, cement, dryden, expel, finances, heats, needy, exclaimed, perdition, agony, subscribers, ne, smallest, council, crowds, tortured, reduced, kingdom, occurred, unwholesome, rabble, forsake, high, farewel, valued, queen, according, thus, detaining, disposing, law, vicinity, circulated, transmit, brethren, withal, beholding, champion, knaves, provokes, presents, apparel, pangs, gloom, salaries, apply, drunken, distempers, traitors, tribe, strives, empires, peaceably, jury, commence, amends, reminds, count, deputies, beau, tempts, murder, forgets, overcame, sermon, lawyers, dilemma, metaphor, finished, wonted, appears, flavour, sir, lengths, corresponding, amazement, inundation, perish, diffusing, compleatly, departure, villain, guides, proudly, renewing, terrific, bonds, bribe, alas, cadence, fairly, enrich, bishop, atmosphere, views, complete, rich, feels, attending, consultation, prelude, series, unbroken, invited, parties, purify, quixote, devastation, proclaimed, mandate, published, pow'rs, redoubled, succeeding, rescue, bishops, included, deathless, ways, priests, aim, specified, defensive, insomuch, order, thank, purchased, consecrate, proclaim, seek, exhaust, son, did, unborn, accompanying, ordained, receiving, salutation, stratagem, combat, divert, suffer'd, committee, trials, labour, vow, sweet, join'd, inquiring, perceives, underwent, defenceless, fatigued, one, furnishing, loathsome, polish, somerset, captive, converted, richest, agonizing, filthy, hume, accidents, ingredients, fain, happens, europe, enormous, prefixed, reside, belongs, occupied, follow, joyless, lucky, materials, manage, performed, inquisition, during, sportive, delineated, insists, importing, baneful, griefs, quarrelling, add, altho, convent, report, mood, beauteous, verse, structure, ensue, sweets, luscious, description, shared, preparations, richer, distract, mathematical, project, lamenting, interposed, duchess, am'rous, commanders, paradise, operations, latest, firm, design'd, indifferently, confin'd, various, form, cardinal""".split(', ')

#fields['Concrete']="""gild, ranging, darker, pall, ralph, instrument, grove, bower, ipse, wolves, neatly, points, until, cures, wake, bowing, hall, tomorrow, sturdy, singers, shifting, flowed, lump, thirsty, mentions, fit, breathing, boy, sting, tour, gilds, port, frames, reach, west, happen'd, precipice, collects, verge, fitted, tools, kneeling, derives, trembled, landed, jail, lusty, damage, houshold, norfolk, filth, footman, fails, fight, wink, ruins, bristol, rays, hiding, wan, tom, signifies, voyage, ordered, clearing, passage, league, plunge, mineral, balmy, carries, presenting, projected, lend, virginia, welch, infants, lurking, print, steer, th, couple, pass'd, rev, home, rags, build, over, missed, indian, lucid, herds, swell, within, thief, fill'd, stray, kills, pasture, dancers, martin, pastures, north, says, joining, perfume, array, slumber, artillery, gunpowder, sly, foliage, reckoning, suspended, woven, game, glossy, canvass, groans, crush, banners, spheres, naples, omnes, holds, stirring, january, dose, rust, week, intercepted, weed, sailor, annum, except, radiant, slumbers, wasted, afternoon, ghosts, magazines, wield, shifts, surrounding, chasm, bell, sinking, terminated, ends, shrieks, onward, ballads, frighted, work, petticoats, tottering, list'ning, job, terminates, mass, ports, bringing, calais, erect, carriages, salts, rooms, mars, killing, sad, thither, cathedral, plague, repair, easter, dawn, flowing, cornwall, villages, markets, travel, presses, thieves, on't, arrives, weeps, distances, tartar, eye, danish, dark, bearer, lib, flaming, tomb, struck, bleeds, paper, touch, aut, hire, consisting, hundreds, dined, d, harvest, builds, dinner, sub, statues, winter, index, keep, setting, robe, hoary, lays, hard, pursuers, unfolds, aerial, gray, windings, sate, sed, horizon, makers, staid, crown, plot, outside, slain, dishes, betwixt, laurel, surround, mist, stay, cards, dispersed, sending, monkey, rainy, morrow, fixt, 18th, heap, stirs, kneel, pace, buried, healed, rob, kept, ear, m, expanse, anne, keeps, fled, fixed, flow'rs, swarm, chain, bills, rend, nt, prepar'd, sussex, christmas, ranged, swallowed, lad, happening, fragrant, drooping, tuesday, threat'ning, nights, terribly, smooth, soldiers, die, sown, heal, scent, ghost, dig, trumpets, disappeared, bait, morsel, turn, brow, richmond, currents, end, chronicle, et, traversed, smell, tides, crew, swine, gazing, frozen, bailiff, built, captains, hungry, circles, vis, sealed, draught, seventh, rent, intercept, stop, fishermen, morning, spectres, dove, entrance, shrink, july, lame, raw, lash, parted, brings, load, twill, wearing, bartholomew, ut, track, boards, daggers, foremost, nec, mihi, moves, east, letting, devouring, expands, vein, gulph, hoarse, took, 25th, german, races, cooper, divided, till, waits, card, circle, midnight, sugar, tow'ring, ey'd, chace, thursday, biting, shake, triumphal, yearly, rushed, dining, snows, myrtle, hercules, waiter, mexico, journies, ben, sound, overtake, pants, loaded, bath, pointed, dressing, chase, masses, gently, seat, ripe, stript, wait, main, coachman, doubled, aching, crops, grow, lands, starved, egypt, sprung, dust, fairies, fox, cum, bring, wrapt, dance, singing, morn, jam, quod, stars, bloody, deals, issued, 17th, left, nicholas, wastes, n, taking, wasting, removes, wear, dutch, altar, plac'd, earth, mariners, rosy, vale, juice, eleventh, 30th, battery, st, lily, liquor, nod, note, ocean, devoured, starts, prints, tables, branch, rested, closet, beneath, gleam, dates, big, slowly, tapestry, ghastly, cavalry, rains, passing, dice, h, dirty, burst, uppermost, choir, labourers, third, guineas, died, fainting, darted, gem, tibi, nods, disabled, building, gales, next, folio, chariots, through, howling, ending, rubbish, breathless, rid, spectacles, parnassus, presently, piece, plunging, crowns, spring, vines, brood, played, worked, stopped, fed, larger, bound, painted, hic, silken, brandy, vallies, together, sell, expanded, brink, dick, farms, dead, friday, burned, travels, whence, wounded, june, coach, quo, month, northward, span, asleep, wanders, surrounds, mornings, coffers, beset, sight, dy'd, trance, sings, peru, nigh, searching, spot, metal, alias, fees, watching, drowning, sink, twentieth, lime, giant, straining, situated, globe, year, kissing, pounds, ray, expired, mouths, takes, k, growing, frost, ascend, fills, carpenter, glitt'ring, buy, fortnight, towers, handed, hen, recruit, fun, boiling, bays, cloud, potatoes, snatched, spotted, ha, mines, wears, diamonds, pushing, shilling, merrily, bears, saxon, si, entry, night, trumpet, winds, giants, flocks, porter, stream, hunting, vest, enters, victuals, pierce, coasts, aye, forests, engraved, marshy, thirds, exit, sleep, sunt, cloaths, coins, brain, started, borders, vermin, shrubs, farm, 14th, nightly, 16th, measuring, lodge, sun, bursts, inserted, bench, scaffold, suns, caught, p, leagues, helm, pines, moor, stepped, 12th, clay, engines, ix, bore, unnumber'd, shorter, din, boys, shrill, gang, resound, wond'ring, whereon, stroke, punch, rests, saw, torch, wreck, wreaths, backward, legions, thunders, graves, september, around, below, fallen, nests, kiss, push, rider, parlour, keeper, doves, icy, grows, ounce, planted, rise, shady, nerve, prominent, regiments, ensign, sixth, beards, indies, hue, wrinkled, feeding, kill, shower, thence, carried, drank, weapon, pistol, start, chaise, shriek, rivulets, packet, space, issues, moveable, 15th, stops, ca, chapel, bridle, pox, lower, snatch, slender, steal, stab, snuff, band, drowned, hyde, lone, cavern, surveying, list, flow'ry, room, threescore, ago, posts, pointing, leaf, camp, fifth, plume, babes, inclosed, wolf, naked, cradle, cloak, club, pierced, front, tread, april, face, lie, torrents, nunc, bleed, drinks, rising, burn, surgeon, pudding, tire, rises, isles, b, wreath, steed, ach, gout, come, subterraneous, stealing, farthest, partition, yon, double, shores, dies, standing, danced, robin, grain, owl, tobacco, eighth, canopy, strikes, bind, devour, serpent, column, islands, mediterranean, ascended, thunder, wore, lamb, sleeps, lance, 20th, links, bar, turkey, china, stand, pan, softly, twilight, shops, jamaica, turns, shrinking, fields, mole, seize, gale, hungary, 2d, wide, scatter, perpendicular, diamond, torches, nerves, jack, dish, insects, seats, starry, cover, repairs, flesh, spreading, line, monday, bows, 24th, carry, hey, spy, divide, erat, corps, 9th, squares, breeze, ivy, strip, ascending, deposited, fleet, stooping, belonging, stake, spacious, plants, marched, fragment, dye, ore, hunt, posted, cell, gown, ass, 26th, cheek, shoals, up, nimble, ire, chariot, fly, mangled, smaller, reins, curtain, linger, cry, streams, dover, mess, mud, van, beams, car, sold, send, dirt, encircled, wounds, tear, groan, snug, bowels, flowers, herbs, tenth, tent, thickest, shillings, watch, marble, atque, trembling, playing, corpse, cups, drink, size, fourth, arch, seals, stands, watered, hold, dashed, rushing, working, extending, vel, 27th, bolts, brought, stomach, spires, sharp, swallow, cook, resembling, vales, lye, plant, garments, reaches, ann, arms, joan, tough, stables, adjacent, largest, burning, drew, olympus, garret, breaks, away, dividing, 8th, stable, grim, butcher, road, driven, link, lads, ii, seizing, nineteen, garden, noon, net, bathe, bent, shrunk, bursting, forest, sounding, getting, briskly, call'd, g, sunk, shook, coast, mead, cuts, wider, thro, tents, kill'd, grains, seas, again, gone, tide, carrying, ragged, trap, strike, clouds, lab'ring, wrap, co, market, plates, flower, poured, goose, separates, seventy, bull, placed, gilded, blazing, set, ale, dews, twelfth, panting, roof, groaning, seal, workmen, gold, verdant, cot, ll, shattered, smith, sleeping, bearing, seizes, reached, sends, shaded, meads, ascends, table, armed, pole, chambers, ruddy, quarter, cypress, newly, bee, ninety, clasp, lamps, area, flute, scattered, corn, look'd, lee, lark, grey, o'er, diameter, columns, border, margin, dagger, lawns, fifty, 4th, close, stocks, ink, meal, sat, toll, woody, babe, petticoat, apace, woods, closes, spread, lamp, rain, closing, descending, sailors, spies, lungs, ice, x, violet, pours, waved, projecting, 6th, go, garlands, bees, patches, shift, armour, squadrons, feed, bulls, honey, roads, salt, dcc, put, key, land, faces, seventeen, moon, valley, dart, casts, called, sixteen, watches, angle, lo, headed, waters, met, lawn, treads, chamber, ay, downward, hundred, descends, stair, wells, catching, fishing, buds, temples, smoke, flanders, fits, meadows, rivers, bare, bled, guinea, emerging, roaring, spreads, 29th, adjoining, surgeons, sounded, hovering, downy, sky, stepping, hunted, waving, inn, sword, marches, blows, womb, rot, swelling, gate, caves, engine, draws, afar, garter, arrows, star, pinions, nut, clothes, ninth, coloured, gets, lightning, see, sweat, rhine, hid, wave, traverse, sow, limbs, shut, rush, darts, urn, palm, swords, drives, sailing, brains, couch, troop, mouse, billows, bud, echoing, olive, lace, swiftly, herb, shield, brown, streaming, going, chicken, trod, wig, casting, lies, goat, thrown, thrice, walks, fold, cream, blown, fragments, apple, march, hearth, weave, stretch, poultry, fifteen, swan, tall, garland, pale, climb, crushed, rushes, hunter, colour, sack, horsemen, birds, heaps, twenty, loads, drive, bleeding, missing, cliffs, sixty, fall, lakes, glass, catches, eagle, square, angles, hat, hanged, stood, pine, covering, circumference, terminating, sprang, rusty, ounces, mare, mortar, crying, meat, drove, hides, beef, grasps, eighteen, crept, issuing, livid, get, j, geese, bowl, ashore, chairs, milk, galleries, sizes, chickens, wipe, clasping, lift, trembles, beer, dry, 10th, highway, watery, alighted, eighty, hound, spears, cap, breaking, killed, laid, pockets, seed, nile, mountain, mount, ox, dog, field, walls, wound, cheeks, horseback, stem, catch, pouring, draw, broader, half, twice, street, fore, circling, drum, swallows, flood, turf, canvas, golden, pool, rivulet, bird, kin, flew, battlements, kick, ere, boiled, steel, swept, roar, shutting, steals, wheat, hand, hunters, glide, pair, hog, eaten, rod, bend, ringing, carcase, cup, rose, board, eat, went, hounds, fix, throats, walk, lambs, mounts, cast, frogs, circuit, feather, watry, hands, upward, gallery, thickness, pebbles, park, near, cells, shaft, needle, glasses, treading, sha, oaks, turrets, dragged, den, cotton, threshold, yonder, grapes, dropped, opens, knot, handle, sees, 7th, washing, pant, penny, folded, tawny, vessel, cannon, windows, alley, stole, coals, joint, large, liver, lightnings, wrapped, seated, turned, garment, pint, pearls, oar, forty, dozen, mantle, scales, rocks, dash, fiddle, fowl, capt, thorns, throw, dashing, cow, gloves, gather, traversing, pond, concave, lay, heads, sheet, ware, eats, cake, balls, crossed, worm, fibres, jump, pound, landing, stopping, bow, gore, ship, bank, circular, glancing, q, canals, drop, flat, massy, dogs, broken, harp, reed, lifts, carpet, crossing, lifted, throws, sailed, driving, powder, parchment, thin, copper, rack, plump, clapping, hogs, off, drawing, purple, rattling, came, darting, pack, banks, miles, poles, rotten, walked, pipes, lake, triple, wing, fourteen, bits, sleeve, bends, gathered, bread, limb, ears, scarlet, muscles, firing, clusters, sleeves, closed, o'clock, cliff, coming, oil, stays, greenwich, tower, shop, mail, waggon, covered, torn, staring, shakes, wire, worn, sands, drag, pig, worms, shore, entrails, crystal, apples, pearl, small, sprinkled, fowls, spear, axe, brows, bars, crest, walking, vault, turn'd, snowy, beds, hats, falls, candle, asunder, gaping, barley, shaped, glides, looking, holding, steep, snake, stag, whistle, sand, pistols, break, ships, rats, sandy, washed, falling, dragon, locked, egg, dusky, rocky, col, goats, paces, piccadilly, chair, wand, quiver, craggy, fitting, mounted, surface, thirty, calf, cattle, blow, web, along, lap, cheese, beard, benches, leaves, wash, forwards, shelves, cave, channel, curtains, tea, kitchen, fro, boats, sable, guns, azure, heaves, whistling, rear, bed, gallows, upwards, nodding, horses, tearing, ladder, beating, gasping, tight, stopt, burnt, climbing, lock, sheep, pitched, about, lying, cock, gliding, knocking, handkerchief, helmet, row, covers, threw, rears, springing, pieces, sit, serpents, boxes, ax, riding, ditches, cords, brass, cats, skirts, behind, waistcoat, wool, five, beam, fling, knees, leaped, thirteen, linen, ash, dropt, hairs, tie, bit, vaults, fell, brick, pillars, tooth, stretches, candles, dung, whip, shoots, leaning, fast, lifting, wax, got, ball, roars, quill, facing, arrow, fishes, silk, hit, mouth, sits, snakes, flies, broke, gravel, bushes, keys, hill, taper, wind, plough, bigger, leaden, huge, meadow, flying, bolt, crosses, flag, silver, bottles, hedges, stairs, veins, peep, pocket, palms, thicker, opening, ducks, clock, hay, heave, edge, dried, oxen, necks, scatter'd, boots, oak, aloft, skins, boughs, sucking, twelve, straw, 11th, straight, reeds, caps, turning, quarters, strait, bough, canal, bush, nostrils, oval, pavement, beach, rode, velvet, folds, boat, mutton, lip, middle, bridges, cows, bruised, mall, plate, outer, apron, floating, foam, blade, bite, pipe, red, runs, feathers, strings, side, toss, crow, picked, river, sweep, cloths, ring, lane, nimbly, locks, tying, trees, leap, sea, corners, huts, shirt, pit, drags, cellar, cloth, downwards, cord, run, tree, bells, beads, rock, shoot, hedge, door, cane, down, knocked, tails, winding, slip, wet, foaming, leaping, hind, ran, gathering, pick, jacket, black, strand, timber, window, paternoster, ride, crack, coats, coffin, dug, thread, arm, eleven, stretched, four, cross, stone, bark, sail, fish, peeping, fetch, lids, seven, waist, coat, swim, ivory, loins, iron, three, eggs, knock, oars, sails, lean, back, beat, trunks, berries, reaching, corner, sliding, shells, drops, block, blowing, jaws, bending, stones, bags, deer, creep, water, bridge, mile, threads, bottom, gathers, white, sitting, knots, cart, thick, crooked, girdle, wheel, bay, blew, pits, inside, tied, deck, horse, thrust, upper, backwards, footed, shoes, bason, sticking, cutting, knee, tumble, picking, swimming, hair, forehead, shooting, staff, basket, inches, stool, roll, stockings, nine, leather, underneath, clap, pot, green, cavity, backs, ground, trip, ditch, saddle, tops, sides, brush, shaking, cat, wood, finger, pulled, blue, grass, rides, wall, sticks, tore, running, mill, shell, rolls, hang, pots, head, twisted, fist, toe, hangs, claws, bag, feet, pulling, hollow, throat, hung, shoulders, knife, horn, knives, stretching, box, heels, stick, chest, stuck, tumbling, yards, yellow, shot, ridge, round, stalk, skin, pull, yard, elbow, skull, pulls, rolling, rounds, broad, horns, nose, joints, edges, hook, rolled, teeth, rub, rings, bones, belly, foot, heel, floor, top, chin, nails, collar, ribs, hole, fastened, tail, tip, wooden, leg, thumb, holes, ropes, bone, thigh, neck, pins, shoulder, pin, nail, cut, toes, legs""".split(', ')


#fields['Vices']="""folly, impiety, injustice, cruelty, avarice, ignorance, profligacy, luxury, hypocrisy, infidelity, treachery, irreligion, insolence, falsehood, ingratitude, villainy, vanity, indolence, imprudence, stupidity, pride, levity, debauchery, barbarity, artifice, arrogance, sensuality, negligence, libertinism, inattention, fraud, extravagance, prodigality, perfidy, obstinacy, intemperance, cowardice, timidity, immorality, effeminacy, dissimulation, venality, sloth, inhumanity, pusillanimity, insensibility, indiscretion, deceit, apostasy, temerity, petulance, knavery, insincerity, imposture, sophistry, flattery, affectation, unbelief, severity, prevarication, indecency, extortion, brutality, misconduct, impenitency, impenitence, gluttony, falsity, fallacy, disobedience, delusion, vainglory, subtlety, ostentation, irresolution, inexperience, indifference, indelicacy, inadvertency, impudence, futility, truth, neglect, irreverence, instability, infatuation, inconstancy, forgery, equivocation, deception, conceit, bribery, vice, subterfuge, servility, rebellion, presumption, parsimony, misrepresentation, mismanagement, inactivity, disregard, disloyalty, diffidence, rigour, mutability, incredulity, illusion, ferocity, extravagancy, evasion, effrontery, credulity, contrivance, vicissitude, untruth, superfluity, stratagem, singularity, roguery, revolt, profusion, omission, obscurity, intrigue, defection, craft, conspiracy, chicanery, assassination, ambiguity, adulation, verity, usury, unconcern, reality, prostitution, oversight, novelty, neg, misbehaviour, invention, insinuation, indifferency, fatality, concupiscence, collusion""".split(', ')

#fields['Virtues']="""generosity, prudence, piety, modesty, probity, moderation, integrity, candour, humility, benevolence, sincerity, courage, affection, magnanimity, humanity, affability, learning, sagacity, friendship, fortitude, clemency, penetration, liberality, honesty, fidelity, compassion, sobriety, ingenuity, gratitude, beneficence, impartiality, dexterity, decency, bravery, valour, temperance, purity, lenity, civility, charity, benignity, ardour, vigour, veracity, loyalty, intrepidity, frugality, constancy, circumspection, understanding, skill, condescension, zeal, vehemence, industry, heroism, fervency, erudition, complaisance, wisdom, vigilance, gallantry, fervour, chastity, bounty, munificence, mercy, love, knowledge, goodwill, diligence, assiduity, agility, warmth, strength, simplicity, sanctity, patriotism, innocence, discernment, cordiality, activity, punctuality, pity, patience, hospitality, foresight, forbearance, decorum, amity, allegiance, alacrity, ability, vehemency, sympathy, rectitude, labour, facility, caution, austerity, virtue, toil, proficiency, precaution, indulgence, fatigue, exertion, courtesy, continence, commiseration, attachment, adherence, wit, spirituality, resolution, policy, perseverance, ledge, infight, grace, exercise, discretion, comprehension, capacity, abstinence""".split(', ')



def datetime():
	import time
	timestr = time.strftime("%Y%m%d-%H%M%S")
	return timestr

def field_freqs(text):
	if not type(text)==list:text=tokenize2(text)
	word_freqs=toks2freq(text)

	odx={}
	totalwords=float(sum(word_freqs.values()))
	for f,words in fields.items():
		odx[f]=sum([word_freqs.get(word,0) for word in words]) / totalwords
	return odx



def get_word_window(text,numwords=100,go_backwards=False):
	import re
	spaces = [match.start() for match in re.finditer(re.compile('\s'), text)]
	spaces = list(reversed(spaces)) if go_backwards else spaces
	spaces = spaces[:numwords]
	return text[:spaces[-1]] if not go_backwards else text[spaces[-1]:]

def index(text,phrase,ignorecase=True):
	compiled = re.compile(phrase, re.IGNORECASE) if ignorecase else re.compile(phrase)
	passage_indices = [(match.start(), match.end()) for match in re.finditer(compiled, text)]
	return passage_indices

def passages(text,phrases=[],window=200,indices=None,ignorecase=True,marker='***'):
	txt_lower = text.lower()
	window_radius=int(window/2)
	for phrase in phrases:
		if phrase.lower() in txt_lower:
			if not indices: indices = index(text,phrase,ignorecase=ignorecase)

			for ia,ib in indices:
				pre,post=text[:ia],text[ib:]
				match = text[ia:ib]
				window=get_word_window(pre,window_radius,True) + marker+match+marker+get_word_window(post,window_radius,False)
				dx={'index':ia, 'index_end':ib, 'passage':window,'phrase':phrase}
				yield dx


def read_ld(fn,keymap={},toprint=True):
	if fn.endswith('.xls') or fn.endswith('.xlsx'):
		return xls2ld(fn,keymap=keymap)
	elif fn.endswith('.csv'):
		#return tsv2ld(fn,tsep=',',keymap=keymap)
		tsep=','
		import csv
		with codecs.open(fn,encoding='utf-8') as f:
			return list(csv.DictReader(f))
	else:
		tsep='\t'

	return list(readgen(fn,tsep=tsep,as_dict=True,toprint=toprint))
		#return tsv2ld(fn,keymap=keymap)



def writegen(fnfn,generator,header=None):
	of = codecs.open(fnfn,'w',encoding='utf-8')
	for i,dx in enumerate(generator()):
		if not header: header=sorted(dx.keys())
		if not i: of.write('\t'.join(header) + '\n')
		of.write('\t'.join([unicode(dx.get(h,'')) for h in header]) + '\n')


def writegengen(fnfn,generator,header=None,save=True):
	if save: of = codecs.open(fnfn,'w',encoding='utf-8')
	for dx in generator():
		if not header:
			header=sorted(dx.keys())
			if save: of.write('\t'.join(header) + '\n')
		if save: of.write('\t'.join([unicode(dx.get(h,'')) for h in header]) + '\n')
		yield dx

def readgen(fnfn,header=None,tsep='\t',keymap={},keymap_all=unicode,encoding='utf-8',as_list=False,as_tuples=False,as_dict=True,toprint=True):
	if tsep=='\t' and toprint:
		print '>> streaming as tsv:',fnfn
	elif tsep==',' and toprint:
		print '>> streaming as csv:',fnfn
	import time
	now=time.time()
	header=None
	if fnfn.endswith('.gz'):
		import gzip
		of=gzip.open(fnfn)
	#of = codecs.open(fnfn,encoding=encoding)
	else:
		of=open(fnfn)

	for line in of:
		line=line.decode(encoding=encoding)[:-1]
		line=line.replace('\r\n','').replace('\r','').replace('\n','')
		if not header:
			header=line.split(tsep)
			continue

		r=data=line.split(tsep)
		if as_list:
			yield r
			continue

		if as_tuples or as_dict:
			r=tuples=zip(header,data)
		if as_dict:
			r=d=dict(tuples)
		yield r
	of.close()
	nownow=time.time()
	if toprint: print '   done ['+str(round(nownow-now,1))+' seconds]'

def header(fnfn,tsep='\t',encoding='utf-8'):
	header=[]

	if fnfn.endswith('.gz'):
		import gzip
		of=gzip.open(fnfn)
	#of = codecs.open(fnfn,encoding=encoding)
	else:
		of=open(fnfn)

	for line in of:
		line = line[:-1]  # remove line end character
		line=line.decode(encoding=encoding)
		header=line.split(tsep)
		break
	of.close()
	return header

def read(fnfn,to_unicode=True):
	if fnfn.endswith('.gz'):
		import gzip
		try:
			with gzip.open(fnfn,'rb') as f:
				x=f.read()
				if to_unicode: x=x.decode('utf-8')
				return x
		except IOError as e:
			print "!! error:",e,
			print "!! opening:",fnfn
			print
			return ''

	elif fnfn.endswith('.txt'):
		if to_unicode:
			try:
				with codecs.open(fnfn,encoding='utf-8') as f:
					return f.read()
			except UnicodeDecodeError:
				return read(fnfn,to_unicode=False)
		else:
			with open(fnfn) as f:
				return f.read()

	return ''

def filesize(fn):
	return sizeof_fmt(os.path.getsize(fn))

def sizeof_fmt(num, suffix='B'):
    for unit in ['','Ki','Mi','Gi','Ti','Pi','Ei','Zi']:
        if abs(num) < 1024.0:
            return "%3.1f%s%s" % (num, unit, suffix)
        num /= 1024.0
    return "%.1f%s%s" % (num, 'Yi', suffix)










# SimpleCalc.py
#
# Demonstration of the parsing module,
# Sample usage
#
#     $ python SimpleCalc.py
#     Type in the string to be parse or 'quit' to exit the program
#     > g=67.89 + 7/5
#     69.29
#     > g
#     69.29
#     > h=(6*g+8.8)-g
#     355.25
#     > h + 1
#     356.25
#     > 87.89 + 7/5
#     89.29
#     > ans+10
#     99.29
#     > quit
#     Good bye!
#
#

def parse_math_str(input_string,variables={}):

	# Uncomment the line below for readline support on interactive terminal
	# import readline
	import re
	from pyparsing import Word, alphas, ParseException, Literal, CaselessLiteral, Combine, Optional, nums, Or, Forward, ZeroOrMore, StringEnd, alphanums
	import math

	# Debugging flag can be set to either "debug_flag=True" or "debug_flag=False"
	debug_flag=False

	exprStack = []
	varStack  = []

	def pushFirst( str, loc, toks ):
		exprStack.append( toks[0] )

	def assignVar( str, loc, toks ):
		varStack.append( toks[0] )

	# define grammar
	point = Literal('.')
	e = CaselessLiteral('E')
	plusorminus = Literal('+') | Literal('-')
	number = Word(nums)
	integer = Combine( Optional(plusorminus) + number )
	floatnumber = Combine( integer + Optional( point + Optional(number) ) + Optional( e + integer ) )

	ident = Word(alphas,alphanums + '_')

	plus  = Literal( "+" )
	minus = Literal( "-" )
	mult  = Literal( "*" )
	div   = Literal( "/" )
	lpar  = Literal( "(" ).suppress()
	rpar  = Literal( ")" ).suppress()
	addop  = plus | minus
	multop = mult | div
	expop = Literal( "^" )
	assign = Literal( "=" )

	expr = Forward()
	atom = ( ( e | floatnumber | integer | ident ).setParseAction(pushFirst) |
					 ( lpar + expr.suppress() + rpar )
				 )

	factor = Forward()
	factor << atom + ZeroOrMore( ( expop + factor ).setParseAction( pushFirst ) )

	term = factor + ZeroOrMore( ( multop + factor ).setParseAction( pushFirst ) )
	expr << term + ZeroOrMore( ( addop + term ).setParseAction( pushFirst ) )
	bnf = Optional((ident + assign).setParseAction(assignVar)) + expr

	pattern =  bnf + StringEnd()

	# map operator symbols to corresponding arithmetic operations
	opn = { "+" : ( lambda a,b: a + b ),
					"-" : ( lambda a,b: a - b ),
					"*" : ( lambda a,b: a * b ),
					"/" : ( lambda a,b: a / b ),
					"^" : ( lambda a,b: a ** b ) }

	# Recursive function that evaluates the stack
	def evaluateStack( s ):
		op = s.pop()
		if op in "+-*/^":
			op2 = evaluateStack( s )
			op1 = evaluateStack( s )
			return opn[op]( op1, op2 )
		elif op == "PI":
			return math.pi
		elif op == "E":
			return math.e
		elif re.search('^[a-zA-Z][a-zA-Z0-9_]*$',op):
			if variables.has_key(op):
				return variables[op]
			else:
				return 0
		elif re.search('^[-+]?[0-9]+$',op):
			return long( op )
		else:
			return float( op )

	# Start with a blank exprStack and a blank varStack
	exprStack = []
	varStack  = []

	if input_string != '':
		# try parsing the input string
		try:
			L=pattern.parseString( input_string )
		except ParseException,err:
			L=['Parse Failure',input_string]

		# show result of parsing the input string
		if debug_flag: print input_string, "->", L
		if len(L)==0 or L[0] != 'Parse Failure':
			if debug_flag: print "exprStack=", exprStack

			# calculate result , store a copy in ans , display the result to user
			result=evaluateStack(exprStack)
			variables['ans']=result
			#print result
			return result

			# Assign result to a variable if required
			if debug_flag: print "var=",varStack
			if len(varStack)==1:
				variables[varStack.pop()]=result
			if debug_flag: print "variables=",variables
		else:
			print 'Parse Failure'
			print err.line
			print " "*(err.column-1) + "^"
			print err
